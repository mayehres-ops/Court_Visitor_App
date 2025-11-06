# guardian_extractor.py
# Clean full script (drop-in). Last updated: 2025-09-05 – address capture fix

import os, io, re, shutil, time
import subprocess
import sys

# CRITICAL FIX: Patch subprocess.Popen GLOBALLY to hide ALL windows (Tesseract/Poppler)
if sys.platform == 'win32':
    _original_popen = subprocess.Popen

    class _HiddenWindowPopen(_original_popen):
        """Popen wrapper that ALWAYS hides console windows on Windows"""
        def __init__(self, *args, **kwargs):
            # Add STARTUPINFO if not present
            if 'startupinfo' not in kwargs:
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                startupinfo.wShowWindow = 0  # SW_HIDE
                kwargs['startupinfo'] = startupinfo
            # Add CREATE_NO_WINDOW if not present
            if 'creationflags' not in kwargs:
                kwargs['creationflags'] = subprocess.CREATE_NO_WINDOW
            # Call original Popen
            super().__init__(*args, **kwargs)

    # Replace globally - affects pdf2image and pytesseract!
    subprocess.Popen = _HiddenWindowPopen

import openpyxl
import pdfplumber
from datetime import datetime
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from pdf2image import convert_from_bytes
import pytesseract
from PIL import Image, ImageFilter, ImageOps

class _SafeWriter:
    def __init__(self, wrapped):
        self._wrapped = wrapped
        self._enc = getattr(wrapped, "encoding", None) or "cp1252"

    def write(self, s):
        try:
            self._wrapped.write(s)
        except UnicodeEncodeError:
            self._wrapped.write(
                s.encode(self._enc, errors="replace").decode(self._enc)
            )

    def flush(self):
        try:
            self._wrapped.flush()
        except Exception:
            pass

# Install wrappers so all print(...) calls are safe
sys.stdout = _SafeWriter(sys.stdout)
sys.stderr = _SafeWriter(sys.stderr)
# --- END: safe console I/O wrapper for Windows ---


# ========= EDIT THESE (your values) =========
# Local input folder for PDFs (no longer using Google Drive folder ID)
from pathlib import Path
def _env(key, default):
    return os.getenv(key, default)

# Try to use dynamic path detection, otherwise fall back to environment variables or hardcoded
try:
    # Add Scripts directory to path for imports
    _script_dir = Path(__file__).parent
    sys.path.insert(0, str(_script_dir / "Scripts"))
    from app_paths import get_app_paths
    _app_paths = get_app_paths()
    BASE = _app_paths.APP_ROOT
    LOCAL_PDF_INPUT_DIR = str(_app_paths.NEW_FILES_DIR)
    LOCAL_EXCEL_PATH = str(_app_paths.EXCEL_PATH)
    VISION_CREDENTIALS_FILE = str(_app_paths.CREDENTIALS_FILE) if _app_paths.CREDENTIALS_FILE.exists() else _env("GOOGLE_SERVICE_ACCOUNT_FILE", r"C:\configlocal\API\google_service_account.json")
except:
    # Fall back to environment variables or hardcoded paths
    BASE = Path(_env("GUARDIAN_BASE", r"C:\GoogleSync\GuardianShip_App"))
    LOCAL_PDF_INPUT_DIR = _env("PDF_INPUT_DIR", str(BASE / "New Files"))
    LOCAL_EXCEL_PATH = _env("EXCEL_PATH", str(BASE / "App Data" / "ward_guardian_info.xlsx"))
    VISION_CREDENTIALS_FILE = _env("GOOGLE_SERVICE_ACCOUNT_FILE", r"C:\configlocal\API\google_service_account.json")

GOOGLE_SERVICE_ACCOUNT_FILE = VISION_CREDENTIALS_FILE

# Backward-compat alias (so any old code that still refers to the old name won't break)

import re
from datetime import datetime, date

# Date headers we will normalize on write
DATE_HEADERS = {"wdob", "gdob", "g2dob", "DateARPfiled", "Dateappointed"}

# Flexible MDY: 6/9/95, 06-09-1995, 6.9.25, etc.
_MDY_RE = re.compile(r"^\s*(\d{1,2})[\/\.\-](\d{1,2})[\/\.\-](\d{2,4})\s*$")

def _norm_mdy_str(s: str) -> str:
    """Return MM/DD/YYYY (zero-padded) or '' if invalid."""
    if not s:
        return ""
    if isinstance(s, (datetime, date)):
        return s.strftime("%m/%d/%Y")
    m = _MDY_RE.match(str(s))
    if not m:
        return ""
    mm, dd, yy = m.groups()
    mm, dd = int(mm), int(dd)
    yy = int(yy)
    # Expand 2-digit years deterministically
    yy = 2000 + yy if yy < 50 else (1900 + yy if yy < 100 else yy)
    # Sanity bounds
    if not (1 <= mm <= 12 and 1 <= dd <= 31 and 1900 <= yy <= 2100):
        return ""
    return f"{mm:02d}/{dd:02d}/{yy}"

def _as_excel_date_or_text(s: str):
    """
    If parsable, return (True, date_obj, 'mm/dd/yyyy').
    Else return (False, normalized_text_or_empty, None).
    """
    norm = _norm_mdy_str(s)
    if not norm:
        return (False, "", None)
    try:
        d = datetime.strptime(norm, "%m/%d/%Y").date()
        return (True, d, "mm/dd/yyyy")
    except Exception:
        return (False, norm, None)


def extract_text_with_ocr_for_arp(
    pdf_path: str,
    page_index: int = 0,
    *,
    psm: int | None = None,           # legacy callers pass psm=4
    psm_primary: int | None = None,   # new style
    psm_alt: int = 6,
    **kwargs
) -> str:
    """
    Robust ARP OCR shim with quiet logging:
      - Mode "path":   convert_from_path only (your previous stable behavior)
      - Mode "bytes":  convert_from_bytes only
      - Mode "mixed":  try bytes, then path
      - Suppresses noisy pdf2image error text if QUIET_PDF2IMAGE_ERRORS = True
      - Runs Tesseract twice and keeps the longer result
    """
    if psm_primary is None:
        psm_primary = psm if psm is not None else 4

    def _quiet_log(msg: str, exc: Exception | None = None, level: str = "NOTE"):
        if QUIET_PDF2IMAGE_ERRORS or exc is None:
            log(f"{level}: {msg}")
        else:
            log(f"{level}: {msg}: {exc}")

    def _from_bytes(_pdf_path: str, _page_idx: int):
        try:
            from pdf2image import convert_from_bytes
        except Exception as e:
            _quiet_log("pdf2image import failed (bytes)", e, "ERROR")
            return None
        try:
            with open(_pdf_path, "rb") as f:
                _bytes = f.read()
            imgs = convert_from_bytes(
                _bytes, dpi=300, poppler_path=POPPLER_BIN,
                first_page=_page_idx + 1, last_page=_page_idx + 1
            )
            return imgs[0] if imgs else None
        except Exception as e:
            _quiet_log("convert_from_bytes failed", e, "NOTE")
            return None

    def _from_path(_pdf_path: str, _page_idx: int):
        try:
            from pdf2image import convert_from_path
        except Exception as e:
            _quiet_log("pdf2image import failed (path)", e, "ERROR")
            return None
        try:
            imgs = convert_from_path(
                _pdf_path, dpi=300, poppler_path=POPPLER_BIN,
                first_page=_page_idx + 1, last_page=_page_idx + 1
            )
            return imgs[0] if imgs else None
        except Exception as e:
            # This is where the library sometimes throws the “local variable 'err'…” message.
            _quiet_log("convert_from_path failed", e, "NOTE")
            return None

    # Choose loader according to mode
    img = None
    if ARP_OCR_MODE == "path":
        img = _from_path(pdf_path, page_index)
    elif ARP_OCR_MODE == "bytes":
        img = _from_bytes(pdf_path, page_index)
    else:  # "mixed"
        img = _from_bytes(pdf_path, page_index) or _from_path(pdf_path, page_index)

    if img is None:
        return ""  # Let Vision or other fallbacks continue

    try:
        proc = preprocess_for_ocr(img)
    except Exception:
        proc = img

    def _ocr(psm_val: int) -> str:
        cfg = f"--psm {psm_val}"
        try:
            raw = pytesseract.image_to_string(proc, config=cfg)
        except Exception:
            raw = ""
        return normalize_unicode_noise(clean_text(raw or ""))

    txt1 = _ocr(psm_primary)
    txt2 = _ocr(psm_alt)
    best_txt, used = (txt1, psm_primary) if len(txt1) >= len(txt2) else (txt2, psm_alt)
    print(f"  ARP OCR used psm {used} (chars={len(best_txt)})")
    return best_txt




def list_local_pdfs(folder: str):
    """
    Return a list of dicts like [{'name': 'file.pdf', 'path': 'C:\\...\\file.pdf'}]
    for all PDFs in the given folder (non-recursive).
    """
    out = []
    log(f"DEBUG: Checking for PDFs in folder: {folder}")
    log(f"DEBUG: Folder exists: {os.path.exists(folder)}")
    log(f"DEBUG: Folder is directory: {os.path.isdir(folder)}")
    
    try:
        files_in_folder = os.listdir(folder)
        log(f"DEBUG: Found {len(files_in_folder)} total files in folder")
        
        pdf_count = 0
        for nm in files_in_folder:
            if nm.lower().endswith(".pdf"):
                pdf_count += 1
                out.append({"name": nm, "path": os.path.join(folder, nm)})
                log(f"DEBUG: Found PDF: {nm}")
        
        log(f"DEBUG: Found {pdf_count} PDF files out of {len(files_in_folder)} total files")
        
    except Exception as e:
        log(f"ERROR listing local PDFs in {folder}: {e}")
        log(f"DEBUG: Exception type: {type(e).__name__}")
        log(f"DEBUG: Exception details: {str(e)}")
    
    return out

def read_pdf_bytes(file_path: str) -> bytes:
    with open(file_path, "rb") as f:
        return f.read()


# Backups + logs
BACKUP_DIR = r"C:\GuardianAutomation\Backup"
LOG_DIR    = r"C:\GuardianAutomation\Log_Files"

# Google Vision service-account key (JSON) moved here:


# Tesseract + Poppler (unchanged)
pytesseract.pytesseract.tesseract_cmd = _env("TESSERACT_CMD", r"C:\Program Files\Tesseract-OCR\tesseract.exe")
POPPLER_BIN = _env("POPPLER_BIN", r"C:\Poppler\Release-25.07.0-0\poppler-25.07.0\Library\bin")
# ===========================================

os.makedirs(LOG_DIR, exist_ok=True)
LOG_FILE = os.path.join(LOG_DIR, "extraction_log.txt")

DEBUG_TEXT_DIR = os.path.join(LOG_DIR, "debug_texts")
os.makedirs(DEBUG_TEXT_DIR, exist_ok=True)

HEADERS = [
    "wardlast","wardfirst","wardmiddle","causeno","visitdate","visittime","wtele","liveswith","waddress","wdob",
    "guardian1","gaddress","gemail","gtele","Relationship","gdob","Guardian2","g2 address","g2eamil",
    "g2tele","g2Relationship","g2dob","datesubmitted","Dateappointed","miles","expense submitted",
    "expensepd","DateARPfiled","Comments","CVR created?","emailsent","Appt_confirmed","Contact_added"
]

# ----- Globals -----
BACKUP_DONE = False  # one backup per run
# ----- Globals -----
BACKUP_DONE = False  # one backup per run

# OCR conversion toggles (keep these single-source-of-truth)
ARP_OCR_MODE = "mixed"          # "path" / "bytes" / "mixed"
QUIET_PDF2IMAGE_ERRORS = True   # True = suppress pdf2image exception text

LABEL_WORDS = {
    # generic
    "phone","address","addresses","relationship","dob","date","time","guardian","guardian(s)",
    "city/state/zip","name","names","email","emails","g1","g2","guardian 1","guardian 2",
    # specific labels that showed up in logs
    "residence","relationship to ward","ward address","ward phone","ward dob",
    "guardian address","guardian phone","guardian dob","guardian email",
    # new anti-noise from checkboxes
    "is (check only one)","check only one","is check only one",
    "ward's home","foster home","group home","nursing home","relative's home","other"
}

PHONE_RE = re.compile(r"\b\(?\d{3}\)?[ \-\.\/]?\d{3}[ \-\.\/]?\d{4}\b")
DATE_RE  = re.compile(r"\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b")
EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")

# --- Exact ARP label regexes (from mapping) ---
WARD_ADDR_LABEL  = re.compile(r"^\s*(?:WARD\s*:)?\s*Address\s*\(no\s*P\.?\s*O\.?\s*Box\)\s*:?\s*", re.I)
WARD_CITY_LABEL  = re.compile(r"^\s*(?:WARD\s*:)?\s*(?:City/State/Zip|CityStateZip|City\s+State\s+Zip)\s*:?\s*", re.I)

GUARD_ADDR_LABEL = re.compile(r"^\s*(?:GUARDIAN\(s\)\s*:)?\s*Address\s*\(no\s*P\.?\s*O\.?\s*Box\)\s*:?\s*", re.I)
GUARD_CITY_LABEL = re.compile(r"^\s*(?:GUARDIAN\(s\)\s*:)?\s*(?:City/State/Zip|CityStateZip|City\s+State\s+Zip)\s*:?\s*", re.I)

# Some forms omit punctuation/spacing; allow a slightly looser fallback line match.
def _line_starts_with(label_re: re.Pattern, line: str) -> bool:
    return bool(label_re.search(line or ""))

def _strip_label(label_re: re.Pattern, line: str) -> str:
    return label_re.sub("", line or "").strip()

def capture_arp_address_by_labels(text: str,
                                  addr_label_re: re.Pattern,
                                  city_label_re: re.Pattern,
                                  max_lines: int = 4) -> str:
    """
    Capture an address from an ARP section by looking for both the "Address" label
    and the "City/State/Zip" label. Returns a single cleaned address string.
    """
    if not text:
        return ""

    lines = (text or "").splitlines()
    raw_street = ""
    raw_city   = ""

    # --- STREET LINE ---
    for i, ln in enumerate(lines):
        if addr_label_re.search(ln):      # use compiled regex directly, no flags
            street = addr_label_re.sub("", ln).strip()
            if not street:
                # Look ahead a few lines for the actual value
                for j in range(i+1, min(i+1+max_lines, len(lines))):
                    nxt = lines[j].strip()
                    if not nxt:
                        continue
                    if re.match(r'^(City/State/Zip|CityStateZip|City\s+State\s+Zip|Phone|Email|DOB|Date of Birth|Relationship)\b',
                                nxt, re.I):
                        continue
                    street = nxt
                    break
            raw_street = street
            break

    # --- CITY/STATE/ZIP LINE ---
    for i, ln in enumerate(lines):
        if city_label_re.search(ln):      # use compiled regex directly, no flags
            city = city_label_re.sub("", ln).strip()
            if not city:
                for j in range(i+1, min(i+1+max_lines, len(lines))):
                    nxt = lines[j].strip()
                    if not nxt:
                        continue
                    # Skip obvious headers / next sections
                    if re.match(r'^(?:\d+\.\s*)?GUARDIAN\(s\)\b', nxt, re.I):
                        continue
                    if re.match(r'^(?:Name\(s\)|Phone|Email|E-?mail|DOB|Date of Birth|Relationship)\b', nxt, re.I):
                        continue
                    # Prefer lines that look like City, ST ZIP (or City, State ZIP)
                    if CITY_STATE_ZIP_RE.search(nxt) or re.search(r",[ ]*[A-Za-z]{2,}[ ]+\d{5}(?:-\d{4})?$", nxt):
                        city = nxt
                        break
                    # else accept as fallback (non-label, non-header)
                    city = nxt
                    break
            raw_city = city
            break

    # --- COMBINE ---
    if raw_street and raw_city:
        return clean_address(f"{raw_street}, {raw_city}")
    return clean_address(raw_street or raw_city)



def capture_labeled_value(text: str, label_re: re.Pattern) -> str:
    """
    Returns the single line right after the label (same line remainder if present,
    else the immediate next line), trimmed.
    """
    if not text: return ""
    lines = (text or "").splitlines()
    for i, ln in enumerate(lines):
        if _line_starts_with(label_re, ln):
            # Prefer same-line remainder after the label
            same_line = _strip_label(label_re, ln)
            if same_line:
                return same_line.strip()
            # Else take the next non-empty line that doesn't look like a label
            for j in range(i+1, min(i+3, len(lines))):
                nxt = (lines[j] or "").strip()
                if not nxt:
                    continue
                if any(_line_starts_with(p, nxt) for p in [WARD_ADDR_LABEL, WARD_CITY_LABEL, GUARD_ADDR_LABEL, GUARD_CITY_LABEL]):
                    break
                return nxt.strip()
    return ""

def join_address_lines(street: str, city_state_zip: str) -> str:
    """
    Clean up artifacts and join the two address parts into one line.
    """
    s1 = (street or "").strip()
    s2 = (city_state_zip or "").strip()

    # Drop the literal “(no P.O. Box)” note if it leaked into OCR
    s1 = re.sub(r"\(?no\s*P\.?\s*O\.?\s*Box\)?\s*:?","", s1, flags=re.I).strip()

    # Remove stray leading 'Address' label residue
    s1 = re.sub(r"^\s*(Address|Residence)\s*:?\s*", "", s1, flags=re.I).strip()

    # Collapse whitespace
    s1 = re.sub(r"\s{2,}", " ", s1)
    s2 = re.sub(r"\s{2,}", " ", s2)

    if s1 and s2:
        return f"{s1}, {s2}"
    return s1 or s2

def split_two_addresses(val: str) -> tuple[str, str]:
    """
    Some ARPs cram both guardians on one line. Split common separators.
    """
    v = (val or "").strip()
    if not v:
        return ("","")
    # Try explicit separators first
    for sep in [" / ", " & ", " and ", ";", " AND "]:
        if sep in v:
            parts = [p.strip(" ,") for p in v.split(sep, 1)]
            return (parts[0], parts[1] if len(parts) > 1 else "")
    # Heuristic: if two addresses look concatenated, try a hard split at double spaces near a comma
    m = re.search(r",\s*[A-Z]{2}\s+\d{5}(?:-\d{4})?\s{2,}", v)
    if m:
        return (v[:m.end()].strip(" ,"), v[m.end():].strip(" ,"))
    return (v, "")

# Address regex helpers
# (Keep STREET_LINE_RE as-is; it's reliable for the first line.)
STREET_LINE_RE = re.compile(
    r"""
    ^\s*
    \d{1,5}                              # Street number
    \s+
    [A-Za-z0-9'.\- ]+                    # Street name
    \s+
    (?:St|Street|Ave|Avenue|Rd|Road|Dr|Drive|Ln|Lane|
       Blvd|Boulevard|Ct|Court|Pl|Place|Pkwy|Parkway|
       Trl|Trail|Ter|Terrace)\.?         # Street type
    (?:\s+[A-Za-z0-9#.\- ]+)?            # Optional trailing unit on same line
    \s*$
    """,
    re.IGNORECASE | re.VERBOSE
)

# Accepts BOTH 2-letter abbreviations (TX) and spelled-out states (Texas), e.g.:
# "Houston, TX 77002" or "Houston, Texas 77002"
CITY_STATE_ZIP_RE = re.compile(
    r"""
    ^\s*
    (?:City/State/Zip\s*:)?\s*           # Optional label prefix
    [A-Za-z][A-Za-z .'\-]+               # City (one or more words)
    \s*,\s*
    [A-Za-z]{2,}                         # State (2+ letters; allows full names)
    \s+
    \d{5}(?:-\d{4})?                     # ZIP (5 or ZIP+4)
    \s*$
    """,
    re.IGNORECASE | re.VERBOSE
)

# Full one-line address (street + optional city + state + zip).
# Allow "TX" or "Texas" for state.
ADDRESS_RE = re.compile(
    r"""
    \b
    \d{1,5}                              # Street number
    [\s,.-]+
    [A-Za-z0-9'.\- ]+                    # Street name
    \s+
    (?:St|Street|Ave|Avenue|Rd|Road|Dr|Drive|Ln|Lane|
       Blvd|Boulevard|Ct|Court|Pl|Place|Pkwy|Parkway|
       Trl|Trail|Ter|Terrace)\.?
    (?:\s+[A-Za-z0-9#.\- ]+)?            # Optional unit (Apt 2B, Ste 300)
    \s*,\s*
    [A-Za-z][A-Za-z .'\-]+               # City (at least one word)
    \s*,\s*
    [A-Za-z]{2,}                         # State (2+ letters; allows full names)
    \s+
    \d{5}(?:-\d{4})?                     # ZIP
    \b
    """,
    re.IGNORECASE | re.VERBOSE
)

# Lines we should stop at (not used as a hard stop inside the capture loop,
# but helpful if you choose to add early exits later).
ADDR_STOP_RE = re.compile(
    r'^(Phone|Email|E-mail|DOB|Date of Birth|Relationship|Guardian(?:\s*\(s\))?|'
    r'Guardian\s*2|G2|Visit\s*Date|Visit\s*Time|Cause\s*No\.?)\b',
    re.I
)

# Intervening unit line that might sit between street and city/state/zip
UNIT_LINE_RE = re.compile(r'\b(?:Apt|Apartment|Unit|Suite|Ste|#)\s*[A-Za-z0-9\-]+\b', re.I)


# ---------- Logging ----------
def log(msg: str):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"[{ts}] {msg}\n")
    print(msg)

def extract_causeno_loose(text: str) -> str:
    """Very loose cause number sniff, then normalize with your existing normalize_causeno().
    Prioritizes cause numbers that follow the C-1-PB pattern (document format, not file stamp)."""
    if not text:
        return ""
    
    t = re.sub(r'\s+', ' ', text)
    
    # First priority: Look for cause numbers that follow the C-1-PB pattern
    # This pattern should only appear in the actual document content, not file stamps
    m = re.search(r'(?i)C\s*-?\s*1\s*-?\s*PB\s*-?\s*(\d{2})\s*-?\s*(\d{6})\b', t)
    if m:
        return normalize_causeno(f"{m.group(1)}-{m.group(2)}")
    
    # Second priority: Look for cause numbers in the document body (skip first few lines which are file stamp)
    lines = text.split('\n')
    if len(lines) > 5:
        document_body = '\n'.join(lines[5:])  # Skip first 5 lines (file stamp area)
        t_body = re.sub(r'\s+', ' ', document_body)
        m = re.search(r'\b(\d{2})-?(\d{6})\b', t_body) or re.search(r'\b(\d{2})-?(\d{5})\b', t_body)
        if m:
            return normalize_causeno(f"{m.group(1)}-{m.group(2)}")
    
    # Fallback: Look in entire text for any cause number pattern
    m = re.search(r'\b(\d{2})-?(\d{6})\b', t) or re.search(r'\b(\d{2})-?(\d{5})\b', t)
    if not m:
        return ""
    return normalize_causeno(f"{m.group(1)}-{m.group(2)}")

def save_debug(name: str, content: str):
    try:
        if not content: return
        safe = re.sub(r'[^A-Za-z0-9._-]+', '_', name)[:120]
        path = os.path.join(DEBUG_TEXT_DIR, f"{safe}.txt")
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
    except Exception as e:
        log(f"(debug save failed for {name}: {e})")

# ---------- Drive ----------
def get_drive_service():
    creds = service_account.Credentials.from_service_account_file(
        GOOGLE_SERVICE_ACCOUNT_FILE,
        scopes=["https://www.googleapis.com/auth/drive.readonly"]
    )
    return build("drive", "v3", credentials=creds)

def list_pdfs(folder_id: str):
    service = get_drive_service()
    q = f"'{folder_id}' in parents and mimeType='application/pdf' and trashed=false"
    resp = service.files().list(
        q=q,
        fields="files(id, name, mimeType, modifiedTime)",
        pageSize=1000,
        includeItemsFromAllDrives=True,
        supportsAllDrives=True
    ).execute()
    return resp.get("files", [])

def download_pdf_bytes(file_id: str) -> bytes:
    service = get_drive_service()
    request = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()

# ---------- Excel helpers ----------
def backup_excel_once(path: str):
    global BACKUP_DONE
    if BACKUP_DONE:
        return
    base, ext = os.path.splitext(path)
    ts = time.strftime("%Y%m%d_%H%M%S")
    bak = f"{base}__backup_{ts}{ext}"
    try:
        shutil.copy2(path, bak)
        log(f"(One-time backup created: {bak})")
    except Exception as e:
        log(f"(Backup skipped: {e})")
    BACKUP_DONE = True

def ensure_headers(ws):
    for i, h in enumerate(HEADERS, start=1):
        if ws.cell(row=1, column=i).value != h:
            ws.cell(row=1, column=i, value=h)

def normalize_causeno(cause: str) -> str:
    """
    Normalize to the canonical 'NN-NNNNNN' tail (2 digits + '-' + 6 digits).
    Accepts values with or without 'C-1-PB-' prefix and with 5 or 6 trailing digits.
    """
    if not cause:
        return ""
    s = re.sub(r'\s+', '', str(cause), flags=re.I)
    # Prefer anchored 6-digit tails if present (with or without C-1-PB prefix)
    m = re.search(r'(?i)(?:\b|^)(?:C-?1-?PB-?)?(\d{2})-?(\d{6})(?:\b|$)', s)
    if not m:
        # Fallback: accept 5 or 6 digits
        m = re.search(r'(\d{2})-?(\d{5,6})', s)
    if m:
        part1 = m.group(1)
        part2 = m.group(2).zfill(6)
        return f"{part1}-{part2}"
    return s

def _save_excel_with_retry(wb, file_path, max_retries=3):
    """
    Save Excel workbook with retry logic for file lock issues.
    """
    for attempt in range(max_retries):
        try:
            wb.save(file_path)
            log(f"Excel file saved successfully: {file_path}")
            return True
        except PermissionError as e:
            if attempt < max_retries - 1:
                log(f"Excel file is locked (attempt {attempt + 1}/{max_retries}). Retrying in 2 seconds...")
                time.sleep(2)
            else:
                log(f"ERROR: Cannot save Excel file - it may be open in Excel. Please close Excel and try again.")
                log(f"File path: {file_path}")
                log(f"Error: {e}")
                return False
        except Exception as e:
            log(f"ERROR: Failed to save Excel file: {e}")
            return False
    return False

def upsert_row_to_excel(row: dict):
    """
    Update an existing row by 'causeno' or append a new row.
    Always overwrites: causeno, Dateappointed, last_updated.
    Fills other fields only if the existing cell is blank.
    (Now normalizes date fields at write-time.)
    """
    if not os.path.exists(LOCAL_EXCEL_PATH):
        raise FileNotFoundError(f"Excel not found at:\n{LOCAL_EXCEL_PATH}")

    backup_excel_once(LOCAL_EXCEL_PATH)

    wb = openpyxl.load_workbook(LOCAL_EXCEL_PATH)
    ws = wb.active
    ensure_headers(ws)

    row["causeno"] = normalize_causeno(row.get("causeno", ""))
    row["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not row["causeno"]:
        _save_excel_with_retry(wb, LOCAL_EXCEL_PATH)
        return

    header_index = {h: i+1 for i, h in enumerate(HEADERS)}
    cause_col = header_index["causeno"]

    # find existing row for this cause
    found_row_idx = None
    for r in range(2, ws.max_row + 1):
        existing = ws.cell(row=r, column=cause_col).value
        if existing and normalize_causeno(str(existing)) == row["causeno"]:
            found_row_idx = r
            break

    always_overwrite = {"causeno", "Dateappointed", "last_updated"}

    if found_row_idx:
        # update in place
        for h in HEADERS:
            col = header_index[h]
            new_val = row.get(h, "")
            cur_val = ws.cell(row=found_row_idx, column=col).value

            def _write(cell, key, val):
                if key in DATE_HEADERS:
                    is_date, payload, numfmt = _as_excel_date_or_text(val)
                    if is_date:
                        cell.value = payload
                        cell.number_format = numfmt
                    else:
                        cell.value = payload  # '' or normalized string
                else:
                    cell.value = val

            if h in always_overwrite:
                if new_val != "" and new_val is not None:
                    _write(ws.cell(row=found_row_idx, column=col), h, new_val)
                continue

            if new_val and (cur_val is None or str(cur_val).strip() == ""):
                _write(ws.cell(row=found_row_idx, column=col), h, new_val)
    else:
        # append a new row (normalize date cells before/after append)
        values = []
        for h in HEADERS:
            v = row.get(h, "")
            if h in DATE_HEADERS:
                is_date, payload, _fmt = _as_excel_date_or_text(v)
                values.append(payload)  # date obj or ''/normalized text
            else:
                values.append(v)
        ws.append(values)

        # apply number formats for date columns on the new row
        new_r = ws.max_row
        for h in DATE_HEADERS:
            if h in header_index:
                c = header_index[h]
                cell = ws.cell(row=new_r, column=c)
                ok, payload, numfmt = _as_excel_date_or_text(cell.value)
                if ok:
                    cell.value = payload
                    cell.number_format = numfmt

    _save_excel_with_retry(wb, LOCAL_EXCEL_PATH)


# ---------- Text cleanup & OCR ----------
def clean_text(s: str) -> str:
    s = re.sub(r"[ \t]+", " ", s or "")
    s = re.sub(r"\n+", "\n", s)
    return s.strip()

def normalize_unicode_noise(s: str) -> str:
    if not s: return ""
    trans = {
        "\u2018": "'", "\u2019": "'", "\u201C": '"', "\u201D": '"',
        "\u2013": "-", "\u2014": "-", "\u2022": " ", "\u00A0": " ",
        "\ufb01": "fi", "\ufb02": "fl", "–": "-", "—": "-", "•": " ",
    }
    for k, v in trans.items():
        s = s.replace(k, v)
    s = re.sub(r"[^\x09\x0A\x0D\x20-\x7E]", " ", s)
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{2,}", "\n", s)
    return s.strip()

def preprocess_for_ocr(img: Image.Image) -> Image.Image:
    g = ImageOps.grayscale(img)
    g = g.filter(ImageFilter.MedianFilter(size=3))
    g = ImageOps.autocontrast(g, cutoff=2)
    th = 175
    bw = g.point(lambda p: 255 if p > th else 0)
    return bw

def extract_text_with_pdfplumber(pdf_bytes: bytes) -> str:
    text_parts = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            txt = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
            text_parts.append(txt)
    raw = "\n".join(text_parts)
    return normalize_unicode_noise(clean_text(raw))

def extract_text_with_vision(pdf_bytes: bytes) -> str:
    """
    Google Vision OCR fallback. Safe if package/creds are missing.
    Converts PDF bytes to images and runs document_text_detection per page.
    Uses a function attribute to avoid repeated attempts in the same run.
    """
    # If we already failed init once this run, don't try again
    if getattr(extract_text_with_vision, "_hard_disabled", False):
        return ""

    # No creds: disable for this run
    if not os.path.exists(VISION_CREDENTIALS_FILE):
        log(f"⚠️ Vision credentials missing at {VISION_CREDENTIALS_FILE}; skipping Vision fallback.")
        extract_text_with_vision._hard_disabled = True
        return ""

    # Lazy import so missing package doesn't crash at import time
    try:
        from google.cloud import vision as _vision  # <- alias to avoid name clashes
        client = _vision.ImageAnnotatorClient.from_service_account_file(VISION_CREDENTIALS_FILE)
    except Exception as e:
        log(f"⚠️ Could not init Vision client: {e}")
        extract_text_with_vision._hard_disabled = True
        return ""

    try:
        # Convert PDF to images then OCR each page
        images = None
        
        # Try different conversion methods
        try:
            images = convert_from_bytes(pdf_bytes, dpi=300, poppler_path=POPPLER_BIN)
        except Exception as e1:
            log(f"  Vision: convert_from_bytes failed: {e1}")
            try:
                # Try without poppler path
                images = convert_from_bytes(pdf_bytes, dpi=300)
            except Exception as e2:
                log(f"  Vision: convert_from_bytes (no poppler) failed: {e2}")
                try:
                    # Try with different DPI
                    images = convert_from_bytes(pdf_bytes, dpi=150)
                except Exception as e3:
                    log(f"  Vision: convert_from_bytes (dpi=150) failed: {e3}")
                    return ""
        
        if not images:
            log("  Vision: No images converted from PDF")
            return ""
            
        parts = []
        for img in images:
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            image = _vision.Image(content=buf.getvalue())

            resp = client.document_text_detection(image=image)

            # Handle API errors cleanly
            if getattr(resp, "error", None) and getattr(resp.error, "message", ""):
                log(f"⚠️ Vision API error: {resp.error.message}")
                continue

            text = getattr(resp.full_text_annotation, "text", "") or ""
            if text:
                parts.append(text)

        raw = "\n".join(parts)
        return normalize_unicode_noise(clean_text(raw))
    except Exception as e:
        log(f"⚠️ Vision request failed: {e}")
        return ""



# ---------- Helpers ----------
# ---- WARD NAME HELPERS ----
QUALIFIERS = [
    r'\bthe\s+ward\b',
    r'\ba\s+minor\b',
    r'\ban?\s+adult\b',
    r'\ban?\s+incapacitated\s+person\b',
    r'\ban?\s+incapacitated\s+adult\b',
    r'\ban?\s+incapacitated\b',
    r'\bdeceased\b',
]
STOP_WORDS_AFTER_NAME = [
    r'\bage\b', r'\bdob\b', r'\baddress\b', r'\bphone\b', r'\bguardian', r'\bcause\s*no\b'
]
SUFFIXES = {"jr","jr.","sr","sr.","ii","iii","iv","v"}

def _strip_qualifiers(s: str) -> str:
    if not s: return ""
    t = s
    for sw in STOP_WORDS_AFTER_NAME:
        t = re.split(sw, t, flags=re.I)[0]
    t = re.sub(r'\bthe\s+estate\s+of\b', '', t, flags=re.I)
    t = re.sub(r'\bestate\s+of\b', '', t, flags=re.I)
    t = re.sub(r'\bthe\s+estate\b', '', t, flags=re.I)
    t = re.sub(r'\bestate\b', '', t, flags=re.I)
    for q in QUALIFIERS:
        t = re.sub(q, "", t, flags=re.I)
    t = re.sub(r'\s+', ' ', t)
    t = re.sub(r'[\s,;:\-]+$', '', t).strip(" -,:;")
    return t.strip()

def _split_first_last(raw: str) -> tuple[str, str, str]:
    """
    Split a name into (first, middle, last) components.
    Returns (first_name, middle_name, last_name) where middle_name can be empty.
    """
    if not raw: return ("", "", "")
    s = re.sub(r'\s+', ' ', raw).strip()
    
    # Filter out generic/placeholder names
    if s.lower() in ["only person", "person only", "ward", "incapacitated person", "a person"]:
        return ("", "", "")
    
    if "," in s:
        left, right = s.split(",", 1)
        last = left.strip()
        first = right.strip().split()[0] if right.strip() else ""
        return (first.title(), "", last.title())
    
    tokens = s.split()
    if len(tokens) < 2:
        return ("", "", "")
    
    # Handle suffixes like "Jr.", "Sr.", "III", etc.
    last = tokens[-1]
    prev = tokens[-2] if len(tokens) >= 2 else ""
    if last.lower() in SUFFIXES and prev:
        last = prev + " " + tokens[-1]
        tokens = tokens[:-1]
    
    # For names without commas, handle: first_name middle_name(s) last_name
    if len(tokens) >= 3:
        # Has middle name(s): "John Michael Smith" -> first="John", middle="Michael", last="Smith"
        first = tokens[0]
        middle = " ".join(tokens[1:-1])  # All tokens between first and last
        last = tokens[-1]
        return (first.title(), middle.title(), last.title())
    else:
        # Only first and last: "John Smith" -> first="John", middle="", last="Smith"
        first = tokens[0]
        last = tokens[-1]
        return (first.title(), "", last.title())

def _looks_like_human_name(s: str) -> bool:
    """
    Stricter: 2–4 words, letters/space/.-', at least 2 TitleCase words,
    no single-letter tokens, not a boilerplate label, no digits.
    """
    if not s:
        return False
    s = s.strip()
    if _NEVER_NAME_RE.search(s):
        return False
    if re.search(r"[^A-Za-z .'\-,\s]", s):  # reject slashes and other symbols, but allow commas
        return False
    if any(ch.isdigit() for ch in s):
        return False

    words = s.split()
    if not (1 <= len(words) <= 4):  # Allow single names for guardians
        return False

    # Must have at least 1 TitleCase token for single names, 2 for multi-word names
    # Also count "Jr.", "Sr.", "III", etc. as valid name parts
    # Allow all-caps names (like "MEENU JAIN")
    titled = sum(
        1 for w in words
        if (re.match(r"[A-Z][a-z]+(?:[.\-'][A-Za-z]+)?$", w) is not None or
            re.match(r"^[A-Z]{2,}$", w) is not None or  # All caps names like "JAIN", "MEENU"
            re.match(r"^(Jr\.?|Sr\.?|III|IV|V)$", w, re.I) is not None)
    )
    if len(words) == 1 and titled < 1:  # Single name needs at least 1 TitleCase or all-caps
        return False
    elif len(words) > 1 and titled < 2:  # Multi-word names need at least 2 TitleCase or all-caps
        return False

    # Avoid single-letter tokens like "C"
    if any(len(w) == 1 for w in words):
        return False

    return True

def extract_ward_name_candidates_from_order(t: str) -> list[str]:
    T = t or ""
    cands = []
    def looks_like_noise(line: str) -> bool:
        L = line.upper()
        bad_tokens = ("COURT", "COUNTY", "STATE", "TEXAS", "CAUSE", "NO", "NUMBER", "PROBATE")
        if any(bt in L for bt in bad_tokens):
            return True
        
        # Filter out generic/placeholder names
        generic_names = ("ONLY PERSON", "PERSON ONLY", "WARD", "INCAPACITATED PERSON", "A PERSON", "THE PERSON")
        if L.strip() in generic_names:
            return True
            
        return False
    
    # HIGHEST PRIORITY: "In the Guardianship of" followed by ward name on next line
    # This is the most reliable source since it's always typed in ORDER documents
    guardianship_probate_pattern = r'In\s+the\s+Guardianship\s+of\s*\n\s*([^\n]+?)(?:\s*\n\s*In\s+Probate\s+Court|\s*\n\s*In\s+the\s+Probate\s+Court|\s*\n\s*In\s+Probate|\s*\n\s*In\s+the\s+Probate)'
    m = re.search(guardianship_probate_pattern, T, re.I)
    if m:
        ward_name = m.group(1).strip()
        # Clean up the ward name (remove extra spaces, punctuation)
        ward_name = re.sub(r'\s+', ' ', ward_name).strip()
        ward_name = re.sub(r'[^\w\s\'-]', '', ward_name).strip()  # Keep only letters, spaces, hyphens, apostrophes
        # Only accept if it looks like a real human name (not OCR noise like "ess")
        if ward_name and len(ward_name) >= 3 and not looks_like_noise(ward_name) and _looks_like_human_name(ward_name):
            cands.append(ward_name)
            print(f"  Found ward name from 'In the Guardianship of' pattern: {ward_name!r}")
    
    # Also check individual lines for this pattern
    for line in T.splitlines():
        m = re.search(guardianship_probate_pattern, line, re.I)
        if m:
            ward_name = m.group(1).strip()
            ward_name = re.sub(r'\s+', ' ', ward_name).strip()
            ward_name = re.sub(r'[^\w\s\'-]', '', ward_name).strip()
            # Only accept if it looks like a real human name (not OCR noise like "ess")
            if ward_name and len(ward_name) >= 3 and not looks_like_noise(ward_name) and _looks_like_human_name(ward_name):
                cands.append(ward_name)
                print(f"  Found ward name from line 'In the Guardianship of' pattern: {ward_name!r}")
    pat_same = r'(?:IN\s+RE|IN\s+THE\s+MATTER\s+OF)\s*:?\s*(?:THE\s+)?(?:GUARDIANSHIP\s+OF|MATTER\s+OF)\s+(.+)$'
    for line in T.splitlines():
        m = re.search(pat_same, line, re.I)
        if m:
            cand = m.group(1).strip()
            if not looks_like_noise(cand):
                cands.append(cand)
    m = re.search(r'(?:IN\s+RE|IN\s+THE\s+MATTER\s+OF)\s*:?\s*(?:THE\s+)?(?:GUARDIANSHIP\s+OF|MATTER\s+OF)\s*$', T, re.I|re.M)
    if m:
        tail = T[m.end():].splitlines()
        if tail:
            nxt = tail[0].strip()
            if len(nxt) >= 3 and not looks_like_noise(nxt):
                cands.append(nxt)
    for line in T.splitlines():
        m = re.search(r'GUARDIANSHIP\s+OF\s+(.+)$', line, re.I)
        if m:
            cand = m.group(1).strip()
            if not looks_like_noise(cand):
                cands.append(cand)
    
    # Additional ORDER patterns for better coverage
    order_patterns = [
        # Look for ward name after "IN THE GUARDIANSHIP OF" on same line
        r'IN\s+THE\s+GUARDIANSHIP\s+OF\s+([A-Za-z\s]+?)(?:\s+IN\s+PROBATE|\s+IN\s+THE\s+PROBATE|$)',
        # Look for ward name in various formats
        r'GUARDIANSHIP\s+OF\s+([A-Za-z\s]+?)(?:\s+IN\s+PROBATE|\s+IN\s+THE\s+PROBATE|$)',
        # Look for ward name after "IN RE" patterns
        r'IN\s+RE\s+([A-Za-z\s]+?)(?:\s+IN\s+PROBATE|\s+IN\s+THE\s+PROBATE|$)',
        # Look for ward name in matter patterns
        r'IN\s+THE\s+MATTER\s+OF\s+([A-Za-z\s]+?)(?:\s+IN\s+PROBATE|\s+IN\s+THE\s+PROBATE|$)',
    ]
    
    for pattern in order_patterns:
        m = re.search(pattern, T, re.I | re.M)
        if m:
            ward_name = m.group(1).strip()
            # Clean up the ward name
            ward_name = re.sub(r'\s+', ' ', ward_name)
            ward_name = re.sub(r'[^\w\s\'-]', '', ward_name)
            if ward_name and len(ward_name) >= 3 and not looks_like_noise(ward_name) and _looks_like_human_name(ward_name):
                cands.append(ward_name)
                print(f"  Found ward name from ORDER pattern: {ward_name!r}")
    out = []
    for c in cands:
        c2 = _strip_qualifiers(c)
        c2 = re.split(r',\s*(an?|the)\b', c2, flags=re.I)[0].strip()
        if _looks_like_human_name(c2):
            out.append(c2)
    return out

def extract_ward_name_candidates_from_arp(t: str) -> list[str]:
    T = t or ""
    cands = []
    
    # PRIORITY 1: WARD block on ARP (most reliable for ARPs)
    # Look for ward name in the WARD block - this is the primary source for ARPs
    ward_block_patterns = [
        r'Ward\s*Name\s*[:\-]?\s*(.+)',
        r'Ward\s*:\s*Name\s*[:\-]?\s*(.+)', 
        r'Ward\s*:\s*(.+)',
        r'Ward\s*Name\s*[:\-]?\s*\n\s*(.+)',
        r'Ward\s*:\s*Name\s*[:\-]?\s*\n\s*(.+)',
        r'Ward\s*:\s*\n\s*(.+)',
        # Look for "WARD" block followed by name on same or next line
        r'WARD\s*[:\-]?\s*(.+)',
        r'WARD\s*[:\-]?\s*\n\s*(.+)',
        # Look for ward name in a block format
        r'Ward\s*[:\-]?\s*([A-Za-z\s]+?)(?:\n|$)',
    ]
    
    for pattern in ward_block_patterns:
        m = re.search(pattern, T, re.I | re.M)
        if m:
            ward_name = m.group(1).strip()
            # Clean up the ward name
            ward_name = re.sub(r'\s+', ' ', ward_name)
            ward_name = re.sub(r'[^\w\s\'-]', '', ward_name)
            if ward_name and len(ward_name) >= 3 and _looks_like_human_name(ward_name):
                cands.append(ward_name)
                print(f"  Found ward name from ARP WARD block: {ward_name!r}")
    
    # PRIORITY 2: Top of ARP (always typed, fallback for handwritten WARD blocks)
    # Look for ward name at the top of ARP documents - this is the most reliable source
    # The top section is always typed and contains the full official court name
    
    # Pattern 1: "In the Guardianship of" followed by ward name on next line
    guardianship_probate_pattern = r'In\s+the\s+Guardianship\s+of\s*\n\s*([^\n]+?)(?:\s*\n\s*In\s+Probate\s+Court|\s*\n\s*In\s+the\s+Probate\s+Court|\s*\n\s*In\s+Probate|\s*\n\s*In\s+the\s+Probate)'
    m = re.search(guardianship_probate_pattern, T, re.I)
    if m:
        ward_name = m.group(1).strip()
        ward_name = re.sub(r'\s+', ' ', ward_name).strip()
        ward_name = re.sub(r'[^\w\s\'-]', '', ward_name).strip()
        if ward_name and len(ward_name) >= 3 and _looks_like_human_name(ward_name):
            cands.append(ward_name)
            print(f"  Found ward name from ARP top 'In the Guardianship of': {ward_name!r}")
    
    # Pattern 2: "In the Guardianship of" on same line as ward name
    # This handles cases like "In the Guardianship of Jayleen Jaimes In Probate Court No. 1"
    guardianship_same_line_pattern = r'In\s+the\s+Guardianship\s+of\s+([A-Za-z\s]+?)(?:\s+In\s+Probate\s+Court|\s+In\s+the\s+Probate\s+Court|\s+In\s+Probate|\s+In\s+the\s+Probate|$)'
    m = re.search(guardianship_same_line_pattern, T, re.I | re.M)
    if m:
        ward_name = m.group(1).strip()
        ward_name = re.sub(r'\s+', ' ', ward_name).strip()
        ward_name = re.sub(r'[^\w\s\'-]', '', ward_name).strip()
        if ward_name and len(ward_name) >= 3 and _looks_like_human_name(ward_name):
            cands.append(ward_name)
            print(f"  Found ward name from ARP top same line: {ward_name!r}")
    
    # Pattern 2b: Handle cases where ward name appears between "In the Guardianship of" and "In Probate Court"
    # This specifically handles the format: "In the Guardianship of [WARD NAME] In Probate Court"
    guardianship_between_pattern = r'In\s+the\s+Guardianship\s+of\s+([A-Za-z\s]+?)\s+In\s+Probate\s+Court'
    m = re.search(guardianship_between_pattern, T, re.I)
    if m:
        ward_name = m.group(1).strip()
        ward_name = re.sub(r'\s+', ' ', ward_name).strip()
        ward_name = re.sub(r'[^\w\s\'-]', '', ward_name).strip()
        if ward_name and len(ward_name) >= 3 and _looks_like_human_name(ward_name):
            cands.append(ward_name)
            print(f"  Found ward name from ARP top between pattern: {ward_name!r}")
    
    # Pattern 2c: Handle cases where ward name appears on the line BEFORE "In the Guardianship of"
    # This handles OCR issues where the ward name gets separated from the guardianship text
    lines = T.splitlines()
    for i, line in enumerate(lines):
        if re.search(r'In\s+the\s+Guardianship\s+of', line, re.I):
            # Check the previous line for a potential ward name
            if i > 0:
                prev_line = lines[i-1].strip()
                # Look for a line that contains what looks like a full name
                # Allow for OCR errors like starting with lowercase letters
                if (re.search(r'^[A-Za-z][a-z]+(?:\s+[A-Za-z][a-z]+){1,3}$', prev_line) and
                    not re.search(r'(?:County|Clerk|Court|Probate|State|Texas|Travis|Filed|No\.)', prev_line, re.I) and
                    _looks_like_human_name(prev_line) and len(prev_line) >= 6):
                    cands.append(prev_line)
                    break
            
            # Also check the next few lines for ward names (for ORDER documents)
            for j in range(i+1, min(i+6, len(lines))):  # Check next 5 lines
                next_line = lines[j].strip()
                if next_line and not re.search(r'(?:ORDER|APPOINTING|COURT|VISITOR)', next_line, re.I):
                    # Clean up the line (remove trailing "OF" and other common OCR artifacts)
                    clean_line = re.sub(r'\s+OF\s*$', '', next_line).strip()
                    # Remove common suffixes like "Incapacitated Person", "an Incapacitated Person", etc.
                    clean_line = re.sub(r'\s+(?:an\s+)?incapacitated\s+person.*$', '', clean_line, flags=re.I).strip()
                    
                    # Check if it looks like a name (all caps or mixed case)
                    if (len(clean_line.split()) >= 2 and  # At least 2 words
                        not re.search(r'(?:COUNTY|CLERK|COURT|PROBATE|STATE|TEXAS|TRAVIS|FILED|NO\.)', clean_line, re.I) and
                        _looks_like_human_name(clean_line.title()) and len(clean_line) >= 6):
                        cands.append(clean_line.title())  # Convert to Title Case
                        break
    
    # Pattern 3: "IN THE GUARDIANSHIP OF" (all caps version)
    guardianship_caps_pattern = r'IN\s+THE\s+GUARDIANSHIP\s+OF\s*\n\s*([^\n]+?)(?:\s*\n\s*IN\s+PROBATE\s+COURT|\s*\n\s*IN\s+THE\s+PROBATE\s+COURT|\s*\n\s*IN\s+PROBATE|\s*\n\s*IN\s+THE\s+PROBATE)'
    m = re.search(guardianship_caps_pattern, T, re.I)
    if m:
        ward_name = m.group(1).strip()
        ward_name = re.sub(r'\s+', ' ', ward_name).strip()
        ward_name = re.sub(r'[^\w\s\'-]', '', ward_name).strip()
        if ward_name and len(ward_name) >= 3 and _looks_like_human_name(ward_name):
            cands.append(ward_name)
            print(f"  Found ward name from ARP top caps: {ward_name!r}")
    
    # Pattern 4: "IN THE GUARDIANSHIP OF" on same line (all caps)
    guardianship_caps_same_line_pattern = r'IN\s+THE\s+GUARDIANSHIP\s+OF\s+([A-Za-z\s]+?)(?:\s+IN\s+PROBATE\s+COURT|\s+IN\s+THE\s+PROBATE\s+COURT|\s+IN\s+PROBATE|\s+IN\s+THE\s+PROBATE|$)'
    m = re.search(guardianship_caps_same_line_pattern, T, re.I | re.M)
    if m:
        ward_name = m.group(1).strip()
        ward_name = re.sub(r'\s+', ' ', ward_name).strip()
        ward_name = re.sub(r'[^\w\s\'-]', '', ward_name).strip()
        if ward_name and len(ward_name) >= 3 and _looks_like_human_name(ward_name):
            cands.append(ward_name)
            print(f"  Found ward name from ARP top caps same line: {ward_name!r}")
    
    # Pattern 5: Look for ward name in the first few lines of the document
    # This catches cases where the format might be different
    # But be more selective to avoid picking up "Travis County Clerk" etc.
    lines = T.splitlines()
    for i, line in enumerate(lines[:10]):  # Check first 10 lines
        line_clean = line.strip()
        # Look for lines that contain what looks like a full name
        # But exclude common non-ward names
        if (re.search(r'^[A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3}$', line_clean) and
            not re.search(r'(?:County|Clerk|Court|Probate|State|Texas|Travis)', line_clean, re.I) and
            _looks_like_human_name(line_clean) and len(line_clean) >= 6):
            cands.append(line_clean)
            print(f"  Found ward name from ARP top lines (line {i+1}): {line_clean!r}")
            break  # Only take the first good match from top lines
    
    # Pattern 6: Look for ward name after common ARP document headers
    # This catches cases where the ward name appears after document headers
    arp_header_patterns = [
        r'(?:ANNUAL\s+REPORT\s+OF\s+PROGRESS|ARP|REPORT\s+OF\s+PROGRESS)\s*[:\-]?\s*\n\s*([A-Za-z\s]+?)(?:\n|$)',
        r'(?:GUARDIANSHIP\s+REPORT|PROGRESS\s+REPORT)\s*[:\-]?\s*\n\s*([A-Za-z\s]+?)(?:\n|$)',
        r'(?:WARD\s+NAME|WARD)\s*[:\-]?\s*\n\s*([A-Za-z\s]+?)(?:\n|$)',
    ]
    
    for pattern in arp_header_patterns:
        m = re.search(pattern, T, re.I | re.M)
        if m:
            ward_name = m.group(1).strip()
            ward_name = re.sub(r'\s+', ' ', ward_name)
            ward_name = re.sub(r'[^\w\s\'-]', '', ward_name)
            if ward_name and len(ward_name) >= 3 and _looks_like_human_name(ward_name):
                cands.append(ward_name)
                print(f"  Found ward name from ARP header pattern: {ward_name!r}")
                break  # Only take the first good match
    
    # PRIORITY 3: ORDER patterns (last resort - only when ORDER exists and ARP fails)
    # Only include ORDER patterns if we haven't found anything from ARP-specific patterns
    if not cands:
        print("  No ward name found in ARP-specific patterns, trying ORDER patterns as last resort...")
        cands += extract_ward_name_candidates_from_order(T)
    out = []
    for c in cands:
        c2 = _strip_qualifiers(c)
        if _looks_like_human_name(c2):
            out.append(c2)
    return out

def choose_best_ward_name(cands: list[str]) -> tuple[str, str, str]:
    """
    Choose the best ward name from candidates and return (first, middle, last).
    """
    if not cands:
        return ("", "", "")
    comma_cands = [c for c in cands if "," in c]
    pool = comma_cands if comma_cands else cands
    def score(s: str):
        tokens = len(s.split())
        return (abs(tokens - 3), -len(s))
    pool.sort(key=score)
    best = pool[0]
    if re.search(r'\bestate\b', best, re.I):
        for c in pool[1:]:
            if not re.search(r'\bestate\b', c, re.I):
                best = c
                break
    first, middle, last = _split_first_last(best)
    return (first, middle, last)

def grab(pattern, text, flags=re.I|re.S):
    m = re.search(pattern, text or "", flags)
    if not m:
        return ""
    try:
        return m.group(1).strip()
    except IndexError:
        return m.group(0).strip()

def to_slashes(date_str: str) -> str:
    if not date_str: return ""
    return date_str.replace(".", "/").replace("-", "/")

def normalize_role(s: str) -> str:
    if not s: return ""
    t = s.lower()
    t = t.replace("motherand father", "father/mother").replace("mother and father", "father/mother").replace("father and mother", "father/mother")
    t = t.replace("motherand", "mother").replace("fatherand", "father")
    t = re.sub(r'^[\s:;,.\-]+', ' ', t)  # strip leading punctuation like ": Mom"
    t = re.sub(r'[^a-z/ ]+', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    if "father/mother" in t or ("father" in t and "mother" in t): return "Father/Mother"
    if "mother" in t and "father" not in t: return "Mother"
    if "mom" in t and "mother" not in t: return "Mother"
    if "dad" in t and "father" not in t: return "Father"
    if "father" in t and "mother" not in t: return "Father"
    if "public guardian" in t or "public guardianship" in t: return "Public Guardian"
    if "parent" in t: return "Parent"
    if "son" in t: return "Son"
    if "daughter" in t: return "Daughter"
    return s.strip()

def fix_date_typos(s: str) -> str:
    if not s: return ""
    s = s.replace("q", "").replace("O", "0").replace("o", "0")
    s = s.replace("..", ".").replace("//", "/")
    return s.strip()


# --- ARP "Filed" date helpers ---
_MONTH_MAP = {
    'jan':'01','january':'01',
    'feb':'02','february':'02',
    'mar':'03','march':'03',
    'apr':'04','april':'04',
    'may':'05',
    'jun':'06','june':'06',
    'jul':'07','july':'07',
    'aug':'08','august':'08',
    'sep':'09','sept':'09','september':'09',
    'oct':'10','october':'10',
    'nov':'11','november':'11',
    'dec':'12','december':'12',
}

def _mm_from_month_name(name: str) -> str:
    return _MONTH_MAP.get((name or "").strip().lower(), "")

def _fmt_mdY(mm: str, dd: str|int, yyyy: str|int) -> str:
    """
    Return MM/DD/YYYY if inputs are sane, else "".
    NOTE: Capital Y in the name. Must be top-level (no indentation).
    """
    try:
        dd = int(dd); yyyy = int(yyyy)
        if 1 <= dd <= 31 and 1900 <= yyyy <= 2100 and mm:
            return f"{mm}/{dd:02d}/{yyyy}"
    except Exception:
        pass
    return ""

log(f"_fmt_mdY visible at import: {'_fmt_mdY' in globals()}")

def normalize_month_text_date(s: str) -> str:
    """
    Convert 'September 16, 2025' or 'Sep 16 2025' to '09/16/2025'.
    If already like 9/6/2025 after to_slashes(), returns zero-padded MM/DD/YYYY.
    Returns the original string if it doesn't look like a month-name date.
    """
    if not s:
        return ""
    t = (s or "").strip()

    # Month DD, YYYY  (optional comma, optional st/nd/rd/th)
    m = re.search(
        r'^\s*(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|'
        r'Aug(?:ust)?|Sep(?:t)?(?:ember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)'
        r'\s+(\d{1,2})(?:st|nd|rd|th)?\s*,?\s*(\d{4})\s*$', t, re.I
    )
    if m:
        mm = _mm_from_month_name(m.group(1))
        dd = int(m.group(2)); yyyy = int(m.group(3))
        return f"{mm}/{dd:02d}/{yyyy}" if mm else ""

    # YYYY Month DD  (e.g., '2025 Jul 22')
    m = re.search(
        r'^\s*(\d{4})\s+'
        r'(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|'
        r'Aug(?:ust)?|Sep(?:t)?(?:ember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)'
        r'\s+(\d{1,2})(?:st|nd|rd|th)?\s*$', t, re.I
    )
    if m:
        yyyy = int(m.group(1))
        mm = _mm_from_month_name(m.group(2))
        dd = int(m.group(3))
        return f"{mm}/{dd:02d}/{yyyy}" if mm else ""

    # Already numeric MM/DD/YYYY (allow 1/2/2025 style; normalize)
    m = re.search(r'^\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$', t)
    if m:
        mo, da, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{mo:02d}/{da:02d}/{yr}"

    # Not a recognized month-text date; return as-is
    return t

def extract_arp_filed_date(t: str) -> str:
    """
    Tries to recognize clerk stamps such as:
      'Filed for Record 2025 Jul 22'
      'Filed for Record Jul 22, 2025'
      'Entered for Record JUL 22 2025'
      'Filed: Jul 22, 2025'
      'Filed 07/22/2025'
    Returns MM/DD/YYYY or "".
    Note: Ignores "Updated" dates as they are just form dates.
    """
    if not t: return ""

    # Numeric near 'Filed/Entered' (including 'Filed for Record')
    m = re.search(r'(Filed(?:\s+for\s+Record)?|Entered(?:\s+for\s+Record)?)\b.{0,60}?(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})', t, re.I|re.S)
    if m:
        raw = m.group(2).replace(".", "/").replace("-", "/")
        parts = raw.split("/")
        if len(parts) == 3 and len(parts[2]) == 2:
            y2 = int(parts[2]); parts[2] = f"20{y2:02d}" if y2 < 50 else f"19{y2:02d}"
            raw = "/".join(parts)
        return raw

    # 'Month 22, 2025' near 'Filed/Entered' (including 'Filed for Record')
    m = re.search(
        r'(Filed(?:\s+for\s+Record)?|Entered(?:\s+for\s+Record)?)\b.{0,60}?'
        r'((Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|'
        r'Aug(?:ust)?|Sep(?:t)?(?:ember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)'
        r'\s+\d{1,2}(?:st|nd|rd|th)?\s*,?\s*\d{4})',
        t, re.I|re.S
    )
    if m:
        return normalize_month_text_date(m.group(2))

    # 'YYYY Mon 22' near 'Filed/Entered' (including 'Filed for Record')
    m = re.search(
        r'(Filed(?:\s+for\s+Record)?|Entered(?:\s+for\s+Record)?)\b.{0,60}?(\d{4})\s+'
        r'(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|'
        r'Aug(?:ust)?|Sep(?:t)?(?:ember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)'
        r'\s+(\d{1,2})(?:st|nd|rd|th)?',
        t, re.I|re.S
    )
    if m:
        mm = _mm_from_month_name(m.group(3))
        out = _fmt_mdY(mm, m.group(4), m.group(2))
        if out: return out

    # 'Mon 22 2025' near 'Filed/Entered' (including 'Filed for Record')
    m = re.search(
        r'(Filed(?:\s+for\s+Record)?|Entered(?:\s+for\s+Record)?)\b.{0,60}?'
        r'(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|'
        r'Aug(?:ust)?|Sep(?:t)?(?:ember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)'
        r'\s+(\d{1,2})(?:st|nd|rd|th)?\s+(\d{4})',
        t, re.I|re.S
    )
    if m:
        mm = _mm_from_month_name(m.group(2))
        out = _fmt_mdY(mm, m.group(3), m.group(4))
        if out: return out

    # Last resort: anywhere 'YYYY Mon DD'
    m = re.search(
        r'(\d{4})\s+'
        r'(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|'
        r'Aug(?:ust)?|Sep(?:t)?(?:ember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)'
        r'\s+(\d{1,2})(?:st|nd|rd|th)?',
        t, re.I
    )
    if m:
        mm = _mm_from_month_name(m.group(2))
        out = _fmt_mdY(mm, m.group(3), m.group(1))
        if out: return out

    return ""

def normalize_ocr_confusions(t: str) -> str:
    s = t or ""
    s = s.replace("—", "-").replace("–", "-").replace("−", "-")
    s = s.replace("\u00A0", " ").replace("：", ":")
    s = re.sub(r'\bS1gned\b', 'Signed', s, flags=re.I)
    s = re.sub(r'\bSlgned\b', 'Signed', s, flags=re.I)
    s = re.sub(r'\bSigned\s*[o0]n\b', 'Signed on', s, flags=re.I)
    s = re.sub(r'Ca(u)?se\s*No\.?', 'Cause No.', s, flags=re.I)
    s = re.sub(r'\bause\s*No\.?', 'Cause No.', s, flags=re.I)
    s = re.sub(r'C\s*[-–—]?\s*1\s*[-–—]?\s*PB', 'C-1-PB', s, flags=re.I)
    s = re.sub(r'[ \t]+', ' ', s)
    return s

def clean_ocr_underscores(text: str) -> str:
    """Clean up OCR text that has underscores inserted between characters."""
    if not text:
        return ""
    
    # Pattern: _char_char_char_ -> charchar
    # But preserve underscores that are part of valid patterns (like email domains)
    
    # First, clean up obvious OCR underscore patterns
    # Pattern: _char_char_char_ where char is a letter or number
    result = re.sub(r'_([a-zA-Z0-9])_([a-zA-Z0-9])_([a-zA-Z0-9])_', r'\1\2\3', text)
    result = re.sub(r'_([a-zA-Z0-9])_([a-zA-Z0-9])_', r'\1\2', result)
    result = re.sub(r'_([a-zA-Z0-9])_', r'\1', result)
    
    # Clean up remaining single underscores that are clearly OCR artifacts
    # But preserve underscores in email addresses and other valid contexts
    result = re.sub(r'_([a-zA-Z0-9])', r'\1', result)
    result = re.sub(r'([a-zA-Z0-9])_', r'\1', result)
    
    return result

def _looks_like_two_addresses(s: str) -> bool:
    """
    Heuristic: return True if the string likely contains two separate addresses.
    Signals:
      - 2+ ZIP codes (##### or #####-####)
      - 2+ street-type tokens *and* a separator like ';', '/', or ' and '
      - 2+ state abbreviations (TX, CA, etc.)
    """
    if not s:
        return False

    # ZIP codes
    if len(re.findall(r'\b\d{5}(?:-\d{4})?\b', s)) >= 2:
        return True

    # Street tokens + a likely separator
    street_tokens = re.findall(
        r'\b('
        r'st|street|rd|road|dr|drive|ln|lane|ct|court|ave|avenue|blvd|boulevard|'
        r'pkwy|parkway|ter|terrace|pl|place|way|loop|trail|pass|cove|cir|circle|'
        r'hwy|highway'
        r')\b',
        s, re.IGNORECASE
    )
    if len(street_tokens) >= 2 and (';' in s or ' / ' in s or ' and ' in s.lower()):
        return True

    # 2+ state abbreviations
    state_re = r'\b(A[LKZR]|C[AOT]|D[EC]|F[LM]|G[AU]|HI|I[ADLN]|K[SY]|LA|M[ADEHINOPST]|N[CDEHJMVY]|O[HKR]|P[AWR]|RI|S[CD]|T[NX]|UT|V[AIT]|W[AIVY])\b'
    if len(re.findall(state_re, s)) >= 2:
        return True

    return False

def normalize_phone(s: str) -> str:
    if not s: return ""
    digits = re.sub(r"\D", "", str(s))
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return str(s).strip()

def clean_address(raw: str) -> str:
    """
    Clean up addresses but keep City, State ZIP.
    Strips '(no P.O. Box)', leading 'Address:' labels, and parenthetical blobs.
    Collapses line breaks into spaces.
    Also removes any accidental section headers like '2. GUARDIAN(s): Name(s)' that leak in.
    """
    s = (raw or "").strip()
    s = re.sub(r'^\s*(Address|Addr\.?|Residence|Mailing\s*Address)\s*[:\-]?\s*', '', s, flags=re.I)
    s = re.sub(r'\(.*?P\.?\s*O\.?\s*Box.*?\)', '', s, flags=re.I)
    s = re.sub(r'^\s*\(.*?\)\s*', '', s)
    s = re.sub(r'[\r\n]+', ' ', s)
    s = re.sub(r'\s{2,}', ' ', s)
    s = s.strip(' ,.-')
    s = s.replace(' ,', ',')

    # Repair missing spaces often caused by OCR:
    # - Insert a space between a leading house number and the street name (e.g., '101Acapulco' -> '101 Acapulco')
    s = re.sub(r'^(\d{1,6})([A-Za-z])', r'\1 \2', s)
    # - Insert spaces between camel-cased tokens (e.g., 'AcapulcoCourt' -> 'Acapulco Court')
    s = re.sub(r'([a-z])([A-Z])', r'\1 \2', s)
    # - Normalize punctuation before state abbreviations (e.g., 'Austin. TX' -> 'Austin, TX')
    s = re.sub(r'([A-Za-z])\.[ ]+([A-Z]{2}\b)', r'\1, \2', s)

    # NEW: if a section header leaked into the address, cut it (and anything after) off.
    s = re.split(
        r'\s*(?:,?\s*)?(?:\d+\.\s*)?(?:GUARDIAN\(s\)|Guardian\(s\)|Name\(s\)|Visit\s*Date|Visit\s*Time|Cause\s*No\.?)\b',
        s, maxsplit=1
    )[0].strip(' ,.-')

    return s

def looks_like_name(s: str) -> bool:
    if not s: return False
    t = s.strip()
    if re.match(r"^[A-Z][A-Za-z'\-]+,\s*[A-Z][A-Za-z'\-]+(?:\s+[A-Z]\.)?$", t):
        return True
    if re.match(r"^[A-Z][A-Za-z'\-]+(?:\s+[A-Z][A-Za-z'\-]+){1,2}$", t):
        return True
    if re.search(r",\s*[A-Z]{2}\b", t):
        return False
    return False

def find_block_after_label(text: str, label_pattern: str, max_lines: int = 3) -> str:
    m = re.search(label_pattern + r'\s*[:\-]?\s*(.+)', text or "", re.I)
    if not m:
        m = re.search(label_pattern + r'\s*[:\-]?\s*\n(.+)', text or "", re.I)
        if not m:
            return ""
    start = m.end(0)
    chunk = (text or "")[start:].splitlines()
    vals = []
    for i in range(min(max_lines, len(chunk))):
        ln = chunk[i].strip()
        if not ln:
            break
        pure = re.sub(r"[:\-]+", "", ln).strip().lower()
        if pure in LABEL_WORDS:
            break
        vals.append(ln)
    return ", ".join(vals).strip(", ").strip()

def find_after_label(text: str, label_pattern: str, max_chars: int = 120) -> str:
    """
    Read same-line or next-line value after a label. Rejects values that look like labels.
    """
    t = text or ""
    def _clean(v: str) -> str:
        v = v.splitlines()[0].strip()
        pure = re.sub(r"[:\-]+", "", v).strip().lower()
        if pure in LABEL_WORDS:
            return ""
        return v
    m = re.search(label_pattern + r'\s*[:\-]?\s*(.+)', t, re.I)
    if m:
        v = _clean(m.group(1))
        if v: return v[:max_chars].strip()
    m = re.search(label_pattern + r'\s*[:\-]?\s*\n\s*(.+)', t, re.I)
    if m:
        v = _clean(m.group(1))
        if v: return v[:max_chars].strip()
    return ""

def parse_liveswith_guardian(text: str) -> str | None:
    """
    Checkbox-only reader for: 'Do you reside with the ward?  [ ] YES  [ ] NO'
    Returns:
      "Guardian"  -> if YES checked (or both YES & NO)
      ""          -> if only NO is checked (means 'does not live with guardian')
      None        -> if the line/boxes can't be confidently found
    """
    if not text:
        return None
    m = re.search(r'Do\s+you\s+(?:reside|live)\s+with\s+the\s+ward', text, re.I)
    if not m:
        return None
    window = text[m.start(): m.start() + 250]
    for ch in ['☑','☒','■','█','✔','✓','✅','❎','❌','✘','✗']:
        window = window.replace(ch, 'X')
    window = re.sub(r'\[(?:x|X)\]', 'X', window)
    window = re.sub(r'\((?:x|X)\)', 'X', window)
    def _checked(word: str) -> bool:
        return bool(
            re.search(rf'\bX\s*{word}\b', window, re.I) or
            re.search(rf'\b{word}\s*X\b', window, re.I)
        )
    yes = _checked('YES')
    no  = _checked('NO')
    if yes and not no:
        return "Guardian"
    if no and not yes:
        return ""
    if yes and no:
        return "Guardian"
    return None

def safe_after_label(text: str, label_pat: str, expect: str = "any", window: int = 300) -> str:
    """
    More defensive version for tricky fields (phones/emails/dates/addresses).
    """
    t = text or ""
    m = re.search(label_pat + r'\s*[:\-]?\s*(.+)', t, re.I)
    val = ""
    if m:
        line = (m.group(1) or "").splitlines()[0].strip()
        pure = re.sub(r"[:\-]+", "", line).strip().lower()
        if line and pure not in LABEL_WORDS:
            val = line
    def _valid(v: str) -> bool:
        if not v: return False
        if expect == "phone":  return PHONE_RE.search(v) is not None
        if expect == "email":  return EMAIL_RE.search(v) is not None
        if expect == "date":   return DATE_RE.search(v) is not None
        if expect == "address":
            return bool(re.search(r"\d", v) or "," in v or
                        re.search(r"\b(po\s*box|suite|apt|unit|st|ave|rd|dr|ln|blvd|ct|pkwy|trl|trail|pass)\b", v, re.I))
        pure = re.sub(r"[:\-]+", "", v).strip().lower()
        return pure not in LABEL_WORDS
    if not _valid(val):
        anchor = re.search(label_pat, t, re.I)
        if anchor:
            tail = t[anchor.end(): anchor.end()+window]
            for ln in tail.splitlines()[:5]:
                s = ln.strip()
                if not _valid(s):
                    continue
                val = s
                break
    return val.strip()

def capture_address_after_label(text: str, label_pattern: str, max_lines: int = 3) -> str:
    """
    Capture an address that follows a label (e.g., '(no P.O. Box) ...').
    Strategy:
      1) Start at the label; take the same-line remainder and the next few lines.
      2) Strip label text ('Address:', 'City/State/Zip:') and '(no P.O. Box)' noise.
      3) Try to find a full one-line address with ADDRESS_RE.
      4) If not found, stitch a street line (+ optional unit) with a nearby city/state/zip line.
      5) Fallback to the best-looking street-ish single line.
    Returns a cleaned single-line address, or "" if nothing confident found.
    """
    if not text:
        return ""

    # 1) Anchor at the label
    m_lab = re.search(label_pattern, text, re.IGNORECASE)
    if not m_lab:
        return ""

    # Small sliding window after the label (keep it local to avoid other sections)
    window = text[m_lab.end(): m_lab.end() + 700]
    raw_lines = [ln.strip() for ln in window.splitlines()]

    # If same-line remainder exists (Label: value...), include it first
    m_same = re.search(label_pattern + r'\s*[:\-]?\s*(.+)', text, re.IGNORECASE)
    preface = []
    if m_same:
        same = (m_same.group(1) or "").splitlines()[0].strip()
        if same:
            preface = [same]

    # Consider only the first few meaningful lines after the label
    lines = [ln for ln in raw_lines if ln.strip()][:max_lines]
    block_lines = preface + lines

    # 2) Clean label text and PO Box notes from each line
    cleaned_lines = []
    for ln in block_lines:
        if not ln:
            continue
        ln2 = ln
        # Remove typical label prefixes that break regex matches
        ln2 = re.sub(r'^\s*(Address|Addr\.?|Residence|Mailing\s*Address|City/State/Zip)\s*[:\-]?\s*', '', ln2, flags=re.I)
        # Remove PO Box note and generic leading parentheses
        ln2 = re.sub(r'\(.*?P\.?\s*O\.?\s*Box.*?\)', '', ln2, flags=re.I)
        ln2 = re.sub(r'^\s*\(.*?\)\s*', '', ln2)
        ln2 = ln2.strip(' ,.-')
        if ln2:
            cleaned_lines.append(ln2)

    if not cleaned_lines:
        return ""

    # Join a paragraph for scanning full one-line addresses
    para = " ".join(cleaned_lines)

    # 3) First choice: full one-line address found anywhere in the short paragraph
    m_full = ADDRESS_RE.search(para)
    if m_full:
        return clean_address(m_full.group(0))

    # 4) Second choice: stitch STREET (+ optional UNIT) + CITY/STATE/ZIP from adjacent lines
    street_idx = None
    for i, ln in enumerate(cleaned_lines):
        if STREET_LINE_RE.match(ln):
            street_idx = i
            break

    if street_idx is not None:
        # Optional unit line immediately after street
        unit_piece = ""
        if street_idx + 1 < len(cleaned_lines) and UNIT_LINE_RE.search(cleaned_lines[street_idx + 1]):
            unit_piece = " " + cleaned_lines[street_idx + 1].strip()
        # Look ahead 1-3 lines for the city/state/zip
        for j in range(street_idx + 1, min(len(cleaned_lines), street_idx + 4)):
            m_city = CITY_STATE_ZIP_RE.match(cleaned_lines[j])
            if m_city:
                stitched = f"{cleaned_lines[street_idx]}{unit_piece}, {cleaned_lines[j]}"
                return clean_address(stitched)

    # 5) Fallback: choose a street-ish single line if nothing else hits
    for ln in cleaned_lines:
        if re.search(r'\d', ln) and re.search(
            r'\b(St|Street|Ave|Avenue|Rd|Road|Dr|Drive|Ln|Lane|Blvd|Boulevard|Ct|Court|Pl|Place|Pkwy|Parkway|Trl|Trail|Ter|Terrace)\b',
            ln, re.IGNORECASE
        ):
            return clean_address(ln)

    return ""

def _slice_between(text: str, start_pat: str, end_pat: str, max_len: int = 1600) -> str:
    """
    Return text between start_pat and end_pat (first matches).
    If end_pat not found, return up to max_len chars after start.
    """
    if not text:
        return ""
    m_start = re.search(start_pat, text, re.I)
    if not m_start:
        return ""
    start = m_start.end()
    m_end = re.search(end_pat, text[start:], re.I)
    if m_end:
        end = start + m_end.start()
    else:
        end = min(len(text), start + max_len)
    return text[start:end]


# ---------- Guardian fallback ----------
def extract_guardians_from_text(t: str):
    t = re.sub(r"[ \t]+", " ", t or "")
    emails = re.findall(EMAIL_RE, t)
    phones = re.findall(PHONE_RE, t)
    dobs   = re.findall(DATE_RE, t)
    name_hits = re.findall(r"[A-Z][A-Za-z'\-]+,\s*[A-Z][A-Za-z'\-]+", t)
    if not name_hits:
        name_hits = re.findall(r"\b[A-Z][a-z]+ [A-Z][a-z]+\b", t)
    def pop_or_empty(lst):
        return lst.pop(0) if lst else ""
    g1 = {"name": "", "email": "", "phone": "", "dob": "", "address": ""}
    g2 = {"name": "", "email": "", "phone": "", "dob": "", "address": ""}
    if name_hits:
        g1["name"] = name_hits[0]
    if len(name_hits) >= 2:
        g2["name"] = name_hits[1]
    g1["email"], g2["email"] = pop_or_empty(emails), pop_or_empty(emails)
    g1["phone"], g2["phone"] = pop_or_empty(phones), pop_or_empty(phones)
    g1["dob"],   g2["dob"]   = pop_or_empty(dobs),   pop_or_empty(dobs)
    g1["phone"] = normalize_phone(g1["phone"])
    g2["phone"] = normalize_phone(g2["phone"])
    g1["dob"]   = to_slashes(g1["dob"])
    g2["dob"]   = to_slashes(g2["dob"])
    return g1, g2

# --- Guardian name filtering helpers ---

# Street-type tokens to exclude lines like "Largo Cove", "Acapulco Court", etc.
_STREET_WORDS = (
    "St|Street|Rd|Road|Dr|Drive|Ln|Lane|Ct|Court|Ave|Avenue|Blvd|Boulevard|"
    "Pkwy|Parkway|Ter|Terrace|Pl|Place|Way|Loop|Trail|Pass|Cove|Circle|Cir|Hwy|Highway"
)

# Things that sometimes appear in OCR where a name should be
_LABEL_BLACKLIST = r"(?:Address|New\s+Address|Same\s+Address)"

def _looks_like_city_state_line(s: str) -> bool:
    s2 = (s or "").strip()
    if re.search(r"\b[A-Z]{2}\b", s2): return True            # TX, CA, etc.
    if re.search(r"\bTexas\b", s2, re.I): return True         # full state name
    if re.search(r"\d{5}(?:-\d{4})?\b", s2): return True      # zip code
    if re.match(r"^[A-Za-z .'\-]+,\s*[A-Za-z]{2,}$", s2): return True  # City, ST
    return False

def _filter_guardian_names(names: list[str]) -> list[str]:
    if not names:
        return []
    # Remove label/placeholder noise
    names = [n for n in names if not re.search(_LABEL_BLACKLIST, n, re.I)]
    # Remove street-like fragments (e.g., "Largo Cove", "Acapulco Court")
    names = [n for n in names if not re.search(rf"\b{_STREET_WORDS}\b", n, re.I)]
    # Remove city/state/zip-ish lines
    names = [n for n in names if not _looks_like_city_state_line(n)]
    # Only keep human-like names using your EXISTING _looks_like_human_name helper
    names = [n for n in names if _looks_like_human_name(n)]
    return names


# ---------- ORDER parsing ----------
def extract_order_date(text: str) -> str:
    t = normalize_ocr_confusions(text)
    m = re.search(r'\bSigned\b\s*:?\s*(?:on\s*)?([A-Za-z]{3,9}\s+\d{1,2}\s*,\s*\d{4}|\d{1,2}[./-]\d{1,2}[./-]\d{2,4})', t, re.I|re.S)
    if m:
        val = m.group(1).strip()
        if re.match(r'^\d{1,2}[./-]\d{1,2}[./-]\d{2,4}$', val): return to_slashes(val)
        return re.sub(r'\s*,\s*', ', ', val)
    m = re.search(r'\bSigned\s*on\b\s*:?\s*([0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{2,4})(?:\s+\d{1,2}:\d{2}\s*(?:AM|PM)?)?', t, re.I|re.S)
    if m: return to_slashes(m.group(1))
    m = re.search(r'\bSigned\s*on\s*this\s*the\s*(\d{1,2})(?:st|nd|rd|th)?\s*day\s*of\s*([A-Za-z]{3,9})\s*,?\s*(\d{4})', t, re.I|re.S)
    if m: return f"{m.group(2)} {m.group(1)}, {m.group(3)}"
    m = re.search(r'\b(Order\s*signed|Ordered\s*on)\s*:?\s*([0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{2,4})', t, re.I|re.S)
    if m: return to_slashes(m.group(2))
    m_anchor = re.search(r'\bSigned\b\s*:?\s*(?:on\b)?', t, re.I|re.S)
    if m_anchor:
        window = t[m_anchor.end(): m_anchor.end()+250]
        m = re.search(r'([A-Za-z]{3,9}\s+\d{1,2}\s*,\s*\d{4})', window, re.I|re.S)
        if m: return re.sub(r'\s*,\s*', ', ', m.group(1).strip())
        m = re.search(r'(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})', window, re.I|re.S)
        if m: return to_slashes(m.group(1))
    j = re.search(r'Judge|Presiding\s*Judge|Court\s*Judge', t, re.I)
    if j:
        window = t[max(0, j.start()-400): j.start()]
        m = re.search(r'([A-Za-z]{3,9}\s+\d{1,2}\s*,\s*\d{4})', window, re.I)
        if m: return re.sub(r'\s*,\s*', ', ', m.group(1).strip())
        m = re.search(r'(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})', window, re.I)
        if m: return to_slashes(m.group(1))
    m = re.search(r'([A-Za-z]{3,9}\s+\d{1,2}\s*,\s*\d{4})', t, re.I|re.S)
    if m: return re.sub(r'\s*,\s*', ', ', m.group(1).strip())
    m = re.search(r'(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})', t, re.I|re.S)
    if m: return to_slashes(m.group(1))
    return ""

def parse_order_fields(text: str, pdf_bytes_for_vision: bytes | None = None) -> dict | None:
    """
    For ORDER PDFs: return row with causeno + Dateappointed + wardfirst/wardlast (if found).
    Accepts optional pdf_bytes_for_vision for a single Vision retry when needed.
    """
    def _try_parse(t: str) -> dict | None:
        data = {h: "" for h in HEADERS}
        tt = normalize_ocr_confusions(t)
        cause = (
            grab(r'(?:Cause\s*No\.?\s*[:#]?\s*)(C[\s\-]?1[\s\-]?PB[\s\-]?\d{2}[\s\-]?\d+)', tt)
            or grab(r'\b(C[\s\-]?1[\s\-]?PB[\s\-]?\d{2}[\s\-]?\d+)\b', tt)
            or grab(r'(?:No\.?\s*)(C[\s\-]?1[\s\-]?PB[\s\-]?\d{2}[\s\-]?\d+)', tt)
        )
        if not cause:
            tail = grab(r'\b(\d{2}[-]\d{5,6})\b', tt)
            if tail:
                cause = f"C-1-PB-{tail}"
        cause = normalize_causeno(cause) if cause else ""
        if not cause:
            return None

        # Normalize appointment date
        date_appt_raw = extract_order_date(tt)
        if not date_appt_raw:
            return None
        date_appt = normalize_month_text_date(to_slashes(date_appt_raw))

        cands = extract_ward_name_candidates_from_order(tt)
        wf, wm, wl = choose_best_ward_name(cands)

        data["causeno"] = cause
        data["Dateappointed"] = date_appt
        if wf and wl:
            data["wardfirst"] = wf
            data["wardmiddle"] = wm
            data["wardlast"] = wl
        return data

    row = _try_parse(text)
    if row:
        return row
    if pdf_bytes_for_vision:
        try:
            t_v = extract_text_with_vision(pdf_bytes_for_vision)
            row = _try_parse(t_v)
            if row:
                return row
        except Exception:
            pass
    return None

# ---------- ARP parsing ----------
def guardian_signal_score(text: str) -> int:
    if not text: return 0
    t = text
    total = 0
    for m in re.finditer(r'Guardian', t, re.I):
        window = t[max(0, m.start()-250): m.end()+600]
        total += len(re.findall(r"[A-Za-z0-9@]", window))
    if total == 0:
        total = len(re.findall(r"[A-Za-z0-9@]", t))
    return total

def best_arp_text_from_tesseract(pdf_bytes: bytes, current_text: str | None = None) -> str:
    t4 = extract_text_with_ocr_for_arp(pdf_bytes, psm=4)
    t6 = extract_text_with_ocr_for_arp(pdf_bytes, psm=6)
    candidates = [(t4, guardian_signal_score(t4)), (t6, guardian_signal_score(t6))]
    if current_text:
        candidates.append((current_text, guardian_signal_score(current_text)))
    best = max(candidates, key=lambda x: x[1])[0]
    return best

# ---------- Post-OCR normalization (gentle) ----------
def improve_mapping(data: dict) -> dict:
    data = {k: (v or "").strip() for k, v in data.items()}

    # Phones
    for k in ("wtele", "gtele", "g2tele"):
        raw = data.get(k, "").strip()
        m = PHONE_RE.search(raw)
        data[k] = normalize_phone(m.group(0)) if m else ("" if raw.lower() in LABEL_WORDS else raw)

    # Dates
    for k in ("wdob", "gdob", "g2dob", "visitdate", "Dateappointed", "DateARPfiled"):
        raw = to_slashes(data.get(k, ""))
        m = DATE_RE.search(raw)
        data[k] = fix_date_typos(m.group(0)) if m else ("" if raw.lower() in LABEL_WORDS else raw)
    # Force Dateappointed into MM/DD/YYYY even when extracted as "Month 16, 2025"
    data["Dateappointed"] = normalize_month_text_date(data.get("Dateappointed", ""))
    # Clamp future years for key dates
    if data.get("DateARPfiled"):
        data["DateARPfiled"] = _clamp_future_year_to_current(data["DateARPfiled"]) or data["DateARPfiled"]
    if data.get("Dateappointed"):
        data["Dateappointed"] = _clamp_future_year_to_current(data["Dateappointed"]) or data["Dateappointed"]

    # Emails
    for k in ("gemail", "g2eamil"):
        raw = data.get(k, "").strip()
        data[k] = raw if ("@" in raw and "." in raw) else ("" if raw.lower() in LABEL_WORDS else raw)

    # Names
    for k in ("guardian1", "Guardian2", "wardfirst", "wardlast"):
        v = re.sub(r"\s+", " ", data.get(k, "")).strip()
        if v and v.lower() not in LABEL_WORDS:
            if "," in v:
                parts = [p.strip() for p in v.split(",")]
                if len(parts) >= 2:
                    v = f"{parts[0].title()}, {parts[1].title()}"
            else:
                v = " ".join(p.capitalize() for p in v.split())
            data[k] = v

    # Drop obvious estate artifacts in ward names
    if re.search(r"\bestate\b", data.get("wardfirst", ""), re.I):
        data["wardfirst"] = ""
    if re.search(r"\bestate\b", data.get("wardlast", ""), re.I):
        data["wardlast"] = ""

    # Addresses & liveswith
    for k in ("waddress", "gaddress", "g2 address", "liveswith"):
        v = data.get(k, "").strip()
        data[k] = "" if v.lower() in LABEL_WORDS else re.sub(r"\s+", " ", v).replace(" ,", ",")[:200]

    # ---- Mirror guardian address to Guardian2 ONLY when it’s clearly a shared address ----
    log(
        "Mirror check -> G2 present="
        + str(bool(data.get("Guardian2")))
        + " | g2 addr blank="
        + str(not bool(data.get("g2 address")))
        + " | g1 addr present="
        + str(bool(data.get("gaddress")))
    )

    # normalize any accidental 'g2address' key to 'g2 address'
    if "g2address" in data and not data.get("g2 address"):
        data["g2 address"] = data.pop("g2address")

    if data.get("Guardian2") and not data.get("g2 address") and data.get("gaddress"):
        if not _looks_like_two_addresses(data["gaddress"]):
            data["g2 address"] = data["gaddress"]
            log("  Mirrored gaddress → g2 address (single/shared address detected).")
        else:
            log("  NOT mirrored: looks like two separate addresses.")

    # If guardian1 looks like "First and/and Second Last", split into two names
    if data.get("guardian1") and not data.get("Guardian2"):
        m = re.match(
            r"^\s*([A-Z][a-z]+)\s+(?:&|and)\s+([A-Z][a-z]+)\s+([A-Z][a-z]+)\s*$",
            data["guardian1"],
        )
        if m:
            first1, first2, last = m.groups()
            data["guardian1"] = f"{first1} {last}"
            data["Guardian2"] = f"{first2} {last}"

    # Cause number
    data["causeno"] = normalize_causeno(data.get("causeno", ""))

    return data


import re
from typing import List, Tuple, Optional

# --- Junk filters ---
_STREET_WORDS = (
    "St|Street|Rd|Road|Dr|Drive|Ln|Lane|Ct|Court|Ave|Avenue|Blvd|Boulevard|"
    "Pkwy|Parkway|Ter|Terrace|Pl|Place|Way|Loop|Trail|Pass|Cove|Circle|Cir|Hwy|Highway"
)
_LABEL_BLACKLIST = r"Address|New\s+Address|Same\s+Address|Guardian(?:s)?\b|Phone|Email|Cause\b|No\b|#:?"
_PHONE_RE = re.compile(r"\(?\d{3}\)?[-/.\s]?\d{3}[-/.\s]?\d{4}")
# If other code expects PHONE_RE (without underscore), expose it:
PHONE_RE = _PHONE_RE

_EMAIL_RE = re.compile(r"[^@ \t\r\n]+@[^@ \t\r\n]+\.[^@ \t\r\n]+")
_CITY_STATE_TOKEN_BLOCK = {"ZIP", "CITY", "STATE", "TX", "TEXAS", "AUSTIN"}  # Expandable
_ZIP_RE = re.compile(r"\b\d{5}(?:-\d{4})?\b")

# Split delimiters when two first names share a last name on the right
_SHARED_LAST_DELIMS = re.compile(r"\s*(?:&|and|/|\+)\s*", re.IGNORECASE)

INSTRUCTION_RE = re.compile(r"\b(check|circle|select)\s+one\b", re.IGNORECASE)
INITIAL_RE = re.compile(r"\binitial\b", re.IGNORECASE)
FORM_HINT_TOKENS = {"PLEASE", "PRINT", "CLEARLY"}



def _looks_like_junk(s):
    s = s.strip()
    junkpat = re.compile(
    r'check\s*one|initial|annual|final|dates\s+covered|guardianship\s+of|please\s+fill\s+out|select\s+one|circle\s+one|filed\s+for\s+record|'
    r'hospital facility|medical facility|name\s|visit date|visit time|cause\s*no|tx\b|austin\b|\bzip\b|\d{5}(?:-\d{4})?\b|@|\d{3}[-/.\s]?\d{3}[-/.\s]?\d{4}|'
    r'both must be listed|must be listed|list both|if applicable|n/?a|none',
    re.I)
    return bool(junkpat.search(s))

def _looks_like_name(s: str) -> bool:
    s2 = s.strip()
    if not s2:
        return False
    if _looks_like_junk(s2):
        return False
    # Only letters, spaces, hyphens, apostrophes, periods
    if not re.match(r"^[A-Za-z][A-Za-z .'\-]+[A-Za-z.]$", s2):
        return False
    # 2-4 tokens, usually capitalized
    words = s2.split()
    if not (2 <= len(words) <= 4):
        return False
    # reject two-letter all-caps tokens (state codes) inside
    if any(len(w) == 2 and w.isupper() for w in words):
        return False
    return True


def _infer_shared_last_two_names(line: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Handle the frequent 'First & First Last' pattern on a single line.
    Returns (g1, g2) or (None, None) if not applicable.
    """
    lm = re.sub(r"\s+", " ", line).strip()
    if not _SHARED_LAST_DELIMS.search(lm):
        return (None, None)

    parts = _SHARED_LAST_DELIMS.split(lm)
    if len(parts) != 2:
        return (None, None)

    left, right = parts[0].strip(), parts[1].strip()
    # Right should contain at least First Last
    right_tokens = [t for t in right.split() if t and t[0].isalpha()]
    left_tokens  = [t for t in left.split() if t and t[0].isalpha()]
    if len(right_tokens) >= 2 and len(left_tokens) >= 1:
        last = right_tokens[-1]
        g1 = f"{left_tokens[-1]} {last}"
        g2 = f"{right_tokens[-2]} {last}"
        # Validate with name checks
        if _looks_like_name(g1) and _looks_like_name(g2):
            return (g1, g2)
    return (None, None)



def parse_arp_fields(text: str) -> dict | None:
    """
    Extract ward/guardian info from ARP PDFs (page 1).
    Leaves visitdate/visittime blank (not reliable on ARP).
    Uses robust multi-line capture for addresses and checkbox-only logic for liveswith.
    """
    t = normalize_ocr_confusions(text)
    t = clean_ocr_underscores(t)
    data = {h: "" for h in HEADERS}

    # --- Cause number ---
    cause = (
        grab(r'(?:[Cc]?ause\s*No\.?\s*[:#]?\s*)(C[\s\-]?1[\s\-]?PB[\s\-]?\d{2}[\s\-]?\d+)', t)
        or grab(r'\b(C[\s\-]?1[\s\-]?PB[\s\-]?\d{2}[\s\-]?\d+)\b', t)
    )
    if not cause:
        tail = grab(r'\b(\d{2}-\d{5,6})\b', t)
        if tail:
            cause = f"C-1-PB-{tail}"

    # NEW: very loose fallback for handwritten / weird OCR
    if not cause:
        cause = extract_causeno_loose(t)

    cause = normalize_causeno(cause) if cause else ""
    if not cause:
        return None

    data["causeno"] = cause

    # --- Ward name ---
    name_cands = extract_ward_name_candidates_from_arp(t)
    wf, wm, wl = choose_best_ward_name(name_cands)
    if wf and wl:
        data["wardfirst"] = wf
        data["wardmiddle"] = wm
        data["wardlast"] = wl

    # --- Visit date/time: intentionally blank for ARP ---
    data["visitdate"] = ""
    data["visittime"] = ""

    # --- Ward phone / address / dob ---
    data["wtele"] = normalize_phone(
        safe_after_label(t, r'(Ward\s*Phone|Phone)', "phone")
    )
    data["wdob"] = to_slashes(
        safe_after_label(t, r'(Ward\s*DOB|DOB|Date\s*of\s*Birth)', "date")
    )

    # --- Liveswith: checkbox-only logic (YES->Guardian; NO->blank; both->Guardian; unknown->blank) ---
    lw = parse_liveswith_guardian(t)
    data["liveswith"] = lw if lw is not None else ""

    # --- Addresses (ARP-specific label stitching, then fallback) ---
    # Peek the exact label lines the OCR produced for Ward
    ward_street_line = capture_labeled_value(t, WARD_ADDR_LABEL)
    ward_city_line   = capture_labeled_value(t, WARD_CITY_LABEL)
    # Capture final Ward address using ARP label stitcher
    data["waddress"] = capture_arp_address_by_labels(
        t, WARD_ADDR_LABEL, WARD_CITY_LABEL
    )

    # If still empty, fallback to scoped chunk near Ward section
    ward_chunk = _slice_between(
        t,
        r'(Ward\s*Information|Ward\s*Name|Ward\s*:\s*Name|1\.\s*WARD\b)',
        r'(Guardian\(s\)|Guardian\s*Information|2\.\s*GUARDIAN\(s\)\b)'
    )
    if not data["waddress"]:
        data["waddress"] = capture_address_after_label(ward_chunk, r'\bAddress\b', max_lines=5)

    # DEBUG: save what we saw vs. what we kept (Ward)
    try:
        cz = data.get("causeno", "").strip() or "unknown"
        save_debug(f"{cz}__WARD_label_lines", f"street_line: {ward_street_line}\ncity_line: {ward_city_line}")
        save_debug(f"{cz}__WARD_address_final", data["waddress"])
        # Optional context: the local Ward chunk
        save_debug(f"{cz}__WARD_chunk", ward_chunk[:1200])
    except Exception as _:
        pass

    # Guardian 1 — peek label lines and capture
    g1_street_line = capture_labeled_value(t, GUARD_ADDR_LABEL)
    g1_city_line   = capture_labeled_value(t, GUARD_CITY_LABEL)
    data["gaddress"] = capture_arp_address_by_labels(
        t, GUARD_ADDR_LABEL, GUARD_CITY_LABEL
    )

    # Fallback to a Guardian 1 chunk if needed
    g1_chunk = _slice_between(
        t,
        r'(Guardian\(s\)|Guardian\s*Information|Guardian\s*Name|2\.\s*GUARDIAN\(s\)\b)',
        r'(Guardian\s*2|Second\s*Guardian|G2\s*Information|Visit\s*Date|Visit\s*Time|Cause\s*No\.?)'
    )
    if not data["gaddress"]:
        data["gaddress"] = capture_address_after_label(g1_chunk, r'\bAddress\b', max_lines=5)
    
    # Clean up address contamination - remove guardian info from address fields
    if data.get("gaddress"):
        # Remove patterns like "2. GUARDIAN(s): Name(s) Matthew & Amy Cox" from addresses
        data["gaddress"] = re.sub(r'\d+\.\s*GUARDIAN\(s\)\s*:\s*Name\(s\)\s*[^,]+(?:,|$)', '', data["gaddress"], flags=re.I).strip()
        # Remove any remaining guardian section headers
        data["gaddress"] = re.sub(r'GUARDIAN\(s\)\s*:\s*Name\(s\)\s*[^,]+(?:,|$)', '', data["gaddress"], flags=re.I).strip()
        # Clean up any trailing commas or extra spaces
        data["gaddress"] = re.sub(r',\s*$', '', data["gaddress"]).strip()

    # DEBUG: save what we saw vs. what we kept (Guardian 1)
    try:
        cz = data.get("causeno", "").strip() or "unknown"
        save_debug(f"{cz}__G1_label_lines", f"street_line: {g1_street_line}\ncity_line: {g1_city_line}")
        save_debug(f"{cz}__G1_address_final", data["gaddress"])
        save_debug(f"{cz}__G1_chunk", g1_chunk[:1200])
    except Exception as _:
        pass

    # Guardian 2 (if present) — still via chunk, since many ARPs keep both guardians in one block
    g2_chunk = _slice_between(
        t,
        r'(Guardian\s*2|Second\s*Guardian|G2\s*Information)',
        r'(Visit\s*Date|Visit\s*Time|Cause\s*No\.?|$)'
    )
    data["g2 address"] = capture_address_after_label(g2_chunk, r'\bAddress\b', max_lines=5)

    # DEBUG: capture G2 result and context even if blank (helps confirm layout)
    try:
        cz = data.get("causeno", "").strip() or "unknown"
        if data.get("g2 address"):
            save_debug(f"{cz}__G2_address_final", data["g2 address"])
        if g2_chunk:
            save_debug(f"{cz}__G2_chunk", g2_chunk[:1200])
    except Exception as _:
        pass



    # --- Relationships ---
    rel   = safe_after_label(t, r'(Relationship\s*to\s*Ward|Relationship|Relation|Rel\.)', "any")
    g2rel = safe_after_label(t, r'(Second\s*Guardian\s*Relationship|Guardian\s*2\s*Relationship|G2\s*Relationship)', "any")
    data["Relationship"]   = sanitize_relationship(normalize_role(rel))
    data["g2Relationship"] = sanitize_relationship(normalize_role(g2rel))

    # --- ARP stamp: Filed date ---
    data["datesubmited"] = ""

    # Prefer robust clerk-stamp reader; if it fails, fall back to the old regex
    date_from_stamp = extract_arp_filed_date(t)
    if not date_from_stamp:
        stamp = grab(
            r'\b(Filed|Entered)\b\s*:?\s*([0-9]{1,2}[./-]\d{1,2}[./-]\d{2,4}|[A-Za-z]{3,9}\s+\d{1,2},\s+\d{4})',
            t
        )
        if stamp:
            mstamp = re.search(
                r'([0-9]{1,2}[./-]\d{1,2}[./-]\d{2,4}|[A-Za-z]{3,9}\s+\d{1,2},\s+\d{4})',
                stamp
            )
            date_from_stamp = mstamp.group(1) if mstamp else stamp

    data["DateARPfiled"] = normalize_month_text_date(to_slashes(date_from_stamp))
  
    # --- Guardian ARP name extraction (names only) ---
    gslice = _slice_guardian_section(t)

    # First try the new name line parser - be more aggressive about finding Guardian 2 patterns
    name_line = None
    for line in gslice:
        # Look for various name line patterns
        if (re.search(r'name(?:\(s\))?s?\s*(?:[:\-]|\s)', line, re.I) or
            re.search(r'name\s*\([0-9]+\)', line, re.I) or  # OCR error: "Name (0)" instead of "Name(s)"
            re.search(r'guardian\(s\)\s*:\s*name\(s\)', line, re.I) or
            re.search(r'2\.\s*guardian\(s\)\s*:\s*name\(s\)', line, re.I) or
            # Enhanced patterns for Guardian 2 detection - PRIORITY PATTERNS
            re.search(r'\b[A-Z][a-z]+\s+[A-Z][a-z]+\s+(?:and|&|\+|/|,)\s*[A-Z][a-z]+', line, re.I) or  # Full name separator name
            re.search(r'\b[A-Z][a-z]+\s+(?:and|&|\+|/|,)\s*[A-Z][a-z]+', line, re.I) or  # Name separator name
            re.search(r'\b(?:and|&|\+|/|,)\s*[A-Z][a-z]+', line, re.I)):  # Lines with separators + names
            name_line = line
            break
    
    g1, g2 = (None, None)
    if name_line:
        log(f"Found name line: {name_line}")
        g1, g2 = _extract_guardian_names_from_name_line(name_line)
        log(f"Name line extraction: g1='{g1}', g2='{g2}'")
    
    # Fallback to old method if name line parser didn't work OR if g2 is missing
    if not g1 and not g2:
        g1, g2 = _extract_guardian_names_from_lines(gslice)
    elif g1 and not g2:
        # Try to find g2 using the old method
        _, g2_fallback = _extract_guardian_names_from_lines(gslice)
        if g2_fallback:
            g2 = g2_fallback
    
    # Special case: if we found a single name in the name line but no Guardian2, 
    # and there's evidence of a second guardian in other fields, try to extract it
    if g1 and not g2:
        # Look for patterns like "Michael and Joslyn Mogonye" in the name line
        if name_line and ('and' in name_line.lower() or '&' in name_line):
            g1_new, g2_new = _extract_guardian_names_from_name_line(name_line)
            if g1_new and g2_new:
                g1, g2 = g1_new, g2_new
        
        # Enhanced Guardian 2 extraction from guardian slice
        if not g2:
            for line in gslice:
                # Look for various Guardian 2 patterns
                patterns = [
                    r'\band\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)',  # "and First Last"
                    r'\b&\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)',   # "& First Last"
                    r'\b/\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)',   # "/ First Last"
                    r'\b\+\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)',  # "+ First Last"
                    r'\b,\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)',   # ", First Last"
                ]
                
                for pattern in patterns:
                    if re.search(pattern, line, re.I):
                        match = re.search(pattern, line, re.I)
                        if match:
                            candidate = match.group(1).strip()
                            # Exclude common non-name words that might appear after separators
                            exclude_words = {'disability', 'services', 'department', 'aging', 'branch', 'commission', 'investigation', 'conducted', 'judicial', 'certification', 'subject', 'professional', 'guardian', 'program', 'reporting', 'year', 'convicted', 'felony', 'misdemeanor', 'traffic', 'offense', 'explain', 'resigning', 'successor', 'identified', 'reside', 'visited', 'ward'}
                            if _looks_like_human_name(candidate) and not any(word in candidate.lower() for word in exclude_words):
                                g2 = candidate
                                break
                if g2:
                    break
    
    # ENHANCED: Try to reconstruct incomplete guardian names
    # If g1 is just a first name, try to find the last name from other parts of the guardian section
    if g1 and len(g1.split()) == 1:  # g1 is just a first name
        # Look for full names in the guardian section that start with the same first name
        for line in gslice:
            # Look for patterns like "Matthew Cox", "Daniel Anthony Davidson", etc.
            words = line.split()
            for i, word in enumerate(words):
                if word == g1 and i + 1 < len(words):  # Found the first name, check if next word is a last name
                    potential_last = words[i + 1]
                    # Check if it looks like a last name (capitalized, not a common word)
                    if (re.match(r'^[A-Z][a-z]+$', potential_last) and 
                        potential_last.lower() not in {'and', 'or', 'the', 'of', 'in', 'on', 'at', 'to', 'for', 'with', 'by', 'age', 'dob', 'email', 'phone', 'address', 'relationship'}):
                        full_name = f"{g1} {potential_last}"
                        if _looks_like_human_name(full_name):
                            g1 = full_name
                            break
            if len(g1.split()) > 1:  # Found a last name, stop looking
                break
        
        # If still just first name, try to find last name from Guardian2 or other context
        if len(g1.split()) == 1 and g2 and len(g2.split()) >= 2:
            # Try to share the last name from Guardian2
            g2_words = g2.split()
            if len(g2_words) >= 2:
                shared_last = g2_words[-1]  # Last word of Guardian2
                full_g1 = f"{g1} {shared_last}"
                if _looks_like_human_name(full_g1):
                    g1 = full_g1

    # Special handling: if g1 is just a first name and g2 has full name, try to construct g1's full name
    if g1 and g2 and not _looks_like_human_name(g1) and _looks_like_human_name(g2):
        # g1 might be just "Michael", g2 is "Joslyn Mogonye"
        # Try to give g1 the same last name as g2
        g1_words = g1.split()
        g2_words = g2.split()
        if len(g1_words) == 1 and len(g2_words) >= 2:
            last_name = g2_words[-1]
            g1 = f"{g1_words[0]} {last_name}"
    
    if g1 and _looks_like_human_name(g1):
        data["guardian1"] = g1
    elif g1:
        pass  # g1 rejected by _looks_like_human_name
    
    if g2 and _looks_like_human_name(g2):
        data["Guardian2"] = g2
    elif g2:
        pass  # g2 rejected by _looks_like_human_name
    
    # CRITICAL: Don't let Guardian2 fallback overwrite Guardian1!
    # If we found Guardian1 from name line, preserve it even if Guardian2 fallback runs
    original_g1 = data.get("guardian1")
    
    # FALLBACK: If no guardian name found in the name line, look for names elsewhere in the guardian section
    if not data.get("guardian1") and not data.get("Guardian2"):
        # Look for any human names in the guardian section
        for line in gslice:
            # Skip lines that are clearly not names
            if re.search(r'(?:age|dob|email|phone|address|relationship|city|state|zip)', line, re.I):
                continue
            
            # Look for potential names (2-4 words, TitleCase)
            words = line.split()
            if 2 <= len(words) <= 4:
                potential_name = ' '.join(words)
                if _looks_like_human_name(potential_name):
                    if not data.get("guardian1"):
                        data["guardian1"] = potential_name
                        break
    
    # FALLBACK 2: If still no guardian name, look in the signature section
    if not data.get("guardian1") and not data.get("Guardian2"):
        # Look for signature section with guardian name
        # Pattern: look for lines before "I," or "the guardian of the person for"
        signature_section = _slice_between(t, r'I,\s*the\s+guardian|I,\s*$', r'Executed\s+on|Guardian\'s\s+signature', max_len=500)
        if signature_section:
            for line in signature_section.split('\n'):
                line = line.strip()
                if _looks_like_human_name(line) and len(line) >= 6:
                    if not data.get("guardian1"):
                        data["guardian1"] = line
                        break
        
        # FALLBACK 3: Look for guardian name in the lines before the signature
        if not data.get("guardian1") and not data.get("Guardian2"):
            # Look for lines that appear before "I," or "the guardian of the person for"
            lines = t.split('\n')
            for i, line in enumerate(lines):
                if re.search(r'I,\s*$|the\s+guardian\s+of\s+the\s+person\s+for', line, re.I):
                    # Check the previous few lines for guardian names
                    for j in range(max(0, i-3), i):
                        prev_line = lines[j].strip()
                        if _looks_like_human_name(prev_line) and len(prev_line) >= 6:
                            if not data.get("guardian1"):
                                data["guardian1"] = prev_line
                                break
                    break

    # --- Guardian contacts attach (split by separators and assign to G1/G2) ---
    try:
        # Extract field values from guardian section
        guardian_fields = {}
        for line in gslice:
            line_lower = line.lower()
            
            # Check for DOB in this line
            if 'dob' in line_lower:
                dob_match = re.search(r'dob(?:\(s\))?\s*[:\-]?\s*(.+)', line, re.I)
                if dob_match:
                    dob_value = dob_match.group(1).strip()
                    
                    # First, clean the DOB value to remove non-date parts
                    # Remove common prefixes like "(6)", "(s)", etc.
                    cleaned_dob_value = re.sub(r'^\([^)]*\)\s*', '', dob_value)
                    # Extract the date part(s) - look for multiple dates separated by spaces
                    # Pattern like "8/16/65 4/15/65" or "8/16/65"
                    date_part = re.search(r'^([\d/\s]+)', cleaned_dob_value)
                    if date_part:
                        clean_dob = date_part.group(1).strip()
                        
                        # Check patterns in order of specificity (most specific first)
                        # Special case: complex line like "11/13/70/3/21/23" - extract two dates
                        if re.search(r'\d{1,2}/\d{1,2}/\d{2,4}/\d{1,2}/\d{1,2}/\d{2,4}', clean_dob):
                            # Extract two dates from pattern like "11/13/70/3/21/23"
                            dates = re.findall(r'\d{1,2}/\d{1,2}/\d{2,4}', clean_dob)
                            if len(dates) >= 2:
                                guardian_fields['dob'] = f"{dates[0]} / {dates[1]}"
                        # Check if it contains multiple dates with clear separators
                        elif re.search(r'\d{1,2}/\d{1,2}/\d{2,4}\s+and\s+\d{1,2}/\d{1,2}/\d{2,4}', clean_dob, re.I):
                            guardian_fields['dob'] = clean_dob
                        elif re.search(r'\d{1,2}/\d{1,2}/\d{2,4}\s*[/&]\s*\d{1,2}/\d{1,2}/\d{2,4}', clean_dob):
                            guardian_fields['dob'] = clean_dob
                        # Check for two dates separated by space (like "8/16/65 4/15/65")
                        elif re.search(r'\d{1,2}/\d{1,2}/\d{2,4}\s+\d{1,2}/\d{1,2}/\d{2,4}', clean_dob):
                            dates = re.findall(r'\d{1,2}/\d{1,2}/\d{2,4}', clean_dob)
                            if len(dates) >= 2:
                                guardian_fields['dob'] = f"{dates[0]} / {dates[1]}"
                        # If it's a single date, capture it for Guardian1 only
                        elif re.search(r'\d{1,2}/\d{1,2}/\d{2,4}', clean_dob) and 'dob' not in guardian_fields:
                            guardian_fields['dob_single'] = clean_dob
            
            # Check for Email in this line (can be in same line as DOB)
            if 'email' in line_lower or re.search(r'[^/\s]+@[^/\s]+\.[^/\s]+', line):
                # First try the specific pattern: email followed by email address
                email_match = re.search(r'email\s*[:\-]?\s*([^/\s]+@[^/\s]+\.[^/\s]+)', line, re.I)
                if email_match:
                    email_value = email_match.group(1).strip()
                    # Check if it contains multiple emails (for splitting)
                    if re.search(r'[/&]|and|\s{2,}', email_value):
                        guardian_fields['email'] = email_value
                    else:
                        guardian_fields['email_single'] = email_value
                else:
                    # Fallback: look for any email pattern in the line
                    emails = re.findall(r'[^/\s]+@[^/\s]+\.[^/\s]+', line)
                    if emails:
                        email_value = emails[0]  # Take the first email found
                        guardian_fields['email_single'] = email_value
                    else:
                        # Try to find incomplete emails and complete them
                        incomplete_emails = re.findall(r'[^/\s]+@[^/\s]+(?:gmail|yahoo|hotmail|outlook|aol)', line, re.I)
                        if incomplete_emails:
                            email_value = incomplete_emails[0]
                            # Complete common domains
                            if email_value.lower().endswith('gmail'):
                                email_value += '.com'
                            elif email_value.lower().endswith(('yahoo', 'hotmail', 'outlook', 'aol')):
                                email_value += '.com'
                            guardian_fields['email_single'] = email_value
                        else:
                            # Try to find very incomplete emails like "Wendy immerson@gmail"
                            very_incomplete = re.findall(r'[A-Za-z]+\s+[A-Za-z]+@[A-Za-z]+', line)
                            if very_incomplete:
                                email_value = very_incomplete[0]
                                # Fix spacing and complete domain
                                email_value = email_value.replace(' ', '')  # Remove space
                                if email_value.lower().endswith('gmail'):
                                    email_value += '.com'
                                guardian_fields['email_single'] = email_value
            
            # Check for Phone in this line (can be in same line as other data)
            if 'phone' in line_lower:
                # Look for phone patterns in the line
                phones = re.findall(r'\(?\d{3}\)?[-/.\s]?\d{3}[-/.\s]?\d{4}', line)
                if phones:
                    phone_value = ' / '.join(phones)
                    if len(phones) >= 2:
                        guardian_fields['phone'] = phone_value
                    else:
                        guardian_fields['phone_single'] = phone_value
            
            # Check for Relationship in this line (can be in same line as other data)
            if 'relationship' in line_lower:
                rel_match = re.search(r'relationship\s*[:\-]?\s*(.+)', line, re.I)
                if rel_match:
                    rel_value = rel_match.group(1).strip()
                    # Check if it contains multiple relationships (for splitting)
                    if re.search(r'[/&]|and|\s{2,}', rel_value):
                        guardian_fields['relationship'] = rel_value
                    else:
                        guardian_fields['relationship_single'] = rel_value

        # Split fields by separators and assign to G1/G2
        if 'dob' in guardian_fields:
            g1_dob, g2_dob = _split_guardian_field_by_separators(guardian_fields['dob'])
            if g1_dob and not data.get("gdob"): 
                data["gdob"] = _clean_dob(g1_dob) or g1_dob
            if g2_dob and not data.get("g2dob"): 
                data["g2dob"] = _clean_dob(g2_dob) or g2_dob
        elif 'dob_single' in guardian_fields:
            # Single DOB - assign to Guardian1 only
            single_dob = _clean_dob(guardian_fields['dob_single']) or guardian_fields['dob_single']
            if single_dob and not data.get("gdob"):
                data["gdob"] = single_dob

        if 'email' in guardian_fields:
            g1_email, g2_email = _split_guardian_field_by_separators(guardian_fields['email'])
            if g1_email and not data.get("gemail"): 
                data["gemail"] = g1_email
            if g2_email and not data.get("g2eamil"): 
                data["g2eamil"] = g2_email
        elif 'email_single' in guardian_fields:
            # Single email - assign to Guardian1 only
            single_email = guardian_fields['email_single']
            if single_email and not data.get("gemail"):
                data["gemail"] = single_email

        if 'phone' in guardian_fields:
            g1_phone, g2_phone = _split_guardian_field_by_separators(guardian_fields['phone'])
            if g1_phone and not data.get("gtele"): 
                try: data["gtele"] = normalize_phone(g1_phone)
                except Exception: data["gtele"] = g1_phone
            if g2_phone and not data.get("g2tele"): 
                try: data["g2tele"] = normalize_phone(g2_phone)
                except Exception: data["g2tele"] = g2_phone
        elif 'phone_single' in guardian_fields:
            # Single phone - assign to Guardian1 only
            single_phone = guardian_fields['phone_single']
            if single_phone and not data.get("gtele"):
                try: data["gtele"] = normalize_phone(single_phone)
                except Exception: data["gtele"] = single_phone

        if 'relationship' in guardian_fields:
            g1_rel, g2_rel = _split_guardian_field_by_separators(guardian_fields['relationship'])
            if g1_rel and not data.get("Relationship"): 
                data["Relationship"] = sanitize_relationship(normalize_role(g1_rel))
            if g2_rel and not data.get("g2Relationship"): 
                data["g2Relationship"] = sanitize_relationship(normalize_role(g2_rel))
        elif 'relationship_single' in guardian_fields:
            # Single relationship - assign to Guardian1 only
            single_rel = guardian_fields['relationship_single']
            if single_rel and not data.get("Relationship"):
                data["Relationship"] = sanitize_relationship(normalize_role(single_rel))

        
        # CRITICAL FIX: Restore Guardian1 if it got lost during field splitting
        if original_g1 and not data.get("guardian1"):
            data["guardian1"] = original_g1
    except Exception as e:
        pass

    # --- Final cleanup: drop addressy/junk survivors like 'Zip Austin' ---
    def _is_addressy_or_junk(val: str) -> bool:
        v = (val or "").strip()
        if not v:
            return False
        # Don't flag names that look like human names - check this FIRST
        if _looks_like_human_name(v):
            return False
        try:
            return _looks_like_junk(v)  # preferred if helper exists
        except NameError:
            pass
        if re.search(r"(?:Address|New\s+Address|Zip|City|State|TX|Texas)\b", v, re.IGNORECASE):
            return True
        if re.search(r"\d{5}(?:-\d{4})?\b", v):    # zip
            return True
        if re.search(r"\(?\d{3}\)?[-/.\s]?\d{3}[-/.\s]?\d{4}", v):  # phone
            return True
        if re.search(r"[^@\s]+@[^@\s]+\.[^@\s]+", v):  # email
            return True
        return False

    for key in ("guardian1", "Guardian2"):
        if key in data and isinstance(data[key], str) and _is_addressy_or_junk(data[key]):
            data[key] = ""

    # Clamp the ARP filed date year if present
    if data.get("DateARPfiled"):
        fixed = _clamp_year_mdY(data["DateARPfiled"])
        if fixed:
            data["DateARPfiled"] = fixed
        else:
            data["DateARPfiled"] = ""

    return data


# =========================
#  Helper (TOP-LEVEL)
# =========================
def _clamp_year_mdY(s: str) -> str | None:
    m = re.match(r"^\s*(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\s*$", s or "")
    if not m:
        return None
    mm, dd, yy = m.groups()
    mm, dd = int(mm), int(dd)
    yy = int(yy)
    if yy < 100:
        yy = 2000 + yy if yy < 50 else 1900 + yy
    if not (1900 <= yy <= 2100):
        return None
    return f"{mm:02d}/{dd:02d}/{yy}"


def _clamp_future_year_to_current(s: str) -> str:
    """
    If date string MM/DD/YYYY has a year > current year, clamp to current year.
    Otherwise return unchanged. Returns original input on parse failure.
    """
    try:
        m = re.match(r"^\s*(\d{1,2})[/-](\d{1,2})[/-](\d{4})\s*$", s or "")
        if not m:
            return s or ""
        mm, dd, yyyy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        cur = datetime.now().year
        if yyyy > cur:
            yyyy = cur
        return f"{mm:02d}/{dd:02d}/{yyyy}"
    except Exception:
        return s or ""


# === Guardian contacts helpers (TOP-LEVEL) ===
_EMAIL_RE = re.compile(r'[^@\s]+@[^@\s]+\.[^@\s]+')
_PHONE_RE = re.compile(r'\(?\d{3}\)?[-/.\s]?\d{3}[-/.\s]?\d{4}')
_DOB_TOKEN_RE = re.compile(r'\b(dob|date\s*of\s*birth)\b', re.IGNORECASE)
_DATE_MDY_RE = re.compile(r'\b(0?[1-9]|1[0-2])[\/\-\.\_](0?[1-9]|[12]\d|3[01])[\/\-\.\_](\d{2,4})\b')

def _collect_guardian_contacts(lines: list[str]) -> dict[str, list[tuple[str,int]]]:
    emails, phones, dobs = [], [], []
    for i, ln in enumerate(lines):
        for e in _EMAIL_RE.findall(ln): emails.append((e, i))
        for p in _PHONE_RE.findall(ln): phones.append((p, i))
        # DOBs only if a DOB token is on this or an adjacent line
        if _DOB_TOKEN_RE.search(ln):
            for m in _DATE_MDY_RE.finditer(ln): dobs.append(('/'.join(m.groups()), i))
        else:
            if (i > 0 and _DOB_TOKEN_RE.search(lines[i-1])) or (i+1 < len(lines) and _DOB_TOKEN_RE.search(lines[i+1])):
                for m in _DATE_MDY_RE.finditer(ln): dobs.append(('/'.join(m.groups()), i))
    return {"emails": emails, "phones": phones, "dobs": dobs}

def _choose_nearest(name_idx: int, items: list[tuple[str,int]], max_window: int = 8) -> str | None:
    best, best_score = None, 10**9
    for val, idx in items:
        d = idx - name_idx
        if abs(d) > max_window:
            continue
        # Penalize obviously invalid NANP exchanges like 000 or 09x
        penalty = 0
        if re.match(r'^\(?\d{3}\)?[-/.\s]?0\d\d', val):
            penalty += 2
        score = (d if d >= 0 else abs(d) + 3) + penalty
        if score < best_score:
            best_score, best = score, val
    return best

def _clean_dob(val: str | None) -> str | None:
    if not val: return None
    m = _DATE_MDY_RE.match(val.strip())
    if not m: return None
    mm, dd, yy = map(int, m.groups())
    if yy < 100:
        yy = 2000 + yy if yy < 50 else 1900 + yy
    if not (1900 <= yy <= 2100): return None
    return f"{mm:02d}/{dd:02d}/{yy}"

import re

# Ignore these common instruction/boilerplate lines after the Guardian header
_INSTR_LINE_RE = re.compile(
    r'(check\s*one|initial|annual|final|dates\s+covered|guardianship\s+of|'
    r'please\s+fill\s+out|select\s+one|circle\s+one|filed\s+for\s+record|'
    r'relationship\s*to\s*ward|guardian\s*1|guardian\s*2|guardian\s*information|guardian\s*name)',
    re.IGNORECASE
)

# Street/address words (to block "Sage Court", "Largo Cove", etc.)
_STREET_WORDS_RE = re.compile(
    r'\b(st|street|rd|road|dr|drive|ln|lane|ct|court|ave|avenue|blvd|boulevard|'
    r'pkwy|parkway|ter|terrace|pl|place|way|loop|trail|pass|cove|cir|circle|'
    r'hwy|highway)\b',
    re.IGNORECASE
)
_NEVER_NAME_RE = re.compile(
    r'(?:'
    r'check\s*one|initial|annual|final|dates?\s+covered|filed\s*for\s*record|'
    r'hospital\s*facility|medical\s*facility|visit\s*date|visit\s*time|cause\s*no|'
    r'address|address\s*\(no\s*p\.?\s*o\.?\s*box\)|phone|email|age(?:\(s\))?|dob(?:\(s\))?|'
    r'^\s*name\b|guardian\s*\(s\)|guardian\s*information|'
    r'relationship\s*(?:to\s*the\s*)?ward|guardian\s*of|'
    r'parents?|mother|father|grand(?:mother|father)|spouse|'
    r'both\s*must\s*be\s*listed|'
    r'\btx\b|\baustin\b|'
    r'\bzip\b|\bcity\b|\bstate\b|c\s*ity|s\s*tate|z\s*ip|'
    r'c\s*rty\s*l\s*state\s*l\s*zip|city\s*/\s*state\s*/\s*zip|'
    r'\d{5}(?:-\d{4})?\b'
    r')',
    re.IGNORECASE
)


def _strip_joiners(s: str) -> str:
    # remove leading joiners like "&", "+", "and"
    return re.sub(r'^\s*(?:&|\+|and)\s+', '', s, flags=re.IGNORECASE).strip()

# Back-compat alias (in case other code still calls your old name)
def _strip_leading_conjunction(s: str) -> str:
    return _strip_joiners(s)

def _has_street_word(s: str) -> bool:
    toks = re.findall(r"[A-Za-z']+", s.lower())
    return any(tok in _STREET_WORDS for tok in toks)


def _slice_guardian_section(t: str) -> list[str]:
    # Anchor: "2. GUARDIAN...", "GUARDIAN(S)", or "Guardian Information/Name(s)"
    anchor = re.search(
        r'(?:\b2\.\s*guardian[^\n:]*:?|guardian\(s\)[^\n:]*:?|guardian\s*(?:information|name\(s\)|names)[^\n:]*:?)',
        t, re.IGNORECASE
    )
    if not anchor:
        return []
    start = anchor.end()

    # End: next numbered section (3.), Visit Date/Time, or Cause No
    m_end = re.search(r'\n\s*\d+\.\s|visit\s*date|visit\s*time|cause\s*no', t[start:], re.IGNORECASE)
    stop = start + (m_end.start() if m_end else len(t) - start)
    raw_lines = [ln.strip() for ln in t[start:stop].splitlines()]
    raw_lines = [ln for ln in raw_lines if ln]  # drop empties

    cleaned: list[str] = []
    for ln in raw_lines:
        if not ln:
            continue
        if _INSTR_LINE_RE.search(ln):
            # Keep lines that likely already contain a human name after the label
            # e.g., "Name(s) Magdalena Wolk"
            if re.search(r"[A-Z][a-z]+\s+[A-Z][a-z]+", ln):
                cleaned.append(ln)
            else:
                continue
        else:
            cleaned.append(ln)
    # allow a larger slice (some scans put names far down)
    return cleaned[:360]

def _clean_extracted_name(n: str | None) -> str | None:
    if not n:
        return None
    n2 = re.sub(r'^\s*name[:\s]+', '', n, flags=re.IGNORECASE).strip()
    n2 = _strip_joiners(n2)
    if _NEVER_NAME_RE.search(n2) or _has_street_word(n2):
        return None
    return n2

def _strip_leading_conjunction(s: str) -> str:
    # turn "and Joslyn Mogonye" / "& Joslyn Mogonye" into "Joslyn Mogonye"
    return re.sub(r'^\s*(?:and|&)\s+', '', s, flags=re.IGNORECASE).strip()

def _looks_like_human_name(s: str) -> bool:
    if not s:
        return False
    s = s.strip()
    if _NEVER_NAME_RE.search(s) or _STREET_WORDS_RE.search(s):
        return False
    if re.search(r"[^A-Za-z .'\-,\s]", s):  # allow commas and spaces
        return False
    if any(ch.isdigit() for ch in s):
        return False

    words = s.split()
    if not (1 <= len(words) <= 4):  # Allow single names for guardians
        return False

    # disallow role-y tokens as a "name" (but allow company names and professional titles)
    if words[0].lower() in ("parents", "parent", "mother", "father", "guardian", "guardians"):
        return False
    
    # Allow company names and professional titles for Guardian 2 extraction
    # This helps with patterns like "Family Eldercare, Inc-Robert Lontkowski, Care Manager"
    if any(word.lower() in ("inc", "llc", "corp", "company", "manager", "care") for word in words):
        return True

    # Count TitleCase words and also "Jr.", "Sr.", "III", etc.
    # Allow all-caps names (like "MEENU JAIN")
    titled = sum(
        1 for w in words
        if (re.match(r"[A-Z][a-z]+(?:[.\-'][A-Za-z]+)?$", w) is not None or
            re.match(r"^[A-Z]{2,}$", w) is not None or  # All caps names like "JAIN", "MEENU"
            re.match(r"^(Jr\.?|Sr\.?|III|IV|V)$", w, re.I) is not None)
    )
    if len(words) == 1 and titled < 1:  # Single name needs at least 1 TitleCase or all-caps
        return False
    elif len(words) > 1 and titled < 2:  # Multi-word names need at least 2 TitleCase or all-caps
        return False

    if _STREET_WORDS_RE.search(words[-1]):  # e.g., "... Court", "... Cove"
        return False

    return True

_ALLOWED_REL = {"Father", "Mother", "Father/Mother", "Parent", "Son", "Daughter", "Public Guardian"}

def sanitize_relationship(val: str) -> str:
    v = (val or "").strip()
    v = re.sub(r'[:;,.]+$', '', v)
    v = re.sub(r'\s+', ' ', v)
    v = normalize_role(v)  # you already have this
    if not v:
        return ""
    if v in _ALLOWED_REL:
        return v
    # discard long sentences or obviously wrong content
    if len(v) > 30:
        return ""
    if re.search(r'\d|visit|convict|report|year', v, re.I):
        return ""
    return v

def _split_name_line(line: str) -> list[str]:
    """
    Given a 'Name(s) ...' line (already roughly trimmed), return a list of 0–2 cleaned names.
    - Handles: 'Wendelynn & James Jimmerson' (shared last), 'First Last and Second Last',
               'Michael and Joslyn Mogonye', and single full-name lines.
    - Never returns None values; always a list[str].
    """
    # 0) Guard against None or non-strings
    if not isinstance(line, str):
        return []

    s = line.strip()
    if not s:
        return []

    # 1) Strip common 'Name' labels so we only see the actual names
    s = re.sub(r'^\s*name(?:\(s\))?s?\s*[:\-]?\s*', '', s, flags=re.I)

    # 2) Normalize weird OCR separators and whitespace
    s = s.replace('|', ' ').replace('/', ' ').replace('\\', ' ')
    s = re.sub(r'\s{2,}', ' ', s).strip()

    # 3) Early exit if the whole thing looks like a non-name line
    if _NEVER_NAME_RE.search(s) and not re.search(r'[A-Z][a-z]+\s+[A-Z][a-z]+', s):
        # if the line contains an embedded full name despite label noise, keep parsing
        return []

    def _clean(n: str) -> str | None:
        if not isinstance(n, str):
            return None
        n = n.strip(" ,.;:-")
        if not n:
            return None
        # Drop trailing non-name garbage tokens that sometimes bleed into the line
        n = re.sub(r'\b(?:age|dob|email|phone|relationship)\b.*$', '', n, flags=re.I).strip(" ,.;:-")
        if not n or _NEVER_NAME_RE.search(n):
            return None
        # Title-case words that look like names; keep all-caps acronyms
        parts = []
        for w in n.split():
            parts.append(w if w.isupper() or len(w) <= 1 else w.capitalize())
        out = " ".join(parts).strip()
        return out or None

    out: list[str] = []

    # 4) Pattern A: shared last name: "First1 [&|and] First2 Last"
    m = re.search(
        r"^\s*([A-Za-z.'\- ]+?)\s*(?:&|and)\s*([A-Za-z.'\- ]+?)\s+([A-Za-z][A-Za-z.'\- ]+)\s*$",
        s, flags=re.I
    )
    if m:
        first1, first2, last = m.group(1), m.group(2), m.group(3)
        g1 = _clean(f"{first1} {last}")
        g2 = _clean(f"{first2} {last}")
        if g1: out.append(g1)
        if g2: out.append(g2)
        return out

    # 5) Pattern B: both have their own last names: "First Last [&|and] First Last"
    m = re.search(
        r"^\s*([A-Za-z.'\-]+(?:\s+[A-Za-z.'\-]+)+)\s*(?:&|and)\s*([A-Za-z.'\-]+(?:\s+[A-Za-z.'\-]+)+)\s*$",
        s, flags=re.I
    )
    if m:
        g1 = _clean(m.group(1))
        g2 = _clean(m.group(2))
        if g1: out.append(g1)
        if g2: out.append(g2)
        return out

    # 6) Pattern C: a single full name only
    m = re.search(r"^\s*([A-Za-z.'\-]+(?:\s+[A-Za-z.'\-]+)+)\s*$", s)
    if m:
        g1 = _clean(m.group(1))
        if g1: out.append(g1)
        return out

    return out

def _extract_guardian_names_from_lines(lines: list[str]) -> tuple[str | None, str | None]:
    """
    Build candidate names from cleaned lines:
      - strip leading 'and ' / '& '
      - drop address/city/state/zip/relationship lines early
      - strip 'Name ' label
      - accept only human-looking names
      - repair 'First & First Last' and stacked 'First First' + 'Last'
    """
    cleaned_lines: list[str] = []
    for line in lines:
        ln = _strip_leading_conjunction(line.strip())
        if not ln:
            continue
        if _NEVER_NAME_RE.search(ln) or _STREET_WORDS_RE.search(ln):
            continue
        ln = re.sub(r'^\s*name[:\s]+', '', ln, flags=re.IGNORECASE).strip()
        if ln and not _NEVER_NAME_RE.search(ln) and not _STREET_WORDS_RE.search(ln):
            cleaned_lines.append(ln)

    raw_parts: list[str] = []
    for ln in cleaned_lines:
        raw_parts.extend(_split_name_line(ln))
    raw_parts = [_strip_leading_conjunction(p) for p in raw_parts if p.strip()]

    candidates: list[str] = []
    for p in raw_parts:
        if _looks_like_human_name(p):
            candidates.append(p)

    # Repair pattern: one full name + one single first → borrow last
    if len(candidates) < 2:
        full_with_last = [p for p in raw_parts if re.match(r"^[A-Z][a-z]+ [A-Z][a-z]+(?: [A-Z][a-z]+)?$", p)]
        single_firsts  = [p for p in raw_parts if re.match(r"^[A-Z][a-z]+$", p)]
        if full_with_last and single_firsts:
            last = full_with_last[0].split()[-1]
            repaired = f"{single_firsts[0]} {last}"
            if _looks_like_human_name(repaired):
                idx_full   = raw_parts.index(full_with_last[0])
                idx_single = raw_parts.index(single_firsts[0])
                candidates = [repaired, full_with_last[0]] if idx_single < idx_full else [full_with_last[0], repaired]

    # Additional repair: two first names on line1, last name appears on line3
    if len(candidates) < 2 and len(cleaned_lines) >= 3:
        firsts = [w for w in cleaned_lines[0].split() if re.match(r"^[A-Z][a-z]+$", w)]
        last_tokens = [w for w in cleaned_lines[2].split() if re.match(r"^[A-Z][a-z]+$", w)]
        if len(firsts) == 2 and len(last_tokens) == 1:
            n1 = f"{firsts[0]} {last_tokens[0]}"
            n2 = f"{firsts[1]} {last_tokens[0]}"
            if _looks_like_human_name(n1) and _looks_like_human_name(n2):
                candidates = [n1, n2]

    # Stacked two firsts on line1, last on line2
    if len(candidates) < 2 and len(cleaned_lines) >= 2:
        firsts = [w for w in cleaned_lines[0].split() if re.match(r"^[A-Z][a-z]+$", w)]
        last_tokens = [w for w in cleaned_lines[1].split() if re.match(r"^[A-Z][a-z]+$", w)]
        if len(firsts) == 2 and len(last_tokens) == 1:
            n1 = f"{firsts[0]} {last_tokens[0]}"
            n2 = f"{firsts[1]} {last_tokens[0]}"
            if _looks_like_human_name(n1) and _looks_like_human_name(n2):
                candidates = [n1, n2]

    seen, out = set(), []
    for n in candidates:
        if n not in seen:
            seen.add(n)
            out.append(n)
        if len(out) == 2:
            break

    while len(out) < 2:
        out.append(None)
    return (out[0], out[1])


def _split_guardian_field_by_separators(value: str) -> tuple[str, str]:
    """
    Split a guardian field value by common separators and return (guardian1_value, guardian2_value).
    Only splits when there are clear separators between two distinct values.
    Special handling for dates: won't split on slashes within date patterns.
    """
    if not value:
        return ("", "")
    
    # Special case for dates: if value contains date patterns, use smart splitting
    # Pattern like "11/13/70 / 3/21/23" or "11/13/70/3/21/23"
    # BUT only if the value is primarily dates, not names with dates mixed in
    if re.search(r'\d{1,2}/\d{1,2}/\d{2,4}', str(value)):
        # Check if this is primarily a date field (like DOB) vs a name field with dates mixed in
        date_count = len(re.findall(r'\d{1,2}/\d{1,2}/\d{2,4}', str(value)))
        name_words = len([w for w in str(value).split() if re.match(r'^[A-Z][a-z]+$', w)])
        
        # Only treat as date field if there are more dates than name words
        if date_count > name_words:
            dates = re.findall(r'\d{1,2}/\d{1,2}/\d{2,4}', str(value))
            if len(dates) >= 2:
                return (dates[0], dates[1])
            elif len(dates) == 1:
                return (dates[0], "")
    
    # Special case for phone numbers: if value contains phone patterns, use smart splitting
    # Pattern like "512-094-6202 / 512-771-1695"
    if re.search(r'\(?\d{3}\)?[-/.\s]?\d{3}[-/.\s]?\d{4}', str(value)):
        # Extract all complete phone numbers
        phones = re.findall(r'\(?\d{3}\)?[-/.\s]?\d{3}[-/.\s]?\d{4}', str(value))
        if len(phones) >= 2:
            return (phones[0], phones[1])
        elif len(phones) == 1:
            return (phones[0], "")
    
    # Look for clear separators that indicate two distinct values
    # Pattern: value1 SEPARATOR value2 (where separator is &, and, +, or, /, @, or multiple spaces)
    # Note: '/' is only used as separator if not part of a date/phone pattern
    separators = [
        r'\s*,\s*',           # comma (highest priority for names)
        r'\s*&\s*',           # ampersand  
        r'\s+and\s+',         # word "and"
        r'\s*\+\s*',          # plus sign
        r'\s+or\s+',          # word "or"
        r'\s*/\s*',           # forward slash (but only if not in date/phone context)
        r'\s*@\s*',           # at symbol (NEW - handles patterns like "x@x")
        r'\s{2,}',            # 2+ spaces (NEW - more flexible than 3+ spaces)
        r'\s{3,}',            # 3+ spaces (keep existing for backward compatibility)
        # Additional patterns for Guardian 2 extraction
        r'\s*-\s*',           # hyphen/dash (for patterns like "Company-Person")
        r'\s*,\s*and\s+',     # comma + and (for "Name, and Name")
        r'\s*,\s*&\s*',       # comma + ampersand (for "Name, & Name")
    ]
    
    for sep_pattern in separators:
        if re.search(sep_pattern, str(value), re.I):
            parts = re.split(sep_pattern, str(value), flags=re.I)
            parts = [p.strip() for p in parts if p.strip()]
            if len(parts) >= 2:
                return (parts[0], parts[1])
    
    # If no clear separator found, return as single value
    return (str(value).strip(), "")


def _extract_guardian_names_from_name_line(name_line: str) -> tuple[str | None, str | None]:
    """
    Extract Guardian1 and Guardian2 names from a "Name(s) ..." line.
    Handles various formats like "Michael and Joslyn Mogonye", "Michael & Joslyn Mogonye", etc.
    """
    if not name_line:
        return (None, None)
    
    # Strip the "Name(s)" label - handle OCR errors like "Name (0)"
    clean_line = re.sub(r'^\s*name(?:\(s\))?s?\s*[:\-]?\s*', '', name_line, flags=re.I)
    clean_line = re.sub(r'^\s*name\s*\([0-9]+\)\s*', '', clean_line, flags=re.I)  # Handle "Name (0)"
    clean_line = clean_line.strip()
    
    if not clean_line:
        return (None, None)
    
    # Clean up common OCR issues and extra content
    # Remove email addresses, ages, DOBs that get mixed into name lines
    clean_line = re.sub(r'\s+email:\s*[^\s@]+@[^\s]+\s*', ' ', clean_line, flags=re.I)  # Full email
    clean_line = re.sub(r'\s+[^\s]+@[^\s]+\s*', ' ', clean_line, flags=re.I)  # Any email pattern
    clean_line = re.sub(r'\s+email:\s*[^\s]+\s*', ' ', clean_line, flags=re.I)  # Incomplete email like "email: brefaye"
    clean_line = re.sub(r'\s+[^\s]*\.com\s*', ' ', clean_line, flags=re.I)  # Remove domain names like "hotmail.com"
    clean_line = re.sub(r'\s+and\s+Age\s+\d+', '', clean_line, flags=re.I)
    clean_line = re.sub(r'\s+DOB:\d{1,2}/\d{1,2}/\d{2,4}', '', clean_line, flags=re.I)
    clean_line = re.sub(r'\s+\d{1,2}/\d{1,2}/\d{2,4}', '', clean_line, flags=re.I)
    clean_line = re.sub(r'\s+Age\s+\d+', '', clean_line, flags=re.I)
    clean_line = re.sub(r'\s+\d+\s*', ' ', clean_line)  # Remove standalone numbers
    clean_line = re.sub(r'^\([0-9]+\)\s*', '', clean_line)  # Remove leading OCR errors like "(0)"
    
    # Enhanced cleaning for Guardian 2 patterns
    # Handle company names with person names (like "Family Eldercare, Inc-Robert Lontkowski, Care Manager")
    # Keep the company part and person name, but remove titles
    clean_line = re.sub(r',\s*Care\s*Manager\s*$', '', clean_line, flags=re.I)
    clean_line = re.sub(r',\s*Manager\s*$', '', clean_line, flags=re.I)
    clean_line = re.sub(r',\s*Inc\.?\s*$', '', clean_line, flags=re.I)
    
    clean_line = re.sub(r'\s+', ' ', clean_line).strip()  # Normalize whitespace
    
    # Try to split by common separators
    log(f"Processing name line: '{clean_line}'")
    g1, g2 = _split_guardian_field_by_separators(clean_line)
    log(f"After separator split: g1='{g1}', g2='{g2}'")
    
    # Special case: handle incomplete names like "Michael andli" 
    if g1 and not g2 and not _looks_like_human_name(g1):
        # If g1 doesn't look like a complete name, try to extract just the first name
        words = g1.split()
        if len(words) >= 1 and re.match(r'^[A-Z][a-z]+$', words[0]):
            # Return just the first name as g1, we'll look for g2 elsewhere
            return (words[0], None)
    
    # If we got two parts, try to construct full names
    if g1 and g2:
        # Check if both parts look like names
        if _looks_like_human_name(g1) and _looks_like_human_name(g2):
            # But first, check if g1 is just a first name and g2 has a last name
            # This handles cases like "Michael and Joslyn Mogonye" where we want "Michael Mogonye"
            g1_words = g1.split()
            g2_words = g2.split()
            
            if len(g1_words) == 1 and len(g2_words) >= 2:
                # g1 is first name only, g2 has last name - try to share the last name
                last_name = g2_words[-1]
                full_g1 = f"{g1_words[0]} {last_name}"
                if _looks_like_human_name(full_g1):
                    return (full_g1, g2)
            
            # If no shared last name logic applies, return as-is
            return (g1, g2)
        
        # If one part is just a first name and the other has a last name, try to share the last name
        g1_words = g1.split()
        g2_words = g2.split()
        
        if len(g1_words) == 1 and len(g2_words) >= 2:  # g1 is first name only, g2 has last name
            last_name = g2_words[-1]
            full_g1 = f"{g1_words[0]} {last_name}"
            if _looks_like_human_name(full_g1) and _looks_like_human_name(g2):
                return (full_g1, g2)
        
        if len(g2_words) == 1 and len(g1_words) >= 2:  # g2 is first name only, g1 has last name
            last_name = g1_words[-1]
            full_g2 = f"{g2_words[0]} {last_name}"
            if _looks_like_human_name(g1) and _looks_like_human_name(full_g2):
                return (g1, full_g2)
        
        # Special case: if both are single words, they might be first names sharing a last name
        # This handles cases like "Michael and Joslyn Mogonye" where the last name is on the end
        if len(g1_words) == 1 and len(g2_words) == 1:
            # Look for a pattern like "Michael and Joslyn Mogonye" in the original clean_line
            shared_last_match = re.search(r'(\w+)\s+(?:and|&)\s+(\w+)\s+(\w+)$', clean_line, re.I)
            if shared_last_match:
                first1, first2, last_name = shared_last_match.groups()
                if first1.lower() == g1.lower() and first2.lower() == g2.lower():
                    full_g1 = f"{first1.title()} {last_name.title()}"
                    full_g2 = f"{first2.title()} {last_name.title()}"
                    if _looks_like_human_name(full_g1) and _looks_like_human_name(full_g2):
                        return (full_g1, full_g2)
            
            # Also try to find the last name from the original clean_line after the separator
            # Pattern: "Michael and Joslyn Mogonye" -> extract "Mogonye"
            remaining_after_sep = re.search(r'(?:and|&)\s+(\w+)\s+(\w+)$', clean_line, re.I)
            if remaining_after_sep:
                second_name, last_name = remaining_after_sep.groups()
                if second_name.lower() == g2.lower():
                    full_g1 = f"{g1.title()} {last_name.title()}"
                    full_g2 = f"{g2.title()} {last_name.title()}"
                    if _looks_like_human_name(full_g1) and _looks_like_human_name(full_g2):
                        return (full_g1, full_g2)
    
    # If only one name found, return it as Guardian1
    if g1 and _looks_like_human_name(g1):
        return (g1, None)
    
    return (None, None)




# -------- Main runner --------
if __name__ == "__main__":
    import sys

    log("Starting extraction run...")
    log(f"Vision key present: {os.path.exists(VISION_CREDENTIALS_FILE)} at {VISION_CREDENTIALS_FILE}")

    # Debug path construction
    log(f"DEBUG: BASE path: {BASE}")
    log(f"DEBUG: LOCAL_PDF_INPUT_DIR: {LOCAL_PDF_INPUT_DIR}")
    log(f"DEBUG: LOCAL_EXCEL_PATH: {LOCAL_EXCEL_PATH}")
    log(f"DEBUG: Current working directory: {os.getcwd()}")

    files = list_local_pdfs(LOCAL_PDF_INPUT_DIR)
    if not files:
        log(f"No PDFs found in: {LOCAL_PDF_INPUT_DIR}")
    else:
        log(f"Found {len(files)} PDF(s) in: {LOCAL_PDF_INPUT_DIR}")

    # Track processing results for exit code
    files_processed = 0
    files_failed = 0

    # Pre-pass: collect ORDER cause numbers to help correct ARP mis-reads
    ORDER_CAUSES: set[str] = set()
    try:
        for f in files:
            nm = (f.get("name", "") or "").lower()
            if "order" in nm:
                try:
                    b = read_pdf_bytes(f.get("path", ""))
                    t = extract_text_with_pdfplumber(b)
                    row_hint = parse_order_fields(t, pdf_bytes_for_vision=b)
                    if row_hint and row_hint.get("causeno"):
                        ORDER_CAUSES.add(row_hint["causeno"])
                except Exception:
                    pass
        if ORDER_CAUSES:
            log(f"ORDER causes seen this run: {sorted(ORDER_CAUSES)}")
    except Exception as _:
        pass

    def _closest_order_cause(cause: str) -> str | None:
        try:
            if not cause or not ORDER_CAUSES:
                return None
            prefix = (cause or "")[:2]
            tail   = (cause or "")[3:]
            best, best_d = None, 10**9
            for oc in ORDER_CAUSES:
                if oc[:2] != prefix:
                    continue
                oc_tail = oc[3:]
                if len(oc_tail) == len(tail):
                    d = sum(1 for a,b in zip(oc_tail, tail) if a!=b)
                else:
                    d = 10**6
                if d < best_d:
                    best, best_d = oc, d
            if best is not None and best_d <= 1:
                return best
            return None
        except Exception:
            return None

    for f in files:
        name = f.get("name", "")
        fpath = f.get("path", "")
        lower_name = (name or "").lower()
        log(f"Processing: {name}")

        try:
            # Skip approvals by filename rule
            if any(tok in lower_name for tok in ("approval", "approved", "approvals")):
                log("  Skipping approval PDF by filename rule.")
                continue

            # Read from local disk now (not Drive)
            pdf_bytes = read_pdf_bytes(fpath)

            # ===== ORDER PDFs =====
            if "order" in lower_name:
                text = extract_text_with_pdfplumber(pdf_bytes)
                log(f"  DEBUG: ORDER pdfplumber extracted {len(text or '')} characters")
                if text:
                    log(f"  DEBUG: First 200 chars: {repr(text[:200])}")
                if len(text or "") < 80:
                    log("  Little text found in ORDER; using OCR fallback (Tesseract)...")
                    text = extract_text_with_ocr_for_arp(pdf_bytes)

                row = parse_order_fields(text, pdf_bytes_for_vision=pdf_bytes)
                if row and row.get("Dateappointed") and row.get("causeno"):
                    row = improve_mapping(row)
                    log(f"  Row to upsert (non-empty): "
                        f"{{'causeno': '{row.get('causeno','')}', "
                        f"'Dateappointed': '{row.get('Dateappointed','')}', "
                        f"'wardfirst': '{row.get('wardfirst','')}', "
                        f"'wardlast': '{row.get('wardlast','')}'}}")
                    upsert_row_to_excel(row)
                    log("  ✅ Upserted Dateappointed from ORDER (by causeno).")
                    files_processed += 1
                else:
                    log("  Skipped ORDER — missing causeno or Signed on date.")
                continue

            # ===== ARP PDFs =====
            if "arp" in lower_name:
                text = extract_text_with_pdfplumber(pdf_bytes)
                log(f"  DEBUG: ARP pdfplumber extracted {len(text or '')} characters")
                if text:
                    log(f"  DEBUG: First 200 chars: {repr(text[:200])}")

                # If pdfplumber failed, try alternative text extraction
                if len(text or "") < 80:
                    log("  Little text found; trying alternative text extraction...")
                    
                    # Try pdfplumber with different settings
                    try:
                        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                            alt_text_parts = []
                            for page in pdf.pages:
                                # Try different extraction methods
                                txt1 = page.extract_text(x_tolerance=5, y_tolerance=5) or ""
                                txt2 = page.extract_text(x_tolerance=1, y_tolerance=1) or ""
                                txt3 = page.extract_text() or ""
                                best_txt = max([txt1, txt2, txt3], key=len)
                                alt_text_parts.append(best_txt)
                            alt_text = "\n".join(alt_text_parts)
                            alt_text = normalize_unicode_noise(clean_text(alt_text))
                            
                            if len(alt_text) > len(text or ""):
                                text = alt_text
                                log(f"  Alternative pdfplumber extracted {len(text)} characters")
                    except Exception as e:
                        log(f"  Alternative pdfplumber failed: {e}")
                    
                    # If still no text, try alternative PDF libraries
                    if len(text or "") < 80:
                        log("  Still little text; trying alternative PDF libraries...")
                        
                        # Try PyPDF2 if available
                        try:
                            import PyPDF2
                            with io.BytesIO(pdf_bytes) as pdf_file:
                                pdf_reader = PyPDF2.PdfReader(pdf_file)
                                pypdf2_text = ""
                                for page in pdf_reader.pages:
                                    pypdf2_text += page.extract_text() or ""
                                pypdf2_text = clean_text(pypdf2_text)
                                if len(pypdf2_text) > len(text or ""):
                                    text = pypdf2_text
                                    log(f"  PyPDF2 extracted {len(text)} characters")
                        except ImportError:
                            log("  PyPDF2 not available")
                        except Exception as e:
                            log(f"  PyPDF2 failed: {e}")
                        
                        # Try pymupdf if available
                        if len(text or "") < 80:
                            try:
                                import fitz  # pymupdf
                                doc = fitz.open(stream=pdf_bytes, filetype="pdf")
                                pymupdf_text = ""
                                for page in doc:
                                    pymupdf_text += page.get_text() or ""
                                doc.close()
                                pymupdf_text = clean_text(pymupdf_text)
                                if len(pymupdf_text) > len(text or ""):
                                    text = pymupdf_text
                                    log(f"  PyMuPDF extracted {len(text)} characters")
                            except ImportError:
                                log("  PyMuPDF not available")
                            except Exception as e:
                                log(f"  PyMuPDF failed: {e}")
                    
                    # If still no text, try Vision API first (better than OCR)
                    if len(text or "") < 80:
                        log("  Still little text; trying Vision API...")
                        try:
                            vision_text = extract_text_with_vision(pdf_bytes)
                            if vision_text and len(vision_text) > len(text or ""):
                                text = vision_text
                                log(f"  Vision API extracted {len(text)} characters")
                        except Exception as e:
                            log(f"  Vision API failed: {e}")
                    
                    # If still no text, try OCR
                    if len(text or "") < 80:
                        log("  Still little text; using ARP OCR (page 1)...")
                        t4 = extract_text_with_ocr_for_arp(pdf_bytes, psm=4)
                        t6 = extract_text_with_ocr_for_arp(pdf_bytes, psm=6)
                        text = t4 if guardian_signal_score(t4) >= guardian_signal_score(t6) else t6
                        log(f"  ARP OCR used {'psm 4' if text==t4 else 'psm 6'} (chars={len(text)})")

                # First parse pass
                row = parse_arp_fields(text)

                # If guardians missing, try Vision
                if (row is None) or (not row.get("guardian1") and not row.get("Guardian2")):
                    log("  Anchored parse weak; trying Vision for handwriting...")
                    try:
                        text_v = extract_text_with_vision(pdf_bytes)
                        if text_v and guardian_signal_score(text_v) > guardian_signal_score(text):
                            text = text_v
                    except Exception as _e:
                        log(f"  Vision error ignored: {_e}")

                # If still weak, try best of psm4/psm6 chooser
                row = parse_arp_fields(text)
                if (row is None) or (not row.get("guardian1") and not row.get("Guardian2")):
                    log("  Vision still weak; retrying Tesseract with both psm 4/psm 6 and picking best...")
                    try:
                        text_best = best_arp_text_from_tesseract(pdf_bytes, current_text=text)
                        if guardian_signal_score(text_best) >= guardian_signal_score(text):
                            text = text_best
                            log(f"  ARP OCR (reselect) chars={len(text)}")
                    except Exception as _e:
                        log(f"  Tesseract reselection error ignored: {_e}")

                  # Final parse
                row = parse_arp_fields(text)
                if row is None or not row.get("causeno"):
                    log("  Skipped — not recognized as ARP or missing cause number.")
                    continue

                # Align ARP cause with nearest ORDER cause if likely off by one digit
                try:
                    corr = _closest_order_cause(row.get("causeno", ""))
                    if corr and corr != row["causeno"]:
                        log(f"  Corrected ARP causeno {row['causeno']} -> {corr} using ORDER hint")
                        row["causeno"] = corr
                except Exception:
                    pass

                # --- B) ARP "Filed" date capture ---
                if not row.get("DateARPfiled"):
                    filed = extract_arp_filed_date(text)
                    if filed:
                        row["DateARPfiled"] = filed

                row = improve_mapping(row)

                # Persist
                non_empty_preview = {k: v for k, v in row.items() if v}
                log(f"  Row to upsert (non-empty): {non_empty_preview}")
                upsert_row_to_excel(row)
                log("  ✅ Upserted row in Excel (by causeno).")
                files_processed += 1
                continue
     


            # ===== Everything else =====
            log("  Unrecognized type (not ORDER/ARP); skipping.")
            continue

        except Exception as e:
            log(f"  ERROR on {name}: {e}")
            files_failed += 1

    log("Done.")

    # Exit with proper error code so GUI can detect failure
    if files_processed == 0 and len(files) > 0:
        print(f"\n[FAIL] OCR Guardian Data FAILED - No files processed successfully (errors: {files_failed})")
        sys.exit(1)
    elif files_processed == 0 and len(files) == 0:
        print("\n[OK] OCR Guardian Data - No PDF files found to process")
        sys.exit(0)
    elif files_failed > 0:
        print(f"\n[WARN] OCR Guardian Data PARTIAL SUCCESS - Processed {files_processed} file(s), {files_failed} failed")
        sys.exit(0)  # Partial success is still success
    else:
        print(f"\n[OK] OCR Guardian Data SUCCESS - Processed {files_processed} file(s)")
        sys.exit(0)