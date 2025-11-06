import openpyxl

# Load mileage template
wb = openpyxl.load_workbook(r'C:\GoogleSync\GuardianShip_App\Templates\Mileage_Reimbursement_Form.xlsx')
ws = wb.active

print('Sheet name:', ws.title)
print('\nSearching for Name/Vendor/GL/Cost Center fields in first 20 rows:')
print('-' * 70)

for row in range(1, 20):
    for col in range(1, 15):
        val = ws.cell(row=row, column=col).value
        if val and isinstance(val, str):
            val_lower = val.lower()
            if any(keyword in val_lower for keyword in ['name', 'vendor', 'cost', 'gl']):
                cell_ref = f"{chr(64+col)}{row}"
                print(f'{cell_ref}: {val}')

print('\n' + '=' * 70)
print('Checking what YOUR current values are:')
print('=' * 70)

# Based on the image you showed, let me check typical cells
check_cells = ['A3', 'B3', 'C3', 'D3', 'E3',
               'A4', 'B4', 'C4', 'D4', 'E4',
               'A5', 'B5', 'C5', 'D5', 'E5',
               'A6', 'B6', 'C6', 'D6', 'E6']

for cell_ref in check_cells:
    val = ws[cell_ref].value
    if val:
        print(f'{cell_ref}: {val}')
