# build_mileage_forms.py
# Builds "2025 Volunteer CV Mileage Reimbursement Form.xlsx" from visit data.
# Uses Google *Directions API* with alternatives=True and chooses the *lowest* miles for each leg.
# Writes rows 15–40, sets E8=MM/YYYY, C41=YES/NO, and creates extra forms as needed.

import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from pathlib import Path
import pandas as pd
import openpyxl
import googlemaps
from typing import List, Dict, Tuple

# =========================
# --- USER CONFIG
# =========================

# Dynamic path detection
_script_dir = Path(__file__).parent.parent.parent.parent  # Go up to app root
try:
    sys.path.insert(0, str(_script_dir / "Scripts"))
    from app_paths import get_app_paths
    from cv_info_manager import get_cv_info
    _app_paths = get_app_paths(str(_script_dir))

    DATA_XLSX = str(_app_paths.EXCEL_PATH)
    TEMPLATE_XLSX = str(_app_paths.APP_ROOT / "Templates" / "Mileage_Reimbursement_Form.xlsx")
    OUTPUT_DIR = str(_app_paths.APP_ROOT / "App Data" / "Output" / "Mileage Logs")
    CONFIG_FILE = str(_app_paths.CONFIG_DIR / "mileage_settings.txt")

except Exception:
    # Fallback to hardcoded paths
    DATA_XLSX = r"C:\GoogleSync\GuardianShip_App\App Data\ward_guardian_info.xlsx"
    TEMPLATE_XLSX = r"C:\GoogleSync\GuardianShip_App\Templates\Mileage_Reimbursement_Form.xlsx"
    OUTPUT_DIR = r"C:\GoogleSync\GuardianShip_App\App Data\Output\Mileage Logs"
    CONFIG_FILE = r"C:\GoogleSync\GuardianShip_App\Config\mileage_settings.txt"

# Default addresses (will be overridden by config file)
# Starting address: where you depart from each day
# Ending address: where you return to each day
DEFAULT_STARTING_ADDRESS = "Default Starting Address"
DEFAULT_ENDING_ADDRESS   = "Default Ending Address"

# Column name aliases in your source sheet (case-insensitive)
HEADER_ALIASES = {
    "visitdate": ["visitdate", "visit_date", "date", "visit date"],
    "visittime": ["visittime", "visit_time", "time", "visit time"],
    "address":   ["waddress", "address", "wardaddress", "ward_address", "location", "toaddress"],
}

# Template layout assumptions (adjust if your form differs)
DATA_START_ROW   = 15
DATA_END_ROW     = 40
ROWS_PER_FORM    = DATA_END_ROW - DATA_START_ROW + 1  # 26
HEADER_ROW_GUESS = 14
CELL_MONTH_YEAR  = "E8"   # MM/YYYY
CELL_MORE_LINES  = "C41"  # YES / NO

# =========================
# --- CONFIG MANAGEMENT
# =========================
def load_user_addresses():
    """Load user's starting and ending addresses from config file, or return defaults."""
    if not os.path.exists(CONFIG_FILE):
        return DEFAULT_STARTING_ADDRESS, DEFAULT_ENDING_ADDRESS

    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            lines = [line for line in f.read().strip().split('\n') if not line.startswith('#')]
            if len(lines) >= 2:
                return lines[0].strip(), lines[1].strip()
    except Exception as e:
        print(f"Error reading config: {e}")

    return DEFAULT_STARTING_ADDRESS, DEFAULT_ENDING_ADDRESS

def save_user_addresses(starting_addr, ending_addr):
    """Save user's starting and ending addresses to config file."""
    try:
        os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            f.write("# Mileage Log Address Settings\n")
            f.write("# Line 1: Starting address (where you depart from each day)\n")
            f.write("# Line 2: Ending address (where you return to each day)\n")
            f.write(f"{starting_addr}\n")
            f.write(f"{ending_addr}\n")
        return True
    except Exception as e:
        print(f"Error saving config: {e}")
        return False

def show_address_settings(current_starting, current_ending):
    """Show dialog to edit default starting and ending addresses. Returns (starting, ending) or None if cancelled."""

    class AddressSettingsDialog:
        def __init__(self, starting_addr, ending_addr):
            self.result = None
            self.root = tk.Tk()
            self.root.title("Mileage Log - Address Settings")
            self.root.geometry("550x350")
            self.root.resizable(False, False)

            # Center window
            self.root.update_idletasks()
            x = (self.root.winfo_screenwidth() // 2) - (275)
            y = (self.root.winfo_screenheight() // 2) - (175)
            self.root.geometry(f'+{x}+{y}')

            # Header
            header = tk.Label(self.root,
                            text="Set Your Starting and Ending Addresses",
                            font=('Segoe UI', 14, 'bold'),
                            pady=15)
            header.pack()

            # Instructions
            instructions = tk.Label(self.root,
                                  text="Enter your starting and ending addresses for mileage calculations.\n"
                                       "These will be used as departure and return points for all trips.",
                                  font=('Segoe UI', 10),
                                  pady=10,
                                  justify='left')
            instructions.pack()

            # Form frame
            form_frame = tk.Frame(self.root, padx=20, pady=10)
            form_frame.pack(fill='x')

            # Starting address
            tk.Label(form_frame, text="Starting Address (departure point):",
                    font=('Segoe UI', 10, 'bold')).pack(anchor='w', pady=(5, 2))
            self.start_entry = tk.Entry(form_frame, font=('Segoe UI', 10), width=60)
            self.start_entry.insert(0, starting_addr)
            self.start_entry.pack(pady=(0, 15))

            # Ending address
            tk.Label(form_frame, text="Ending Address (return point):",
                    font=('Segoe UI', 10, 'bold')).pack(anchor='w', pady=(5, 2))
            self.end_entry = tk.Entry(form_frame, font=('Segoe UI', 10), width=60)
            self.end_entry.insert(0, ending_addr)
            self.end_entry.pack(pady=(0, 10))

            # Note
            note = tk.Label(form_frame,
                          text="Note: Use full address including city, state, and ZIP code\n"
                               "Example: 123 Main St, Austin, TX 78701",
                          font=('Segoe UI', 8),
                          fg='gray',
                          justify='left')
            note.pack(anchor='w')

            # Buttons
            btn_frame = tk.Frame(self.root, pady=15)
            btn_frame.pack()

            save_btn = tk.Button(btn_frame, text="Save",
                               command=self.save,
                               font=('Segoe UI', 10, 'bold'),
                               bg='#667eea', fg='white',
                               padx=20, pady=5)
            save_btn.pack(side='left', padx=5)

            cancel_btn = tk.Button(btn_frame, text="Cancel",
                                  command=self.cancel,
                                  font=('Segoe UI', 10),
                                  padx=20, pady=5)
            cancel_btn.pack(side='left', padx=5)

            self.root.protocol("WM_DELETE_WINDOW", self.cancel)

        def save(self):
            start = self.start_entry.get().strip()
            end = self.end_entry.get().strip()

            if not start or not end:
                messagebox.showerror("Error", "Both addresses are required!")
                return

            self.result = (start, end)
            self.root.destroy()

        def cancel(self):
            self.result = None
            self.root.destroy()

        def show(self):
            self.root.mainloop()
            return self.result

    dialog = AddressSettingsDialog(current_starting, current_ending)
    return dialog.show()

# =========================
# --- HELPERS
# =========================
def find_col_by_header_fuzzy(ws, keywords, max_scan_rows=60):
    """
    Search the top of the sheet for a cell whose text CONTAINS any of the keywords (case-insensitive).
    Works even if headers are not on a fixed row or are merged.
    Returns 1-based column index.
    """
    kw = [k.strip().lower() for k in keywords]
    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                lv = v.strip().lower()
                if any(k in lv for k in kw):
                    return c
    raise ValueError(f"Could not find a column with any of keywords: {keywords}")

def guess_header_map(ws) -> dict:
    """
    Hard-coded mapping for this specific template:
    B = Date, C = From, E = To, F = Miles
    """
    return {"date": 2, "from": 3, "to": 5, "miles": 6}


def resolve_source_headers(df: pd.DataFrame) -> Dict[str, str]:
    cols_lower = {c.lower(): c for c in df.columns}
    resolved = {}
    for logical, aliases in HEADER_ALIASES.items():
        for a in aliases:
            if a in cols_lower:
                resolved[logical] = cols_lower[a]
                break
        if logical not in resolved:
            raise KeyError(
                f"Could not find a column for '{logical}'. "
                f"Tried: {aliases}. Columns present: {list(df.columns)}"
            )
    return resolved

def show_month_picker() -> str:
    """Show GUI picker to select month for mileage log. Returns month string or None if cancelled."""

    class MonthPickerDialog:
        def __init__(self):
            self.result = None
            self.root = tk.Tk()
            self.root.title("Select Month for Mileage Log")
            self.root.geometry("400x280")
            self.root.resizable(False, False)

            # Center the window
            self.root.update_idletasks()
            x = (self.root.winfo_screenwidth() // 2) - (400 // 2)
            y = (self.root.winfo_screenheight() // 2) - (280 // 2)
            self.root.geometry(f'+{x}+{y}')

            # Title
            title_label = tk.Label(self.root, text="Generate Mileage Reimbursement Log",
                                 font=('Segoe UI', 12, 'bold'))
            title_label.pack(pady=15)

            # Month selection frame
            frame = ttk.Frame(self.root, padding=10)
            frame.pack(fill='both', expand=True)

            # Month dropdown
            ttk.Label(frame, text="Select Month:", font=('Segoe UI', 10)).grid(row=0, column=0, sticky='w', pady=5)

            # Generate month options (current month + 11 previous months)
            today = date.today()
            self.month_options = []
            for i in range(12):
                month_date = today.replace(day=1) + relativedelta(months=-i)
                label = month_date.strftime("%B %Y")
                if i == 0:
                    label += " (current)"
                elif i == 1:
                    label += " (previous)"
                # Store as "MM/YYYY" format for the month_prompt_to_range function
                value = month_date.strftime("%m/%Y")
                self.month_options.append((label, value))

            self.month_var = tk.StringVar(value=self.month_options[0][0])  # Default to current month
            month_dropdown = ttk.Combobox(frame, textvariable=self.month_var,
                                        values=[m[0] for m in self.month_options],
                                        state='readonly', width=30)
            month_dropdown.grid(row=0, column=1, pady=5, padx=10)

            # Info label
            info_label = ttk.Label(frame, text="Generates mileage log using Google Maps\n"
                                             "to calculate distances for each visit.",
                                 font=('Segoe UI', 9), foreground='gray')
            info_label.grid(row=1, column=0, columnspan=2, pady=15)

            # Buttons
            btn_frame = ttk.Frame(self.root)
            btn_frame.pack(pady=10)

            ttk.Button(btn_frame, text="Generate", command=self.on_generate,
                      width=12).pack(side='left', padx=5)
            ttk.Button(btn_frame, text="Cancel", command=self.on_cancel,
                      width=12).pack(side='left', padx=5)

            self.root.protocol("WM_DELETE_WINDOW", self.on_cancel)

        def on_generate(self):
            selected_label = self.month_var.get()
            # Find the matching month value
            for label, month_value in self.month_options:
                if label == selected_label:
                    self.result = month_value
                    break
            self.root.destroy()

        def on_cancel(self):
            self.result = None
            self.root.destroy()

        def show(self):
            self.root.mainloop()
            return self.result

    # Show picker dialog
    picker = MonthPickerDialog()
    result = picker.show()

    if result is None:
        print("Mileage log generation cancelled by user")
        sys.exit(0)

    print(f"Selected month: {result}")
    return result

def month_prompt_to_range(month_input: str) -> Tuple[date, date, str, str]:
    """
    Input like '9/2025', '09-2025', '2025-9', '2025/9', 'YYYY-MM', blank (=current), or 'last'.
    Returns (month_start, next_month, label_compact_MMYYYY, label_slash_MM/YYYY).
    """
    today = date.today()
    if not month_input or month_input.strip().lower() == "":
        month_start = date(today.year, today.month, 1)
    elif month_input.strip().lower() == "last":
        this_month = date(today.year, today.month, 1)
        month_start = this_month - relativedelta(months=1)
    else:
        s = month_input.strip().replace("\\", "/").replace("_", "/").replace("-", "/")
        parts = [p for p in s.split("/") if p]
        if len(parts) == 2:
            if len(parts[0]) == 4:
                yyyy = int(parts[0]); mm = int(parts[1])
            else:
                mm = int(parts[0]); yyyy = int(parts[1])
        elif len(parts) == 1 and len(parts[0]) == 7 and parts[4] == "-":
            yyyy = int(parts[0][0:4]); mm = int(parts[0][5:7])
        else:
            dt = pd.to_datetime(month_input, errors="coerce")
            if pd.isna(dt):
                raise ValueError(f"Could not parse month input: '{month_input}'")
            yyyy = int(dt.year); mm = int(dt.month)
        month_start = date(yyyy, mm, 1)

    next_month = month_start + relativedelta(months=1)
    label_compact = f"{month_start.month:02d}{month_start.year:04d}"
    label_slash   = f"{month_start.month:02d}/{month_start.year:04d}"
    return month_start, next_month, label_compact, label_slash

def normalize_time(value):
    # Accepts time-only cells from Excel too
    import datetime as _dt
    if pd.isna(value):
        return None
    if isinstance(value, _dt.time):       # <-- NEW: handles Excel time-only cells
        return value
    if isinstance(value, _dt.datetime):
        return value.time()
    if isinstance(value, pd.Timestamp):
        return value.time()
    if isinstance(value, str):
        try:
            return pd.to_datetime(value).time()
        except Exception:
            return None
    return None

def normalize_date(value):
    if pd.isna(value): return None
    if isinstance(value, datetime): return value.date()
    if isinstance(value, pd.Timestamp): return value.date()
    if isinstance(value, date): return value
    if isinstance(value, str):
        try: return pd.to_datetime(value).date()
        except Exception: return None
    return None

def plan_daily_legs(day_df: pd.DataFrame, date_col: str, time_col: str, addr_col: str,
                   starting_address: str, ending_address: str) -> List[Dict]:
    """
    Build legs:
      STARTING_ADDRESS -> first visit (earliest time)
      visit i -> visit i+1 (in time order)
      last visit -> ENDING_ADDRESS
    """
    keep = []
    drops = []  # rows missing time or address
    for idx, r in day_df.iterrows():
        addr = str(r[addr_col]).strip() if not pd.isna(r[addr_col]) else ""
        t = normalize_time(r[time_col])
        d = normalize_date(r[date_col])
        if addr and t and d:
            keep.append({"date": d, "time": t, "address": addr})
        else:
            drops.append({"row_index": idx, "date": d, "time": t, "address": addr})
    if not keep:
        return [], drops

    keep.sort(key=lambda x: x["time"])
    legs = []
    legs.append({"date": keep[0]["date"], "from": starting_address, "to": keep[0]["address"]})
    for i in range(len(keep) - 1):
        legs.append({"date": keep[i]["date"], "from": keep[i]["address"], "to": keep[i+1]["address"]})
    legs.append({"date": keep[-1]["date"], "from": keep[-1]["address"], "to": ending_address})
    return legs, drops

def chunk(lst: List, size: int) -> List[List]:
    return [lst[i:i+size] for i in range(0, len(lst), size)]

def miles_lowest_of_alternatives(gmaps_client, origin: str, dest: str) -> float:
    """
    Use Google Directions API with alternatives=True.
    Return the *lowest* route distance (in miles) among all returned routes.
    NOTE: No rounding here; let Excel handle formatting/rounding.
    """
    routes = gmaps_client.directions(
        origin=origin,
        destination=dest,
        mode="driving",
        alternatives=True,
        units="imperial"
    )
    if not routes:
        raise RuntimeError(f"No routes returned for {origin} -> {dest}")

    # Each route has legs; for a simple origin->dest there is 1 leg.
    best_meters = None
    for r in routes:
        total_meters = 0
        for leg in r.get("legs", []):
            total_meters += leg["distance"]["value"]
        if best_meters is None or total_meters < best_meters:
            best_meters = total_meters

    if best_meters is None:
        raise RuntimeError(f"Routes returned but no leg distances for {origin} -> {dest}")

    miles = best_meters / 1609.344
    return miles

def fill_one_form(trips: List[Dict], template_path: str, out_path: str, month_label_slash: str,
                  write_yes_in_c41: bool) -> None:
    """
    Write up to ROWS_PER_FORM trips into a copy of the ORIGINAL template,
    respecting merged cells. We keep your form exactly as-is.

    Hard-coded intended columns for your template:
      A = Date (1)
      C = From (3)
      E = To   (5)   # Header likely spans D+E; data cells may be merged D:E
      F = Miles(6)

    We detect merged ranges per row and write to the merged range's ANCHOR (top-left) cell.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Helper: if (row, col) is inside a merged range, return that range's top-left column.
    # Otherwise return col unchanged.
    def anchor_col_for(ws_, row, col):
        for cr in ws_.merged_cells.ranges:
            if cr.min_row <= row <= cr.max_row and cr.min_col <= col <= cr.max_col:
                return cr.min_col  # top-left column of that merged block
        return col

    # Put month & yes/no (safe even if those cells are merged; they're referenced by coordinate)
    ws[CELL_MONTH_YEAR] = month_label_slash      # E8 = MM/YYYY
    ws[CELL_MORE_LINES] = "YES" if write_yes_in_c41 else "NO"   # C41

    # Fill Court Visitor information from config
    try:
        cv_info = get_cv_info()
        ws['B8'] = cv_info.get('name', '')           # Court Visitor Name
        ws['B9'] = cv_info.get('vendor_number', '')  # Vendor Number
        ws['B10'] = cv_info.get('gl_number', '')     # GL #
        ws['B11'] = cv_info.get('cost_center', '')   # Cost Center #
    except Exception as e:
        # Gracefully continue if CV config not available
        print(f"Note: Could not load CV info: {e}")

    # Intended columns (don't change your template!)
    date_col_intended  = 1   # A
    from_col_intended  = 3   # C
    to_col_intended    = 5   # E (header may span D+E, data cells may be merged D:E)
    miles_col_intended = 6   # F

    # Write rows 15–40, honoring merges on each row
    row = DATA_START_ROW
    for t in trips:
        # For each target, use the merged anchor (top-left) if the cell is in a merged range
        dcol = anchor_col_for(ws, row, date_col_intended)
        fcol = anchor_col_for(ws, row, from_col_intended)
        tcol = anchor_col_for(ws, row, to_col_intended)
        mcol = anchor_col_for(ws, row, miles_col_intended)

        ws.cell(row=row, column=dcol).value = t["date"].strftime("%m/%d/%Y")
        ws.cell(row=row, column=fcol).value = t["from"]
        ws.cell(row=row, column=tcol).value = t["to"]
        ws.cell(row=row, column=mcol).value = float(t["miles"])  # leave unrounded; Excel formats it
        row += 1

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)


def main():
    print("=== Volunteer CV Mileage Form Builder ===")

    # Load user's starting and ending addresses (or use defaults)
    STARTING_ADDRESS, ENDING_ADDRESS = load_user_addresses()

    # Check if first time running (no config file exists)
    if not os.path.exists(CONFIG_FILE):
        print("\n[First Run] Please set your starting and ending addresses for mileage calculations...")
        result = show_address_settings(STARTING_ADDRESS, ENDING_ADDRESS)
        if result:
            STARTING_ADDRESS, ENDING_ADDRESS = result
            if save_user_addresses(STARTING_ADDRESS, ENDING_ADDRESS):
                print("[OK] Addresses saved!")
            else:
                print("[WARNING] Could not save addresses. Using defaults for this run.")
        else:
            print("[WARNING] No addresses set. Using default addresses.")

    # Show GUI picker to select month
    month_in = show_month_picker()

    month_start, next_month, label_compact, label_slash = month_prompt_to_range(month_in)

    api_key = os.environ.get("GOOGLE_MAPS_API_KEY", "").strip()
    if not api_key:
        print("ERROR: GOOGLE_MAPS_API_KEY environment variable not set.")
        print("Set it permanently with:")
        print('  setx GOOGLE_MAPS_API_KEY "YOUR_REAL_KEY_HERE"')
        print("…then reopen Command Prompt; or use a .BAT that loads the key from C:\\configlocal\\keys.")
        return
    gmaps_client = googlemaps.Client(key=api_key)

    if not os.path.exists(DATA_XLSX):
        print(f"ERROR: Visit data not found at {DATA_XLSX}")
        return
    df = pd.read_excel(DATA_XLSX)
    header_map = resolve_source_headers(df)

    df["_visitdate_norm"] = df[header_map["visitdate"]].apply(normalize_date)
    df["_visittime_norm"] = df[header_map["visittime"]].apply(normalize_time)
    df["_address_str"]    = df[header_map["address"]].astype(str)

    # Restrict to chosen month
    mask = (df["_visitdate_norm"] >= month_start) & (df["_visitdate_norm"] < next_month)
    mdf = df.loc[mask].copy()

    # ---- diagnostics (debug prints) ----
    print(f"Rows in month {month_start:%m/%Y}: {len(mdf)}")

    # Count missing fields
    missing_time = mdf["_visittime_norm"].isna().sum()
    missing_addr = mdf[header_map["address"]].isna().sum() + (mdf["_address_str"].str.strip() == "").sum()
    missing_date = mdf["_visitdate_norm"].isna().sum()

    print(f"Missing visit time in month: {missing_time}")
    print(f"Missing address in month:   {missing_addr}")
    print(f"Missing date in month:      {missing_date}")

    # Show a small sample of problematic rows (first 5)
    bad_rows = mdf[(mdf["_visittime_norm"].isna()) | (mdf["_address_str"].str.strip() == "") | (mdf["_visitdate_norm"].isna())]
    if not bad_rows.empty:
        cols_to_show = [header_map["visitdate"], header_map["visittime"], header_map["address"]]
        print("\nSample rows with missing fields (up to 5):")
        print(bad_rows[cols_to_show].head(5).to_string(index=False))

    # If the month filter left no rows, stop here
    if mdf.empty:
        print(f"No visits found for {label_compact}.")
        return

    # Build legs per day, also collect what we had to drop
    all_legs = []
    missing_by_day = {}  # date -> list of dropped rows (missing time/address)
    for d, g in mdf.groupby("_visitdate_norm"):
        legs, drops = plan_daily_legs(g, "_visitdate_norm", "_visittime_norm", "_address_str",
                                     STARTING_ADDRESS, ENDING_ADDRESS)
        all_legs.extend(legs)
        if drops:
            missing_by_day[d] = drops

    # If no legs, print why and exit
    if not all_legs:
        print(f"\nNo valid (time+address) legs for {label_compact}.")
        if missing_by_day:
            print("\nDetails of skipped rows by date (missing time or address):")
            for d, drops in sorted(missing_by_day.items()):
                print(f"  {d.strftime('%m/%d/%Y')}: {len(drops)} row(s) skipped")
            print("Fix times/addresses in the source Excel and re-run.")
        else:
            print("No rows were grouped; check column headers and month filter.")
            print(f"Resolved headers: visitdate='{header_map['visitdate']}', visittime='{header_map['visittime']}', address='{header_map['address']}'")
        return

    print(f"Computing Google *lowest* miles for {len(all_legs)} legs…")

    for leg in all_legs:
        try:
            leg["miles"] = miles_lowest_of_alternatives(gmaps_client, leg["from"], leg["to"])
        except Exception as e:
            print(f"Warning: {e}")
            leg["miles"] = 0.0

    parts = chunk(all_legs, ROWS_PER_FORM)

    base_name = f"{label_compact[:2]}_{label_compact[2:]} Volunteer CV Mileage Reimbursement Form"
    outputs = []
    for i, chunk_legs in enumerate(parts, start=1):
        fname = f"{base_name}.xlsx" if i == 1 else f"{base_name} ({i}).xlsx"
        out_path = os.path.join(OUTPUT_DIR, fname)
        write_yes = (i == 1 and len(parts) > 1)
        fill_one_form(chunk_legs, TEMPLATE_XLSX, out_path, label_slash, write_yes)
        outputs.append(out_path)

    print("\nDone. Created:")
    for p in outputs:
        print(" - " + p)

    # Auto-open the finished form(s) on Windows
    print("\nOpening generated mileage form(s)...")
    for p in outputs:
        try:
            os.startfile(p)  # Windows-only
            print(f"  Opened: {p}")
        except Exception as e:
            print(f"  WARNING: Could not open {p}: {e}")


if __name__ == "__main__":
    import sys
    import traceback
    try:
        main()
        print("\n[OK] Mileage Forms SUCCESS")
        sys.exit(0)
    except Exception as e:
        print(f"\n[FAIL] Mileage Forms FAILED: {e}")
        traceback.print_exc()
        sys.exit(1)
