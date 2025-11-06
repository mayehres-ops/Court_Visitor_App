import os
import re
import sys
import time
import shutil
from datetime import datetime
from dateutil.tz import gettz

import pandas as pd

# --- COM automation for Word (pywin32) ---
try:
    import win32com.client as win32
    import pythoncom
except ImportError:
    print("pywin32 is not installed. Run: python -m pip install pywin32")
    sys.exit(1)

# --- openpyxl for non-destructive Excel updates ---
try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl is not installed. Run: python -m pip install openpyxl")
    sys.exit(1)

# =========================
# USER SETTINGS
# =========================
EXCEL_PATH   = r"C:\GoogleSync\GuardianShip_App\App Data\ward_guardian_info.xlsx"
SHEET_NAME   = "Sheet1"  # your tab name

TEMPLATE_PATH = r"C:\GoogleSync\GuardianShip_App\App Data\Templates\Court Visitor Report fillable new.docx"

# Staging (generate here first)
STAGING_DIR  = r"C:\GoogleSync\GuardianShip_App\App Data\Staging"

# Destination base: person folders already exist as subfolders here; we DO NOT create them.
DEST_BASE_FOLDER = r"C:\GoogleSync\GuardianShip_App\New Files"

# Script backups (Excel safety copies go here)
BACKUP_DIR   = r"C:\GoogleSync\GuardianShip_App\App Data\Backup"

# === DONE logic ===
# We ONLY use this column as the gate & marker:
# Blank -> needs new doc; after success -> write "Y".
DONE_COLUMN        = "CVR created?"   # must match Excel header exactly
DONE_MARK_VALUE    = "Y"
DONE_ANCHOR_HEADER = "comments"       # if DONE column missing, create it after this header
TIMEZONE           = "America/Chicago"

# What to do if target file already exists in destination
OVERWRITE_POLICY = "rename"           # "skip" | "rename" | "overwrite"

# Default dry-run; can be overridden by --dry-run / --real
DRY_RUN = False

# Accept header variants for data fill & naming (case/space-insensitive)
COLUMN_ALIASES = {
    "wlast": "wardlast",
    "wfirst": "wardfirst",
    "datearpfiled": "DateARPfiled",
    "cause": "causeno",
    "cause no": "causeno",
    "cause number": "causeno",
}

# =========================
# HELPERS
# =========================
def normalize_key(s: str) -> str:
    return re.sub(r"[\s_]+", "", str(s).strip().lower())

def build_header_map(columns):
    """Map normalized key -> actual Excel column name, honoring aliases when targets exist."""
    m = {}
    norm_cols = {normalize_key(c): c for c in columns}
    for c in columns:
        m[normalize_key(c)] = c
    for alias, target in COLUMN_ALIASES.items():
        tgt_norm = normalize_key(target)
        if tgt_norm in norm_cols:
            m[normalize_key(alias)] = norm_cols[tgt_norm]
    return m

def value_from_row(row, header_map, key):
    col = header_map.get(normalize_key(key))
    if col is None:
        return ""
    v = row.get(col, "")
    if pd.isna(v):
        return ""
    if isinstance(v, (pd.Timestamp, datetime)):
        return v.strftime("%m/%d/%Y")
    return str(v).strip()

def first_nonempty(row, header_map, *keys):
    for k in keys:
        v = value_from_row(row, header_map, k)
        if v:
            return v
    return ""

def clean_segment(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

def join_nonempty(parts, sep=", "):
    parts = [clean_segment(p) for p in parts if clean_segment(p)]
    return sep.join(parts)

def safe_name(name: str) -> str:
    name = (name or "").strip().strip(".")
    return re.sub(r'[<>:"/\\|?*]+', "_", name)

def ensure_ext_docx(name: str) -> str:
    return name if name.lower().endswith(".docx") else name + ".docx"

def save_with_retry(doc, path, retries=3, delay=0.8):
    for i in range(retries):
        try:
            doc.SaveAs2(path)
            return
        except Exception:
            if i < retries - 1:
                time.sleep(delay)
                pythoncom.CoInitialize()
                continue
            raise

# ---------- Folder matching by fields (cause first, then name) ----------
def norm_text(s: str) -> str:
    s = (s or "").lower()
    s = s.replace("–", "-").replace("—", "-")
    s = re.sub(r"[^\w\s\-]", " ", s)   # drop punctuation to spaces
    s = re.sub(r"\s+", " ", s).strip()
    return s

def digits_only(s: str) -> str:
    return re.sub(r"\D+", "", s or "")

def find_person_folder(base: str, cause: str, last: str, first: str) -> str | None:
    """Return best match subfolder path using cause # first, then last/first tokens."""
    try:
        entries = [d for d in os.listdir(base) if os.path.isdir(os.path.join(base, d))]
    except Exception:
        return None

    if not entries:
        return None

    # Precompute normalized forms of each folder
    folders = []
    for d in entries:
        p = os.path.join(base, d)
        folders.append({
            "path": p,
            "name": d,
            "norm": norm_text(d),
            "digits": digits_only(d)
        })

    # 1) Prefer cause number match (unique)
    cause_norm = norm_text(cause)
    cause_dig  = digits_only(cause)

    by_cause = []
    if cause_norm or cause_dig:
        for f in folders:
            score = 0
            # hyphen/spacing tolerant: substring in normalized text
            if cause_norm and cause_norm in f["norm"]:
                score += 2
            # punctuation tolerant: digits-only containment
            if cause_dig and cause_dig and cause_dig in f["digits"]:
                score += 3
            if score:
                by_cause.append((score, f))

        if by_cause:
            # tie-break with last/first hints if present
            last_tok  = norm_text(last)
            first_tok = norm_text(first)
            boosted = []
            for score, f in by_cause:
                if last_tok and last_tok in f["norm"]:
                    score += 1
                if first_tok and first_tok in f["norm"]:
                    score += 1
                boosted.append((score, f))
            boosted.sort(key=lambda x: (-x[0], x[1]["name"]))
            return boosted[0][1]["path"]

    # 2) If no cause hit, try last+first together, then last-only, then first-only
    last_tok  = norm_text(last)
    first_tok = norm_text(first)

    def pick_best(cands):
        if not cands:
            return None
        cands.sort(key=lambda x: (-x[0], x[1]["name"]))
        return cands[0][1]["path"]

    both = []
    if last_tok and first_tok:
        for f in folders:
            if last_tok in f["norm"] and first_tok in f["norm"]:
                both.append((2, f))
        p = pick_best(both)
        if p:
            return p

    only_last = []
    if last_tok:
        for f in folders:
            if last_tok in f["norm"]:
                only_last.append((1, f))
        p = pick_best(only_last)
        if p:
            return p

    only_first = []
    if first_tok:
        for f in folders:
            if first_tok in f["norm"]:
                only_first.append((1, f))
        p = pick_best(only_first)
        if p:
            return p

    return None

# ---------- Excel (non-destructive) ----------
def find_col_index_by_header(ws, header_name: str):
    target = normalize_key(header_name)
    for cell in ws[1]:
        if normalize_key(cell.value) == target:
            return cell.column
    return None

def ensure_done_column(ws, done_col_header: str, anchor_header: str):
    """Ensure DONE column exists; if missing, create it *after* anchor_header. Return its 1-based index."""
    idx = find_col_index_by_header(ws, done_col_header)
    if idx:
        return idx
    anchor_idx = find_col_index_by_header(ws, anchor_header)
    if anchor_idx is None:
        new_idx = ws.max_column + 1
        ws.cell(row=1, column=new_idx).value = done_col_header
        return new_idx
    insert_at = anchor_idx + 1
    ws.insert_cols(insert_at, amount=1)
    ws.cell(row=1, column=insert_at).value = done_col_header
    # carry column width from anchor (best-effort)
    try:
        anchor_letter = ws.cell(row=1, column=anchor_idx).column_letter
        new_letter    = ws.cell(row=1, column=insert_at).column_letter
        ws.column_dimensions[new_letter].width = ws.column_dimensions.get(anchor_letter, ws.column_dimensions[new_letter]).width
    except Exception:
        pass
    return insert_at

def write_values(ws, col_idx: int, df_index_list, value: str):
    """Write given value into column for the provided DataFrame indices."""
    for df_i in df_index_list:
        excel_row = int(df_i) + 2  # +2 because row 1 is header
        ws.cell(row=excel_row, column=col_idx).value = value

# =========================
# MAIN
# =========================
def main():
    global DRY_RUN
    # CLI overrides for dry-run
    if "--dry-run" in sys.argv or "-n" in sys.argv:
        DRY_RUN = True
    if "--real" in sys.argv:
        DRY_RUN = False

    # Pre-flight
    if not os.path.isfile(EXCEL_PATH):
        print(f"ERROR: Excel not found: {EXCEL_PATH}")
        sys.exit(1)
    if not os.path.isfile(TEMPLATE_PATH):
        print(f"ERROR: Template not found: {TEMPLATE_PATH}")
        sys.exit(1)
    for p in [BACKUP_DIR, STAGING_DIR]:
        os.makedirs(p, exist_ok=True)

    print(f"Workbook: {EXCEL_PATH}")
    print(f"Sheet: {SHEET_NAME}")
    print(f"Template: {TEMPLATE_PATH}")
    print(f"Staging: {STAGING_DIR}")
    print(f"Dest base (folders must already exist): {DEST_BASE_FOLDER}")
    print(f"Mode: {'DRY RUN' if DRY_RUN else 'REAL RUN'}")

    # Read (pandas only for reading)
    df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME, engine="openpyxl")
    if DONE_COLUMN not in df.columns:
        df[DONE_COLUMN] = ""

    header_map = build_header_map(df.columns)

    # Rows to process: DONE column is blank (no 'Y' yet)
    done_blank = df[DONE_COLUMN].isna() | (df[DONE_COLUMN].astype(str).str.strip() == "")
    todo = df[done_blank].copy()
    if todo.empty:
        print("No new rows to process.")
        return

    # Start Word
    pythoncom.CoInitialize()
    word = win32.Dispatch("Word.Application")
    word.Visible = False
    try:
        try:
            word.DisplayAlerts = 0
        except Exception:
            pass

        processed_idx = []  # rows we will mark with 'Y'

        for idx, row in todo.iterrows():
            # Pull fields
            last  = first_nonempty(row, header_map, "wardlast", "wlast", "last", "lastname", "ward last")
            first = first_nonempty(row, header_map, "wardfirst", "wfirst", "first", "firstname", "ward first")
            cause = first_nonempty(row, header_map, "causeno", "cause number", "cause no", "cause")

            display_folder = join_nonempty([last, first, cause])
            if not display_folder:
                print("Row missing name/cause — skipping.")
                continue

            folder_key = safe_name(display_folder)
            file_name  = ensure_ext_docx(safe_name(f"{display_folder} Court Visitor Report"))

            # === NEW: find person folder using cause first, then name tokens ===
            match_folder = find_person_folder(DEST_BASE_FOLDER, cause, last, first)
            final_folder = match_folder if match_folder else DEST_BASE_FOLDER
            dest_path    = os.path.join(final_folder, file_name)

            # Always generate in staging first
            staging_path = os.path.join(STAGING_DIR, file_name)
            out_path = staging_path

            print(f"\nRow {idx}:")
            print(f"  Folder key: {folder_key}")
            print(f"  Cause     : {cause}")
            print(f"  Match dir : {match_folder if match_folder else '(none - using base root)'}")
            print(f"  File name : {file_name}")
            print(f"  Final path: {dest_path}")

            # Create the Word doc in STAGING
            if DRY_RUN:
                print(f"  [DRY] Would create in staging: {staging_path}")
            else:
                # Handle staging overwrite for repeated tests
                if os.path.exists(staging_path):
                    if OVERWRITE_POLICY == "skip":
                        print("  Staging exists -> skipping generation (will still attempt move).")
                    elif OVERWRITE_POLICY == "rename":
                        base, ext = os.path.splitext(staging_path)
                        n = 2
                        while True:
                            cand = f"{base} ({n}){ext}"
                            if not os.path.exists(cand):
                                out_path = cand
                                print(f"  Staging exists -> renaming to: {out_path}")
                                break
                            n += 1
                    elif OVERWRITE_POLICY == "overwrite":
                        print("  Staging exists -> overwriting.")

                # Generate
                doc = word.Documents.Add(Template=TEMPLATE_PATH)
                try:
                    for cc in doc.ContentControls:
                        name = (cc.Title or cc.Tag or "").strip()
                        if not name:
                            continue
                        val = value_from_row(row, header_map, name)
                        if not val:
                            alias = COLUMN_ALIASES.get(name.lower())
                            if alias:
                                val = value_from_row(row, header_map, alias)
                        try:
                            cc.Range.Text = val
                        except Exception:
                            pass
                except Exception as e:
                    print(f"  WARNING: Could not iterate/fill content controls: {e}")

                save_with_retry(doc, out_path)
                doc.Close(False)
                print(f"  Saved in staging: {out_path}")

            # Move to final destination (we do NOT create person folders)
            target_path = dest_path
            if os.path.exists(dest_path):
                if OVERWRITE_POLICY == "skip":
                    print("  Destination exists -> skipping move.")
                    processed_idx.append(idx)
                    continue
                elif OVERWRITE_POLICY == "rename":
                    base, ext = os.path.splitext(dest_path)
                    n = 2
                    while True:
                        cand = f"{base} ({n}){ext}"
                        if not os.path.exists(cand):
                            target_path = cand
                            break
                        n += 1
                elif OVERWRITE_POLICY == "overwrite":
                    pass  # overwrite

            if DRY_RUN:
                print(f"  [DRY] Would move to: {target_path}")
            else:
                # Only ensure folder exists if it's the base root; don't create person folders
                if os.path.normcase(os.path.abspath(final_folder)) == os.path.normcase(os.path.abspath(DEST_BASE_FOLDER)):
                    os.makedirs(final_folder, exist_ok=True)
                shutil.move(out_path, target_path)
                print(f"  Moved to: {target_path}")

                # Optional cleanup of stray root copy if we moved into a person folder
                try:
                    if match_folder:
                        stray = os.path.join(DEST_BASE_FOLDER, file_name)
                        if os.path.exists(stray):
                            os.remove(stray)
                            print("  Removed stray root copy:", stray)
                except Exception:
                    pass

            # Mark as processed (we generated and moved)
            processed_idx.append(idx)

    finally:
        try:
            word.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()

    # Non-destructive Excel write-back (only in REAL RUN)
    if processed_idx and not DRY_RUN:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        backup_path = os.path.join(
            BACKUP_DIR,
            f"ward_guardian_info.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        try:
            shutil.copyfile(EXCEL_PATH, backup_path)
            print(f"Backup created: {backup_path}")
        except Exception as e:
            print(f"Backup skipped: {e}")

        wb = load_workbook(EXCEL_PATH)
        if SHEET_NAME not in wb.sheetnames:
            wb.close()
            print(f"ERROR: Sheet '{SHEET_NAME}' not found.")
            sys.exit(1)
        ws = wb[SHEET_NAME]

        done_col_idx = ensure_done_column(ws, DONE_COLUMN, DONE_ANCHOR_HEADER)
        write_values(ws, done_col_idx, processed_idx, DONE_MARK_VALUE)

        wb.save(EXCEL_PATH)
        wb.close()
        print(f"Updated Excel in-place: {EXCEL_PATH} — {len(processed_idx)} row(s) marked '{DONE_MARK_VALUE}' in '{DONE_COLUMN}' on '{SHEET_NAME}'.")
    elif DRY_RUN:
        print("DRY RUN: skipped Excel updates.")

    print("\nDone.")

if __name__ == "__main__":
    main()
