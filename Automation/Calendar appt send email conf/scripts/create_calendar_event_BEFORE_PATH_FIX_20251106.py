#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
Create Google Calendar events from ward_guardian_info.xlsx

- Test mode: process ONLY the last row that has BOTH visitdate & visittime (no Excel writes)
- Live mode: process all eligible rows where:
    Appt_confirmed is blank AND visitdate & visittime present AND
    visit datetime is not in the past and >= 15 minutes from now

- Adds guardian emails (if present and not 'none') as attendees.
- 60-minute events, America/Chicago, reminders at 24h and 2h.
- Logs to: C:\\GoogleSync\\Automation\\Calendar appt send email conf\\Logs\\appt_confirm.log

THIS SCRIPT DOES NOT WRITE 'Y' TO Appt_confirmed.
That will be done by Script #2 (email sender) after it successfully sends.

Usage:
  Test last row (dry run; no OAuth):
    py -3 "C:\\GoogleSync\\Automation\\Calendar appt send email conf\\scripts\\create_calendar_event.py" --mode test_last_row --dry-run

  Test last row (actually create):
    py -3 "C:\\GoogleSync\\Automation\\Calendar appt send email conf\\scripts\\create_calendar_event.py" --mode test_last_row

  Live (all eligible rows):
    py -3 "C:\\GoogleSync\\Automation\\Calendar appt send email conf\\scripts\\create_calendar_event.py" --mode live
"""

import argparse
import datetime as dt
import json
import logging
import os
import re
import sys
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
from dateutil import tz  # noqa: F401 (kept if later needed)
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow

# ------------------------------
# CONFIG (edit if needed)
# ------------------------------
EXCEL_PATH = r"C:\GoogleSync\GuardianShip_App\App Data\ward_guardian_info.xlsx"
BASE_GUARDIAN_FOLDER = r"C:\GoogleSync\GuardianShip_App\New Clients"
LOG_DIR = r"C:\GoogleSync\GuardianShip_App\Automation\Calendar appt send email conf\Logs"

# Check app Config folder first, then fall back to legacy configlocal
APP_CONFIG_DIR = r"C:\GoogleSync\GuardianShip_App\Config\API"
LEGACY_CONFIG_DIR = r"C:\configlocal\API"

# Use app config if it exists, otherwise legacy
if os.path.exists(APP_CONFIG_DIR):
    CREDENTIALS_DIR = APP_CONFIG_DIR
else:
    CREDENTIALS_DIR = LEGACY_CONFIG_DIR

CALENDAR_ID = "primary"
DRIVE_FOLDER_NAME = "Court Visitor ARP Documents"  # Google Drive folder for ARP files
TZ_NAME = "America/Chicago"
EVENT_DURATION_MIN = 60
REMINDERS_MINUTES = [24 * 60, 2 * 60]  # 24h, 2h

# Excel column names in your sheet
COL_CAUSENO = "causeno"
COL_VISITDATE = "visitdate"
COL_VISITTIME = "visittime"
COL_WADDRESS = "waddress"
COL_WARDFIRST = "wardfirst"
COL_WARDMIDDLE = "wardmiddle"
COL_WARDLAST = "wardlast"
COL_G1_NAME = "guardian1"
COL_G1_EMAIL = "gemail"
COL_G1_PHONE = "gtele"
COL_G2_NAME = "Guardian2"
COL_G2_EMAIL_PRIMARY = "g2email"   # accept either of these for g2
COL_G2_EMAIL_TYPO = "g2eamil"
COL_G2_PHONE = "g2tele"
COL_APPT_CONFIRMED = "Appt_confirmed"  # check this column
COL_CALENDARED = "Calendared"  # mark this column

# Minimal Google API scopes for events only
SCOPES = ["https://www.googleapis.com/auth/calendar.events"]

# ------------------------------
# Logging
# ------------------------------
os.makedirs(LOG_DIR, exist_ok=True)
LOG_PATH = os.path.join(LOG_DIR, "appt_confirm.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s: %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)

# ------------------------------
# Helpers
# ------------------------------
def find_client_secret_json(creds_dir: str) -> Path:
    """
    Find a Google OAuth *desktop client* JSON usable for InstalledAppFlow.
    Priority:
      1) GOOGLE_OAUTH_CLIENT_JSON env var (if valid)
      2) Any client_secret*.json / credentials*.json / *.json in creds_dir that has 'installed'
    """
    env_path = os.environ.get("GOOGLE_OAUTH_CLIENT_JSON")
    if env_path and Path(env_path).exists():
        try:
            with open(env_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict) and "installed" in data and "client_id" in data["installed"]:
                return Path(env_path)
        except Exception:
            pass

    p = Path(creds_dir)
    if not p.exists():
        raise FileNotFoundError(f"Credentials dir not found: {creds_dir}")

    candidates = list(p.glob("client_secret*.json")) + list(p.glob("credentials*.json")) + list(p.glob("*.json"))
    for cand in candidates:
        try:
            with open(cand, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict) and "installed" in data and "client_id" in data["installed"]:
                return cand
        except Exception:
            continue

    raise FileNotFoundError(
        f"No OAuth desktop client JSON found in {creds_dir}. "
        f"Create an OAuth Client ID (Desktop app) and place it here, "
        f"or set GOOGLE_OAUTH_CLIENT_JSON to its full path."
    )


def get_calendar_service() -> any:
    """
    Build and return an authenticated Google Calendar service.
    Uses token_calendar.json in CREDENTIALS_DIR (separate from any Gmail token).
    """
    token_path = Path(CREDENTIALS_DIR) / "token_calendar.json"
    creds = None
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            logging.info("Refreshing Google Calendar token...")
            try:
                creds.refresh(Request())
            except Exception as e:
                # Token refresh failed (expired/revoked) - delete and re-authenticate
                logging.warning(f"Token refresh failed: {e}")
                logging.warning("Deleting expired token and starting fresh OAuth flow...")
                if token_path.exists():
                    token_path.unlink()
                creds = None  # Force re-auth below

        if not creds:
            secret_json = find_client_secret_json(CREDENTIALS_DIR)
            logging.info(f"Starting OAuth for Calendar using: {secret_json}")
            flow = InstalledAppFlow.from_client_secrets_file(str(secret_json), SCOPES)
            creds = flow.run_local_server(port=0)

        with open(token_path, "w", encoding="utf-8") as f:
            f.write(creds.to_json())
            logging.info(f"Saved calendar token: {token_path}")
    # cache_discovery=False silences the oauth2client cache warning
    service = build("calendar", "v3", credentials=creds, cache_discovery=False)
    return service


def read_workbook(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, engine="openpyxl")
    df.columns = [c.strip() for c in df.columns]
    return df


def coalesce_g2_email(row: pd.Series) -> str | None:
    """Guardian2 email lives in g2eamil/g2email; normalize; treat 'none' as missing."""
    val = None
    if COL_G2_EMAIL_PRIMARY in row.index and pd.notna(row[COL_G2_EMAIL_PRIMARY]):
        val = str(row[COL_G2_EMAIL_PRIMARY]).strip()
    if (not val) and (COL_G2_EMAIL_TYPO in row.index) and pd.notna(row[COL_G2_EMAIL_TYPO]):
        val = str(row[COL_G2_EMAIL_TYPO]).strip()
    if not val:
        return None
    if val.lower() in {"none", "n/a", "na", "null", "nil"} or "@" not in val:
        return None
    return val


def clean_email(s: str | None) -> str | None:
    if not s:
        return None
    s = str(s).strip()
    if not s or s.lower() in {"none", "n/a", "na", "null", "nil"} or "@" not in s:
        return None
    return s


def pick_attendees(g1_email: str | None, g2_email: str | None) -> list[dict]:
    attendees = []
    if g1_email:
        attendees.append({"email": g1_email})
    if g2_email and g2_email != g1_email:
        attendees.append({"email": g2_email})
    return attendees


# ---- Robust date/time parsing ----
def parse_date_cell(d):
    """Return a date object from mixed Excel/Pandas/string inputs."""
    if pd.isna(d):
        return None
    if isinstance(d, pd.Timestamp):
        return d.date()
    if isinstance(d, dt.datetime):
        return d.date()
    if isinstance(d, dt.date):
        return d
    if isinstance(d, (int, float)):
        try:
            return pd.to_datetime(d, unit="d", origin="1899-12-30").date()
        except Exception:
            pass
    s = str(d).strip()
    if not s:
        return None
    dt_obj = pd.to_datetime(s, errors="coerce", dayfirst=False)
    if pd.isna(dt_obj):
        return None
    return dt_obj.date()


def parse_time_cell(t):
    """Return a time object from mixed Excel/Pandas/string/serial inputs."""
    if pd.isna(t):
        return None
    if isinstance(t, pd.Timestamp):
        return t.time()
    if isinstance(t, dt.datetime):
        return t.time()
    if isinstance(t, dt.time):
        return t
    if isinstance(t, (int, float)):
        try:
            return pd.to_datetime(t, unit="d", origin="1899-12-30").time()
        except Exception:
            pass
    s = str(t).strip().lower()
    if not s or s in {"tbd", "na", "n/a", "none"}:
        return None
    s = s.replace(" ", "").replace(".", ":")
    m = re.fullmatch(r"(\d{1,2})(?::?(\d{2}))?(am|pm)?", s)
    if m:
        hh = int(m.group(1))
        mm = int(m.group(2) or 0)
        ampm = m.group(3)
        if ampm == "pm" and hh < 12:
            hh += 12
        if ampm == "am" and hh == 12:
            hh = 0
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return dt.time(hour=hh, minute=mm)
    parsed = pd.to_datetime(str(t), errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.time()


def combine_date_time_local(d, t, tzname: str) -> dt.datetime | None:
    """Combine separate date/time cells into a tz-aware datetime (America/Chicago)."""
    date_part = parse_date_cell(d)
    time_part = parse_time_cell(t)
    if not date_part or not time_part:
        logging.info(f"DEBUG: raw visit cells -> visitdate={d!r}, visittime={t!r}; parsed date={date_part!r} time={time_part!r}")
        return None
    z = ZoneInfo(tzname)
    return dt.datetime.combine(date_part, time_part).replace(tzinfo=z)


def row_is_eligible(row: pd.Series, now_local: dt.datetime) -> tuple[bool, str]:
    """
    Eligibility for live mode:
     - Appt_confirmed=Y (appointment confirmed)
     - Calendared blank (not yet calendared)
     - visitdate & visittime present
     - datetime >= now + 15 min
    """
    # Must have confirmed appointment first
    appt_val = str(row.get(COL_APPT_CONFIRMED, "")).strip() if COL_APPT_CONFIRMED in row.index else ""
    if appt_val.upper() != "Y":
        return False, "Appt_confirmed not Y"
    
    # Must not already be calendared
    cal_val = str(row.get(COL_CALENDARED, "")).strip() if COL_CALENDARED in row.index else ""
    if cal_val.upper() == "Y":
        return False, "Calendared=Y"

    start_dt = combine_date_time_local(row.get(COL_VISITDATE), row.get(COL_VISITTIME), TZ_NAME)
    if not start_dt:
        return False, "Missing/invalid visitdate/visittime"

    if start_dt < now_local + dt.timedelta(minutes=15):
        return False, "Start < now+15min"

    return True, ""


def build_event(row: pd.Series) -> dict:
    ward_first = str(row.get(COL_WARDFIRST, "") or "").strip()
    ward_last = str(row.get(COL_WARDLAST, "") or "").strip()
    causeno = str(row.get(COL_CAUSENO, "") or "").strip()
    waddress = str(row.get(COL_WADDRESS, "") or "").strip()

    start_dt = combine_date_time_local(row.get(COL_VISITDATE), row.get(COL_VISITTIME), TZ_NAME)
    if not start_dt:
        raise ValueError("Invalid start datetime")

    end_dt = start_dt + dt.timedelta(minutes=EVENT_DURATION_MIN)

    title = f"Court Visit — {ward_last}, {ward_first} — Cause {causeno}".strip(" —")
    description_lines = [
        f"Visit regarding {ward_first} {ward_last} (Cause {causeno}).",
        "Created by automation.",
    ]
    description = "\n".join([ln for ln in description_lines if ln])

    event = {
        "summary": title,
        "description": description,
        "start": {"dateTime": start_dt.isoformat(), "timeZone": TZ_NAME},
        "end": {"dateTime": end_dt.isoformat(), "timeZone": TZ_NAME},
        "location": waddress or None,
        "guestsCanInviteOthers": False,
        "guestsCanModify": False,
        "guestsCanSeeOtherGuests": True,
        "reminders": {
            "useDefault": False,
            "overrides": [{"method": "popup", "minutes": m} for m in REMINDERS_MINUTES],
        },
    }
    return {k: v for k, v in event.items() if v not in (None, "", [])}


def find_case_folder(base_folder: str, causeno: str) -> Path | None:
    """Find a subfolder whose name contains the causeno (one level deep)."""
    base = Path(base_folder)
    if not base.exists():
        return None
    cn = (causeno or "").lower()
    for child in base.iterdir():
        if child.is_dir() and cn in child.name.lower():
            return child
    return None


def list_arp_order_files(case_folder: Path) -> tuple[list[Path], list[Path]]:
    """
    Return (arp_files, order_files) accepting suffixes like ARP, ARP1, ARP2, etc.
    """
    arp, order = [], []
    if not case_folder or not case_folder.exists():
        return arp, order
    for p in case_folder.glob("*.pdf"):
        name = p.name.lower()
        # Look for standardized naming pattern: *_ARP.pdf, *_ORDER.pdf
        if name.endswith("_arp.pdf") or re.search(r"_arp\.pdf$", name):
            arp.append(p)
        elif name.endswith("_order.pdf") or re.search(r"_order\.pdf$", name):
            order.append(p)
        # Fallback to original patterns for backward compatibility
        elif re.search(r"\barp(\d+)?\.pdf$", name) or re.search(r"\barp\b", name) or "arp" in name:
            arp.append(p)
        elif re.search(r"\border(\d+)?\.pdf$", name) or re.search(r"\border\b", name) or "order" in name:
            order.append(p)
    arp.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    order.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return arp, order


def find_existing_event(service, causeno: str, start_dt: dt.datetime) -> str | None:
    """Find existing calendar event for this cause number and return event ID if found."""
    try:
        # Search for events with the cause number in the title
        events_result = service.events().list(
            calendarId=CALENDAR_ID,
            timeMin=start_dt.replace(hour=0, minute=0, second=0).isoformat(),
            timeMax=start_dt.replace(hour=23, minute=59, second=59).isoformat(),
            singleEvents=True,
            orderBy='startTime'
        ).execute()
        
        events = events_result.get('items', [])
        for event in events:
            summary = event.get('summary', '')
            if causeno in summary and 'Court Visit' in summary:
                logging.info(f"Found existing event for cause {causeno}: {event.get('id')}")
                return event.get('id')
        return None
    except Exception as e:
        logging.warning(f"Error searching for existing events: {e}")
        return None


def delete_existing_event(service, event_id: str, causeno: str) -> bool:
    """Delete existing calendar event."""
    try:
        service.events().delete(calendarId=CALENDAR_ID, eventId=event_id).execute()
        logging.info(f"Deleted existing event for cause {causeno}")
        return True
    except Exception as e:
        logging.warning(f"Failed to delete existing event for cause {causeno}: {e}")
        return False


def get_drive_service():
    """Get authenticated Google Drive service."""
    SCOPES = ['https://www.googleapis.com/auth/drive.file']
    creds = None

    # Load existing credentials
    token_path = Path(CREDENTIALS_DIR) / 'token_drive.json'
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)

    # If there are no (valid) credentials available, let the user log in
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception as e:
                # Token refresh failed (expired/revoked) - delete and re-authenticate
                logging.warning(f"Drive token refresh failed: {e}")
                logging.warning("Deleting expired token and starting fresh OAuth flow...")
                if token_path.exists():
                    token_path.unlink()
                creds = None  # Force re-auth below

        if not creds:
            # Use the calendar client secret file for Drive API
            secret_json = find_client_secret_json(CREDENTIALS_DIR)
            flow = InstalledAppFlow.from_client_secrets_file(str(secret_json), SCOPES)
            creds = flow.run_local_server(port=0)

        # Save the credentials for the next run
        with open(token_path, 'w') as token:
            token.write(creds.to_json())

    return build('drive', 'v3', credentials=creds)


def find_or_create_drive_folder(drive_service, folder_name: str) -> str:
    """Find or create a Google Drive folder and return its ID."""
    try:
        # Search for existing folder
        results = drive_service.files().list(
            q=f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false",
            fields="files(id, name)"
        ).execute()
        
        folders = results.get('files', [])
        if folders:
            folder_id = folders[0]['id']
            logging.info(f"Found existing Drive folder: {folder_name} (ID: {folder_id})")
            return folder_id
        
        # Create new folder
        folder_metadata = {
            'name': folder_name,
            'mimeType': 'application/vnd.google-apps.folder'
        }
        folder = drive_service.files().create(body=folder_metadata, fields='id').execute()
        folder_id = folder.get('id')
        logging.info(f"Created new Drive folder: {folder_name} (ID: {folder_id})")
        return folder_id
        
    except Exception as e:
        logging.error(f"Error finding/creating Drive folder: {e}")
        return None


def upload_arp_to_drive(drive_service, arp_file_path: Path, causeno: str, folder_id: str) -> str:
    """Upload ARP file to Google Drive and return the file ID."""
    try:
        file_metadata = {
            'name': f"{causeno}_ARP.pdf",
            'parents': [folder_id]
        }
        
        media = MediaFileUpload(str(arp_file_path), mimetype='application/pdf')
        file = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()
        
        file_id = file.get('id')
        logging.info(f"Uploaded ARP to Drive: {file_id}")
        return file_id
        
    except Exception as e:
        logging.error(f"Error uploading ARP to Drive: {e}")
        return None


def upload_order_to_drive(drive_service, order_file_path: Path, causeno: str, folder_id: str) -> str:
    """Upload ORDER file to Google Drive and return the file ID."""
    try:
        file_metadata = {
            'name': f"{causeno}_ORDER.pdf",
            'parents': [folder_id]
        }
        
        media = MediaFileUpload(str(order_file_path), mimetype='application/pdf')
        file = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id'
        ).execute()
        
        file_id = file.get('id')
        logging.info(f"Uploaded ORDER to Drive: {file_id}")
        return file_id
        
    except Exception as e:
        logging.error(f"Error uploading ORDER to Drive: {e}")
        return None


def main():
    parser = argparse.ArgumentParser(description="Create Google Calendar events from Excel.")
    parser.add_argument("--mode", choices=["test_last_row", "live"], required=True, help="Run mode")
    parser.add_argument("--dry-run", action="store_true", help="Do not create events; just log")
    args = parser.parse_args()

    logging.info("=== Start: create_calendar_event ===")
    logging.info(f"Excel: {EXCEL_PATH}")
    logging.info(f"Mode: {args.mode}  Dry-run: {args.dry_run}")

    # Load data
    try:
        df = read_workbook(EXCEL_PATH)
    except Exception as e:
        logging.exception(f"Failed to read workbook: {e}")
        sys.exit(2)

    # Ensure basic columns exist (don't require emails)
    required = [COL_CAUSENO, COL_VISITDATE, COL_VISITTIME, COL_WADDRESS, COL_WARDFIRST, COL_WARDLAST]
    missing = [c for c in required if c not in df.columns]
    if missing:
        logging.error(f"Missing expected columns: {missing}")
        sys.exit(3)

    # Determine rows to process
    if args.mode == "test_last_row":
        candidates = df[df[[COL_VISITDATE, COL_VISITTIME]].notna().all(axis=1)]
        if candidates.empty:
            logging.warning("No rows with BOTH visitdate and visittime set. Nothing to test.")
            return
        idx = candidates.index[-1]
        rows = [df.loc[idx]]
        logging.info(f"Test mode: using last row with BOTH date & time index={idx}")
    else:
        rows = [r for _, r in df.iterrows()]

    now_local = dt.datetime.now(ZoneInfo(TZ_NAME))
    created = 0
    skipped = 0

    # Only build the Calendar service if we are going to create something
    service = None
    if not args.dry_run:
        service = get_calendar_service()

    for row in rows:
        causeno = str(row.get(COL_CAUSENO, "") or "").strip()
        ward_first = str(row.get(COL_WARDFIRST, "") or "").strip()
        ward_last = str(row.get(COL_WARDLAST, "") or "").strip()

        if args.mode == "live":
            ok, reason = row_is_eligible(row, now_local)
            if not ok:
                skipped += 1
                logging.info(f"Skip cause {causeno or '<none>'}: {reason}")
                continue
        else:
            # In test mode, still bail early if datetime can't be parsed
            start_dt = combine_date_time_local(row.get(COL_VISITDATE), row.get(COL_VISITTIME), TZ_NAME)
            if not start_dt:
                skipped += 1
                logging.info(f"Skip cause {causeno or '<none>'}: invalid date/time in test row")
                continue

        # Attendees
        g1_email = clean_email(row.get(COL_G1_EMAIL)) if COL_G1_EMAIL in row.index else None
        g2_email = coalesce_g2_email(row)
        attendees = pick_attendees(g1_email, g2_email)

        # Log raw cells to help troubleshoot
        logging.info(f"Raw cells: visitdate={row.get(COL_VISITDATE)!r}, visittime={row.get(COL_VISITTIME)!r}")

        # Build event
        try:
            event = build_event(row)
        except Exception as e:
            skipped += 1
            logging.warning(f"Row build failed for cause {causeno}: {e}")
            continue

        if attendees:
            event["attendees"] = attendees

        # Case folder & docs - upload both ARP and ORDER to Drive and attach to event
        case_folder = find_case_folder(BASE_GUARDIAN_FOLDER, causeno) if causeno else None
        arp_files, order_files = ([], [])
        arp_file_id = None
        order_file_id = None
        
        if case_folder:
            arp_files, order_files = list_arp_order_files(case_folder)
            
            # Upload files to Google Drive if found
            if arp_files or order_files:
                if args.dry_run:
                    # In dry-run mode, just log what would be attached
                    if arp_files:
                        logging.info(f"DRY-RUN: Would upload ARP file: {arp_files[0].name}")
                    if order_files:
                        logging.info(f"DRY-RUN: Would upload ORDER file: {order_files[0].name}")
                    logging.info(f"DRY-RUN: Would attach files to calendar event")
                else:
                    try:
                        drive_service = get_drive_service()
                        folder_id = find_or_create_drive_folder(drive_service, DRIVE_FOLDER_NAME)
                        if folder_id:
                            # Upload ARP file if found
                            if arp_files:
                                arp_file_id = upload_arp_to_drive(drive_service, arp_files[0], causeno, folder_id)
                            
                            # Upload ORDER file if found
                            if order_files:
                                order_file_id = upload_order_to_drive(drive_service, order_files[0], causeno, folder_id)
                            
                            # Add attachments to event
                            if arp_file_id or order_file_id:
                                if 'attachments' not in event:
                                    event['attachments'] = []
                                
                                if arp_file_id:
                                    event['attachments'].append({
                                        'fileUrl': f"https://drive.google.com/file/d/{arp_file_id}/view",
                                        'title': f"{causeno}_ARP.pdf"
                                    })
                                
                                if order_file_id:
                                    event['attachments'].append({
                                        'fileUrl': f"https://drive.google.com/file/d/{order_file_id}/view",
                                        'title': f"{causeno}_ORDER.pdf"
                                    })
                    except Exception as e:
                        logging.warning(f"Failed to upload files to Drive for cause {causeno}: {e}")

        logging.info(
            f"Prepared event for cause={causeno}, ward={ward_last}, {ward_first}; "
            f"attendees={[a['email'] for a in attendees] if attendees else []}; "
            f"case_folder={case_folder or 'n/a'}; "
            f"ARP_found={len(arp_files)}; Order_found={len(order_files)}; "
            f"ARP_attached={'Yes' if arp_file_id else 'No'}; ORDER_attached={'Yes' if order_file_id else 'No'}"
        )

        if args.dry_run:
            logging.info("DRY-RUN: Would insert calendar event.")
            continue

        try:
            # Check for existing event and delete if found (rescheduling)
            start_dt = combine_date_time_local(row.get(COL_VISITDATE), row.get(COL_VISITTIME), TZ_NAME)
            existing_event_id = find_existing_event(service, causeno, start_dt)
            if existing_event_id:
                logging.info(f"Found existing event for cause {causeno}, deleting old event...")
                delete_existing_event(service, existing_event_id, causeno)
            
            # Create new event with attachments support
            created_event = (
                service.events()
                .insert(calendarId=CALENDAR_ID, body=event, sendUpdates="all", supportsAttachments=True)
                .execute()
            )
            created += 1
            logging.info(f"Created event: {created_event.get('htmlLink')}")
            
            # Mark Calendared as "Y" to prevent duplicates
            try:
                import openpyxl
                wb = openpyxl.load_workbook(EXCEL_PATH)
                ws = wb["Sheet1"]
                
                # Find the row with this cause number (causeno is in column 4)
                for row_idx in range(2, ws.max_row + 1):
                    if str(ws.cell(row=row_idx, column=4).value).strip() == causeno:
                        # Find or create Calendared column
                        cal_col = None
                        for col_idx in range(1, ws.max_column + 1):
                            if ws.cell(row=1, column=col_idx).value == "Calendared":
                                cal_col = col_idx
                                break
                        
                        if cal_col is None:
                            # Create Calendared column at the end
                            cal_col = ws.max_column + 1
                            ws.cell(row=1, column=cal_col, value="Calendared")
                        
                        # Mark Calendared as "Y"
                        ws.cell(row=row_idx, column=cal_col, value="Y")
                        wb.save(EXCEL_PATH)
                        logging.info(f"Marked Calendared=Y for cause {causeno}")
                        break
            except Exception as excel_error:
                logging.warning(f"Failed to mark Calendared for cause {causeno}: {excel_error}")
                
        except Exception as e:
            skipped += 1
            logging.exception(f"Failed to create event for cause {causeno}: {e}")

    logging.info(f"Done. Created={created}, Skipped={skipped}")


if __name__ == "__main__":
    import sys
    import traceback
    try:
        main()
        print("\n[OK] Calendar Events SUCCESS")
        sys.exit(0)
    except KeyboardInterrupt:
        logging.warning("Interrupted by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\n[FAIL] Calendar Events FAILED: {e}")
        traceback.print_exc()
        sys.exit(1)
