#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
Send confirmation emails to guardians, attach ARP/Order if found, and mark Appt_confirmed='Y'.

- Test mode: process ONLY the last row that has BOTH visitdate & visittime (no writes in --dry-run)
- Live mode: process all eligible rows where:
    Appt_confirmed is blank AND visitdate & visittime present AND
    visit datetime is not in the past and >= 15 minutes from now

Email rules:
  * TO: Guardian1 email if present; else Guardian2
  * CC: the other guardian (if present) + YOU (always, on real sends)
  * Never email the ward
  * If ARP/Order missing: DO NOT mention it to guardians; send a private FYI email to you
  * Save a text copy to client folder or _Correspondence_Pending

Excel writes:
  * Adds 'Appt_confirmed' at the END if missing
  * After a successful send, writes 'Y' in that row
  * First write of a run makes a timestamped workbook backup

Usage:
  Test last row (dry run):
    py -3 "C:\\GoogleSync\\Automation\\Appt Email Confirm\\scripts\\send_confirmation_email.py" --mode test_last_row --dry-run

  Test last row (send one + mark Y):
    py -3 "C:\\GoogleSync\\Automation\\Appt Email Confirm\\scripts\\send_confirmation_email.py" --mode test_last_row

  Live (all eligible rows):
    py -3 "C:\\GoogleSync\\Automation\\Appt Email Confirm\\scripts\\send_confirmation_email.py" --mode live
"""

import argparse
import base64
import datetime as dt
import json
import logging
import mimetypes
import os
import re
import shutil
import sys
import urllib.parse
from email.message import EmailMessage
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
from googleapiclient.discovery import build
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from openpyxl import load_workbook

# Dynamic path management - works from any installation location
_script_dir = Path(__file__).parent.parent.parent.parent  # Go up to app root

try:
    sys.path.insert(0, str(_script_dir / "Scripts"))
    from app_paths import get_app_paths

    _app_paths = get_app_paths(str(_script_dir))
    EXCEL_PATH = str(_app_paths.EXCEL_PATH)
    BASE_GUARDIAN_FOLDER = str(_app_paths.NEW_CLIENTS_DIR)
    BASE_DIR = str(_app_paths.APP_ROOT / "Automation" / "Appt Email Confirm")
    LOG_DIR = str(Path(BASE_DIR) / "Logs")
    _CONFIG_DIR = _app_paths.CONFIG_DIR
except Exception:
    # Fallback to hardcoded paths
    EXCEL_PATH = r"C:\GoogleSync\GuardianShip_App\App Data\ward_guardian_info.xlsx"
    BASE_GUARDIAN_FOLDER = r"C:\GoogleSync\GuardianShip_App\New Clients"
    BASE_DIR = r"C:\GoogleSync\GuardianShip_App\Automation\Appt Email Confirm"
    LOG_DIR = rf"{BASE_DIR}\Logs"
    _CONFIG_DIR = Path(r"C:\GoogleSync\GuardianShip_App\Config")

# Check app Config folder first, then fall back to legacy configlocal
APP_CONFIG_DIR = str(_CONFIG_DIR / "API")
LEGACY_CONFIG_DIR = r"C:\configlocal\API"

# Use app config if it exists, otherwise legacy
if os.path.exists(APP_CONFIG_DIR):
    CREDENTIALS_DIR = APP_CONFIG_DIR
else:
    CREDENTIALS_DIR = LEGACY_CONFIG_DIR

TZ_NAME = "America/Chicago"

YOUR_MOBILE = "317-339-9963"  # e.g., "512-555-1234" (shows in email body)
GOOGLE_FORM_URL = "https://docs.google.com/forms/d/e/1FAIpQLSd4MojVDY_iFlwntWqCtkL2wj7LiYwrgmm5CbG5EpKkexBkXQ/viewform?usp=dialog"

# Excel columns
COL_CAUSENO = "causeno"
COL_VISITDATE = "visitdate"
COL_VISITTIME = "visittime"
COL_WADDRESS = "waddress"
COL_WARDFIRST = "wardfirst"
COL_WARDLAST = "wardlast"
COL_G1_NAME = "guardian1"
COL_G1_EMAIL = "gemail"
COL_G2_NAME = "Guardian2"
COL_G2_EMAIL_PRIMARY = "g2email"
COL_G2_EMAIL_TYPO = "g2eamil"
COL_EMAIL_SENT = "emailsent"
COL_APPT_CONFIRMED = "Appt_confirmed"

# OAuth scopes (minimal)
SCOPES_GMAIL = ["https://www.googleapis.com/auth/gmail.send"]
SCOPES_PROFILE = ["https://www.googleapis.com/auth/gmail.readonly"]  # to get your own address for CC

# =========================
# Logging & dirs
# =========================
os.makedirs(LOG_DIR, exist_ok=True)
LOG_PATH = os.path.join(LOG_DIR, "appt_confirm.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s: %(message)s",
    handlers=[logging.FileHandler(LOG_PATH, encoding="utf-8"), logging.StreamHandler(sys.stdout)],
)

# =========================
# Helpers
# =========================
def find_client_secret_json(creds_dir: str) -> Path:
    env_path = os.environ.get("GOOGLE_OAUTH_CLIENT_JSON")
    if env_path and Path(env_path).exists():
        try:
            data = json.load(open(env_path, "r", encoding="utf-8"))
            if "installed" in data and "client_id" in data["installed"]:
                return Path(env_path)
        except Exception:
            pass
    p = Path(creds_dir)
    if not p.exists():
        raise FileNotFoundError(f"Credentials dir not found: {creds_dir}")
    for cand in list(p.glob("client_secret*.json")) + list(p.glob("credentials*.json")) + list(p.glob("*.json")):
        try:
            data = json.load(open(cand, "r", encoding="utf-8"))
            if "installed" in data and "client_id" in data["installed"]:
                return cand
        except Exception:
            continue
    raise FileNotFoundError("No OAuth desktop client JSON found.")

def get_gmail_service():
    token_path = Path(CREDENTIALS_DIR) / "token_gmail.json"
    creds = None
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES_GMAIL + SCOPES_PROFILE)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            logging.info("Refreshing Gmail token...")
            try:
                creds.refresh(Request())
            except Exception as e:
                # Token refresh failed (expired/revoked) - delete and re-authenticate
                logging.warning(f"Token refresh failed: {e}")
                logging.info("Deleting expired token and starting fresh OAuth flow...")
                if token_path.exists():
                    token_path.unlink()
                creds = None  # Force re-auth below

        if not creds:
            secret_json = find_client_secret_json(CREDENTIALS_DIR)
            logging.info(f"Starting OAuth for Gmail using: {secret_json}")
            flow = InstalledAppFlow.from_client_secrets_file(str(secret_json), SCOPES_GMAIL + SCOPES_PROFILE)
            creds = flow.run_local_server(port=0)

        open(token_path, "w", encoding="utf-8").write(creds.to_json())
        logging.info(f"Saved Gmail token: {token_path}")
    return build("gmail", "v1", credentials=creds, cache_discovery=False)

def get_my_email_address(gmail_service) -> str:
    profile = gmail_service.users().getProfile(userId="me").execute()
    return profile.get("emailAddress")

def read_workbook_df(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, engine="openpyxl")
    df.columns = [c.strip() for c in df.columns]
    return df

def clean_email(s):
    if not s:
        return None
    s = str(s).strip()
    if not s or s.lower() in {"none", "n/a", "na", "null", "nil"} or "@" not in s:
        return None
    return s

def coalesce_g2_email(row: pd.Series) -> str | None:
    v = None
    if COL_G2_EMAIL_PRIMARY in row.index and pd.notna(row[COL_G2_EMAIL_PRIMARY]):
        v = str(row[COL_G2_EMAIL_PRIMARY]).strip()
    if (not v) and (COL_G2_EMAIL_TYPO in row.index) and pd.notna(row[COL_G2_EMAIL_TYPO]):
        v = str(row[COL_G2_EMAIL_TYPO]).strip()
    if not v or v.lower() in {"none", "n/a", "na", "null", "nil"} or "@" not in v:
        return None
    return v

# ---- date/time parsing (same as Script #1) ----
def parse_date_cell(d):
    if pd.isna(d): return None
    if isinstance(d, pd.Timestamp): return d.date()
    if isinstance(d, dt.datetime): return d.date()
    if isinstance(d, dt.date): return d
    if isinstance(d, (int, float)):
        try: return pd.to_datetime(d, unit="d", origin="1899-12-30").date()
        except Exception: pass
    s = str(d).strip()
    if not s: return None
    dt_obj = pd.to_datetime(s, errors="coerce", dayfirst=False)
    if pd.isna(dt_obj): return None
    return dt_obj.date()

def parse_time_cell(t):
    if pd.isna(t): return None
    if isinstance(t, pd.Timestamp): return t.time()
    if isinstance(t, dt.datetime): return t.time()
    if isinstance(t, dt.time): return t
    if isinstance(t, (int, float)):
        try: return pd.to_datetime(t, unit="d", origin="1899-12-30").time()
        except Exception: pass
    s = str(t).strip().lower()
    if not s or s in {"tbd", "na", "n/a", "none"}: return None
    s = s.replace(" ", "").replace(".", ":")
    m = re.fullmatch(r"(\d{1,2})(?::?(\d{2}))?(am|pm)?", s)
    if m:
        hh = int(m.group(1)); mm = int(m.group(2) or 0); ampm = m.group(3)
        if ampm == "pm" and hh < 12: hh += 12
        if ampm == "am" and hh == 12: hh = 0
        if 0 <= hh <= 23 and 0 <= mm <= 59: return dt.time(hour=hh, minute=mm)
    parsed = pd.to_datetime(str(t), errors="coerce")
    if pd.isna(parsed): return None
    return parsed.time()

def combine_date_time_local(d, t, tzname: str):
    date_part = parse_date_cell(d); time_part = parse_time_cell(t)
    if not date_part or not time_part:
        logging.info(f"DEBUG: raw visit cells -> visitdate={d!r}, visittime={t!r}; parsed date={date_part!r} time={time_part!r}")
        return None
    return dt.datetime.combine(date_part, time_part).replace(tzinfo=ZoneInfo(tzname))

def row_is_eligible(row: pd.Series, now_local: dt.datetime):
    appt_val = str(row.get(COL_APPT_CONFIRMED, "")).strip() if COL_APPT_CONFIRMED in row.index else ""
    if appt_val.upper() == "Y": return False, "Appt_confirmed=Y"
    start_dt = combine_date_time_local(row.get(COL_VISITDATE), row.get(COL_VISITTIME), TZ_NAME)
    if not start_dt: return False, "Missing/invalid visitdate/visittime"
    if start_dt < now_local + dt.timedelta(minutes=15): return False, "Start < now+15min"
    return True, ""

# ---- ARP/Order locate ----
def find_case_folder(base_folder: str, causeno: str):
    base = Path(base_folder)
    if not base.exists(): return None
    cn = (causeno or "").lower()
    for child in base.iterdir():
        if child.is_dir() and cn in child.name.lower(): return child
    return None

def list_arp_order_files(case_folder: Path):
    arp, order = [], []
    if not case_folder or not case_folder.exists(): return arp, order
    for p in case_folder.glob("*.pdf"):
        n = p.name.lower()
        # Look for standardized naming pattern: *_ARP.pdf, *_ORDER.pdf
        if n.endswith("_arp.pdf") or re.search(r"_arp\.pdf$", n):
            arp.append(p)
        elif n.endswith("_order.pdf") or re.search(r"_order\.pdf$", n):
            order.append(p)
        # Fallback to original patterns for backward compatibility
        elif re.search(r"\barp(\d+)?\.pdf$", n) or re.search(r"\barp\b", n): 
            arp.append(p)
        elif re.search(r"\border(\d+)?\.pdf$", n) or re.search(r"\border\b", n) or "order" in n: 
            order.append(p)
    arp.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    order.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return arp, order

def pick_arp_order(arp_files, order_files):
    # For emails, only attach ORDER files (not ARP)
    # ARP files will be attached to calendar events instead
    return None, (order_files[0] if order_files else None)

# ---- Email rendering ----
def mk_maps_link(address: str) -> str:
    return f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote_plus(address)}"

def render_email_html(ward_first, ward_last, waddress):
    maps_link = mk_maps_link(waddress) if waddress else "#"
    form_link = GOOGLE_FORM_URL
    mobile = YOUR_MOBILE or "your phone"
    return f"""<p>Dear Guardian,</p>
<p>I’m looking forward to getting to know <strong>{ward_first} {ward_last}</strong> and speaking with you.</p>
<p><strong>When:</strong> {{Date}} at {{Time}} (about 60 minutes)<br>
<strong>Where:</strong> {waddress if waddress else "[address pending]"} [ <a href="{maps_link}">map</a> ]</p>
<p>Quick logistics so I arrive without surprises:</p>
<ul>
  <li>If you’re in a gated community or have a gated home, how should I get in?</li>
  <li>If the home doesn’t have a visible house number, what should I look for?</li>
  <li>If parking is tight, where’s the best place to park?</li>
</ul>
<p>I often have several visits back-to-back. I aim to be right on time, but if I’m more than <strong>15 minutes</strong> early or late, I’ll call you—unless I’m actively driving and can’t call safely.</p>
<p><em>Optional (zero pressure):</em> This short pre-visit questionnaire covers many of the topics I usually ask about. It’s completely optional—if you’re pressed for time or prefer to discuss at the visit, that’s absolutely fine. It can save a few minutes if you do have time.</p>
<p><a href="{form_link}">Open the optional pre-visit questionnaire</a></p>
<p>If you need to reschedule, you can reply to this email, <strong>text</strong> me at {mobile}, or call. (Text is best.)</p>
<p>Thank you—looking forward to meeting you and {ward_first}.</p>
<p>May Ehresman</p>
<p style="font-size: 0.9em; color:#666">You’ll also receive a calendar invitation for the appointment time.</p>"""

def render_email_text(ward_first, ward_last, waddress):
    maps_link = mk_maps_link(waddress) if waddress else "#"
    form_link = GOOGLE_FORM_URL
    mobile = YOUR_MOBILE or "your phone"
    return f"""Dear Guardian,

I’m looking forward to getting to know {ward_first} {ward_last} and speaking with you.

When: {{Date}} at {{Time}} (about 60 minutes)
Where: {waddress if waddress else "[address pending]"} [map: {maps_link}]

Quick logistics so I arrive without surprises:
- If you’re in a gated community or have a gated home, how should I get in?
- If the home doesn’t have a visible house number, what should I look for?
- If parking is tight, where’s the best place to park?

I often have several visits back-to-back. I aim to be right on time, but if I’m more than 15 minutes early or late, I’ll call you—unless I’m actively driving and can’t call safely.

Optional (zero pressure): This short pre-visit questionnaire covers many of the topics I usually ask about. It’s completely optional—if you’re pressed for time or prefer to discuss at the visit, that’s absolutely fine. It can save a few minutes if you do have time.

Form: {form_link}

If you need to reschedule, you can reply to this email, text me at {mobile}, or call. (Text is best.)

Thank you—looking forward to meeting you and {ward_first}.

May Ehresman

You’ll also receive a calendar invitation for the appointment time.
"""

def build_mime(from_addr, to_addr, cc_addrs, subject, html_body, text_body, attachments):
    from email.message import EmailMessage
    msg = EmailMessage()
    msg["From"] = from_addr
    msg["To"] = to_addr
    if cc_addrs: msg["Cc"] = ", ".join(cc_addrs)
    msg["Subject"] = subject
    msg.set_content(text_body)
    msg.add_alternative(html_body, subtype="html")
    for path in attachments:
        ctype, enc = mimetypes.guess_type(str(path))
        if ctype is None or enc is not None: ctype = "application/octet-stream"
        maintype, subtype = ctype.split("/", 1)
        with open(path, "rb") as f: data = f.read()
        msg.add_attachment(data, maintype=maintype, subtype=subtype, filename=path.name)
    return msg

def gmail_send_raw(gmail_service, mime_msg):
    raw = base64.urlsafe_b64encode(mime_msg.as_bytes()).decode("utf-8")
    return gmail_service.users().messages().send(userId="me", body={"raw": raw}).execute()

def find_case_folder(base_dir, cause_no):
    """Find client folder by cause number, returns Path object or None"""
    try:
        for name in os.listdir(base_dir):
            full = os.path.join(base_dir, name)
            if os.path.isdir(full) and str(cause_no) in name:
                return Path(full)
    except FileNotFoundError:
        pass
    return None

def save_text_copy(folder, filename, subject, body, to_email, cc_emails):
    """Save text copy to client folder or _Correspondence_Pending"""
    os.makedirs(folder, exist_ok=True)
    out_path = os.path.join(folder, filename)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"TO: {to_email}\n")
        if cc_emails:
            f.write(f"CC: {cc_emails}\n")
        f.write(f"SUBJECT: {subject}\n\n")
        f.write(body)
    return out_path

def backup_workbook(xlsx_path: str):
    p = Path(xlsx_path)
    ts = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = p.with_name(p.stem + f".backup-{ts}" + p.suffix)
    shutil.copy2(p, backup)
    logging.info(f"Workbook backup created: {backup}")
    return backup

def ensure_appt_confirmed_column(wb_path: str, header_name: str) -> int:
    wb = load_workbook(wb_path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if header_name in headers:
        col_idx = headers.index(header_name) + 1
    else:
        col_idx = len(headers) + 1
        ws.cell(row=1, column=col_idx, value=header_name)
    wb.save(wb_path); wb.close()
    return col_idx

def write_Y_at_row(wb_path: str, col_idx: int, df_index: int):
    wb = load_workbook(wb_path)
    ws = wb.active
    excel_row = df_index + 2  # header is row 1
    ws.cell(row=excel_row, column=col_idx, value="Y")
    wb.save(wb_path); wb.close()

# =========================
# Main
# =========================
def main():
    parser = argparse.ArgumentParser(description="Send guardian confirmation emails and mark Appt_confirmed.")
    parser.add_argument("--mode", choices=["test_last_row", "live"], required=True)
    parser.add_argument("--dry-run", action="store_true", help="Log only; do not send; do not write Excel")
    args = parser.parse_args()

    logging.info("=== Start: send_confirmation_email ===")
    logging.info(f"Excel: {EXCEL_PATH}")
    logging.info(f"Mode: {args.mode}  Dry-run: {args.dry_run}")

    # Gmail service
    gmail = None; my_email = None
    if not args.dry_run:
        gmail = get_gmail_service()
        my_email = get_my_email_address(gmail)
        logging.info(f"Sending as: {my_email}")
    else:
        logging.info("DRY-RUN: Skipping Gmail OAuth")

    # Load data
    try:
        df = read_workbook_df(EXCEL_PATH)
    except Exception as e:
        logging.exception(f"Failed to read workbook: {e}"); sys.exit(2)

    required = [COL_CAUSENO, COL_VISITDATE, COL_VISITTIME, COL_WADDRESS, COL_WARDFIRST, COL_WARDLAST]
    missing = [c for c in required if c not in df.columns]
    if missing:
        logging.error(f"Missing expected columns: {missing}"); sys.exit(3)

    # Pick rows
    if args.mode == "test_last_row":
        candidates = df[df[[COL_VISITDATE, COL_VISITTIME]].notna().all(axis=1)]
        if candidates.empty:
            logging.warning("No rows with BOTH visitdate and visittime set. Nothing to test."); return
        idx = candidates.index[-1]
        rows = [(idx, df.loc[idx])]
        logging.info(f"Test mode: using last row with BOTH date & time index={idx}")
    else:
        rows = list(df.iterrows())

    now_local = dt.datetime.now(ZoneInfo(TZ_NAME))
    sent = 0; skipped = 0
    appt_col_idx = None; backup_done = False

    # Helpers inside main
    def combine(d, t):
        # reuse the parsing helpers above
        from datetime import datetime as _dt
        date_part = parse_date_cell(d); time_part = parse_time_cell(t)
        if not date_part or not time_part: return None
        return _dt.combine(date_part, time_part).replace(tzinfo=ZoneInfo(TZ_NAME))

    for idx, row in rows:
        causeno = str(row.get(COL_CAUSENO, "") or "").strip()
        ward_first = str(row.get(COL_WARDFIRST, "") or "").strip()
        ward_last = str(row.get(COL_WARDLAST, "") or "").strip()
        waddress = str(row.get(COL_WADDRESS, "") or "").strip()

        ok, reason = row_is_eligible(row, now_local)
        if not ok:
            if args.mode == "live":
                skipped += 1; logging.info(f"Skip cause {causeno or '<none>'}: {reason}")
            else:
                logging.info(f"Skip (test row invalid): {reason}")
            continue

        # Recipients
        g1_email = clean_email(row.get(COL_G1_EMAIL)) if COL_G1_EMAIL in row.index else None
        g2_email = coalesce_g2_email(row)
        to_addr = g1_email or g2_email
        cc_addrs = []
        if g1_email and g2_email:
            to_addr = g1_email; cc_addrs.append(g2_email)
        if not to_addr:
            skipped += 1; logging.info(f"Skip cause {causeno}: no guardian email present"); continue
        if (not args.dry_run) and my_email:
            cc_addrs.append(my_email)

        # Subject & bodies
        subject = f"Court Visit — {ward_last}, {ward_first} — Cause {causeno}"
        start_dt = combine(row.get(COL_VISITDATE), row.get(COL_VISITTIME))
        # Portable date formatting (strip leading zero from day)
        date_str = start_dt.strftime("%B %d, %Y").replace(" 0", " ") if start_dt else ""
        time_str = start_dt.strftime("%I:%M %p").lstrip("0") if start_dt else ""

        html_body = render_email_html(ward_first, ward_last, waddress).replace("{Date}", date_str).replace("{Time}", time_str)
        text_body = render_email_text(ward_first, ward_last, waddress).replace("{Date}", date_str).replace("{Time}", time_str)

        # Attachments
        case_folder = find_case_folder(BASE_GUARDIAN_FOLDER, causeno) if causeno else None
        arp_files, order_files = ([], [])
        if case_folder:
            arp_files, order_files = list_arp_order_files(case_folder)
        arp_path, order_path = pick_arp_order(arp_files, order_files)
        attachments = [p for p in [arp_path, order_path] if p]
        attach_names = [p.name for p in attachments]

        # Build message
        from_addr = my_email or "me"
        msg = build_mime(from_addr, to_addr, cc_addrs, subject, html_body, text_body, attachments)

        # Send guardian email
        if args.dry_run:
            logging.info(f"DRY-RUN: Would send to={to_addr}, cc={cc_addrs}, attachments={attach_names}")
        else:
            try:
                gmail_send_raw(gmail, msg)
                sent += 1
                logging.info(f"Sent confirmation email to {to_addr} (cc={cc_addrs}, attachments={attach_names})")

                # Save text copy to client folder
                if case_folder:
                    save_folder = str(case_folder)
                else:
                    save_folder = os.path.join(os.path.dirname(BASE_GUARDIAN_FOLDER), "_Correspondence_Pending")
                txt_filename = f"Confirmation Email - {ward_last}, {ward_first} - {dt.datetime.now().strftime('%Y-%m-%d')}.txt"
                cc_str = ", ".join(cc_addrs) if cc_addrs else ""
                txt_path = save_text_copy(save_folder, txt_filename, subject, text_body, to_addr, cc_str)
                logging.info(f"Saved text copy: {txt_path}")

            except Exception as e:
                skipped += 1
                logging.exception(f"Failed to send email for cause {causeno}: {e}")
                continue

        # Private FYI if Order is missing (to you only - don't notify for ARP)
        if order_path is None and (not args.dry_run) and my_email:
            try:
                fyi = EmailMessage()
                fyi["From"] = my_email; fyi["To"] = my_email
                fyi["Subject"] = f"FYI: Missing Order for {causeno}"
                fyi.set_content(f"Cause: {causeno}\nWard: {ward_first} {ward_last}\nMissing: Order\nFolder: {case_folder}\n")
                gmail_send_raw(gmail, fyi)
                logging.info(f"Sent private FYI re missing Order")
            except Exception as e:
                logging.warning(f"Failed to send FYI email: {e}")

        # Mark Y in Excel (only on real sends)
        if not args.dry_run:
            if not backup_done:
                backup_workbook(EXCEL_PATH); backup_done = True
            if appt_col_idx is None:
                appt_col_idx = ensure_appt_confirmed_column(EXCEL_PATH, COL_APPT_CONFIRMED)
            try:
                write_Y_at_row(EXCEL_PATH, appt_col_idx, idx)
                logging.info(f"Marked Appt_confirmed='Y' at row index {idx}")
            except Exception as e:
                logging.warning(f"Failed to write 'Y' for row index {idx}: {e}")

    logging.info(f"Done. Sent={sent}, Skipped={skipped}")

if __name__ == "__main__":
    import sys
    import traceback
    try:
        main()
        print("\n[OK] Confirmation Emails SUCCESS")
        sys.exit(0)
    except KeyboardInterrupt:
        logging.warning("Interrupted by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\n[FAIL] Confirmation Emails FAILED: {e}")
        traceback.print_exc()
        sys.exit(1)
