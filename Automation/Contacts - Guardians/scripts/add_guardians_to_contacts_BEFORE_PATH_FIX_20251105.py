#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
Add/Update Google Contacts for guardians so calls/texts show names.

What this does
- Creates/updates up to two contacts (Guardian1, Guardian2) per row
- Display: "<guardian cell value> for <wardfirst> <wardlast>" (exact guardian cell text; no splitting)
- Adds phones/emails if present
- Sets the guardian's own address as contact "home" address (if present)
- Adds a NOTE with ward/case + visit date/time + ward address (ward address is NOT put in Address field)
- Files into contact group "Guardians (Court Visits)" (set to "" to disable)
- Marks Excel "Contact_added" = "Y" after a successful create/update

Modes
- Test (dry run): only the last row that has at least one guardian email/phone; no Excel write in --dry-run
- Live: all rows where Contact_added is blank AND at least one guardian email/phone exists

One-time
- Enable "People API" in the same Google Cloud project as your OAuth JSON.
- OAuth JSON at C:\configlocal\API\client_secret_calendar.json

Run
  Dry run:
    py -3 "C:\\GoogleSync\\Automation\\Contacts - Guardians\\scripts\\add_guardians_to_contacts.py" --mode test_last_row --dry-run
  Test (create for one row + mark Y):
    py -3 "C:\\GoogleSync\\Automation\\Contacts - Guardians\\scripts\\add_guardians_to_contacts.py" --mode test_last_row
  Live (all eligible):
    py -3 "C:\\GoogleSync\\Automation\\Contacts - Guardians\\scripts\\add_guardians_to_contacts.py" --mode live
"""

import argparse
import datetime as dt
import json
import logging
import os
import re
import sys
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
from googleapiclient.discovery import build
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from openpyxl import load_workbook

# =========================
# CONFIG
# =========================
EXCEL_PATH = r"C:\GoogleSync\GuardianShip_App\App Data\ward_guardian_info.xlsx"

BASE_DIR = r"C:\GoogleSync\GuardianShip_App\Automation\Contacts - Guardians"
LOG_DIR = rf"{BASE_DIR}\Logs"

# Check app Config folder first, then fall back to legacy configlocal
APP_CONFIG_DIR = r"C:\GoogleSync\GuardianShip_App\Config\API"
LEGACY_CONFIG_DIR = r"C:\configlocal\API"

# Use app config if it exists, otherwise legacy
if os.path.exists(APP_CONFIG_DIR):
    CREDENTIALS_DIR = APP_CONFIG_DIR
else:
    CREDENTIALS_DIR = LEGACY_CONFIG_DIR

TZ_NAME = "America/Chicago"
CONTACT_GROUP_LABEL = "Guardians (Court Visits)"   # set to "" to skip grouping

# Excel columns (known)
COL_CAUSENO   = "causeno"
COL_WARDFIRST = "wardfirst"
COL_WARDLAST  = "wardlast"
COL_WADDRESS  = "waddress"     # Ward address (meeting location) -> goes to Notes only

COL_G1_NAME   = "guardian1"    # guardian1 cell contains the visible label (last, first, etc.)
COL_G1_EMAIL  = "gemail"
COL_G1_PHONE  = "gtele"

COL_G2_NAME           = "Guardian2"  # guardian2 cell contains the visible label
COL_G2_EMAIL_PRIMARY  = "g2email"
COL_G2_EMAIL_TYPO     = "g2eamil"
COL_G2_PHONE          = "g2tele"

# Include visit date/time in Notes (optional fields)
COL_VISITDATE = "visitdate"
COL_VISITTIME = "visittime"

COL_CONTACT_ADDED = "Contact_added"  # created/appended at end if missing

# OAuth scope
SCOPES_PEOPLE = ["https://www.googleapis.com/auth/contacts"]

# =========================
# Logging
# =========================
os.makedirs(LOG_DIR, exist_ok=True)
LOG_PATH = os.path.join(LOG_DIR, "contacts.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s: %(message)s",
    handlers=[logging.FileHandler(LOG_PATH, encoding="utf-8"), logging.StreamHandler(sys.stdout)],
)

# =========================
# Auth
# =========================
def find_client_secret_json(creds_dir: str) -> Path:
    env_path = os.environ.get("GOOGLE_OAUTH_CLIENT_JSON")
    if env_path and Path(env_path).exists():
        with open(env_path, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
                if "installed" in data and "client_id" in data["installed"]:
                    return Path(env_path)
            except Exception:
                pass
    p = Path(creds_dir)
    if not p.exists():
        raise FileNotFoundError(f"Credentials dir not found: {creds_dir}")
    for cand in list(p.glob("client_secret*.json")) + list(p.glob("credentials*.json")) + list(p.glob("*.json")):
        try:
            data = json.load(open(cand, "r", encoding="utf-8"))
            if "installed" in data and "client_id" in data["installed"]:
                return cand
        except Exception:
            continue
    raise FileNotFoundError("No OAuth desktop client JSON found.")

def get_people_service():
    token_path = Path(CREDENTIALS_DIR) / "token_people.json"
    creds = None
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES_PEOPLE)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            logging.info("Refreshing People token...")
            try:
                creds.refresh(Request())
            except Exception as e:
                # Token refresh failed (expired/revoked) - delete and re-authenticate
                logging.warning(f"Token refresh failed: {e}")
                logging.warning("Deleting expired token and starting fresh OAuth flow...")
                if token_path.exists():
                    token_path.unlink()
                creds = None  # Force re-auth below

        if not creds:
            secret_json = find_client_secret_json(CREDENTIALS_DIR)
            logging.info(f"Starting OAuth for People API using: {secret_json}")
            flow = InstalledAppFlow.from_client_secrets_file(str(secret_json), SCOPES_PEOPLE)
            creds = flow.run_local_server(port=0)

        open(token_path, "w", encoding="utf-8").write(creds.to_json())
        logging.info(f"Saved People token: {token_path}")
    return build("people", "v1", credentials=creds, cache_discovery=False)

# =========================
# Excel helpers
# =========================
def read_df(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, engine="openpyxl")
    df.columns = [c.strip() for c in df.columns]
    return df

def ensure_contact_added_column(wb_path: str, header_name: str) -> int:
    wb = load_workbook(wb_path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if header_name in headers:
        col_idx = headers.index(header_name) + 1
    else:
        col_idx = len(headers) + 1
        ws.cell(row=1, column=col_idx, value=header_name)
    wb.save(wb_path); wb.close()
    return col_idx

def write_Y_at_row(wb_path: str, col_idx: int, df_index: int):
    wb = load_workbook(wb_path)
    ws = wb.active
    excel_row = df_index + 2  # header row is 1
    ws.cell(row=excel_row, column=col_idx, value="Y")
    wb.save(wb_path); wb.close()

# =========================
# Date/time parsing (aligned with Scripts #1/#2)
# =========================
def parse_date_cell(d):
    if pd.isna(d): return None
    if isinstance(d, pd.Timestamp): return d.date()
    if isinstance(d, dt.datetime): return d.date()
    if isinstance(d, dt.date): return d
    if isinstance(d, (int, float)):
        try: return pd.to_datetime(d, unit="d", origin="1899-12-30").date()
        except Exception: pass
    s = str(d).strip()
    if not s: return None
    dt_obj = pd.to_datetime(s, errors="coerce", dayfirst=False)
    if pd.isna(dt_obj): return None
    return dt_obj.date()

def parse_time_cell(t):
    if pd.isna(t): return None
    if isinstance(t, pd.Timestamp): return t.time()
    if isinstance(t, dt.datetime): return t.time()
    if isinstance(t, dt.time): return t
    if isinstance(t, (int, float)):
        try: return pd.to_datetime(t, unit="d", origin="1899-12-30").time()
        except Exception: pass
    s = str(t).strip().lower()
    if not s or s in {"tbd", "na", "n/a", "none"}: return None
    s = s.replace(" ", "").replace(".", ":")
    m = re.fullmatch(r"(\d{1,2})(?::?(\d{2}))?(am|pm)?", s)
    if m:
        hh = int(m.group(1)); mm = int(m.group(2) or 0); ampm = m.group(3)
        if ampm == "pm" and hh < 12: hh += 12
        if ampm == "am" and hh == 12: hh = 0
        if 0 <= hh <= 23 and 0 <= mm <= 59: return dt.time(hour=hh, minute=mm)
    parsed = pd.to_datetime(str(t), errors="coerce")
    if pd.isna(parsed): return None
    return parsed.time()

def combine_date_time_local(d, t, tzname: str):
    date_part = parse_date_cell(d); time_part = parse_time_cell(t)
    if not date_part or not time_part: return None
    return dt.datetime.combine(date_part, time_part).replace(tzinfo=ZoneInfo(tzname))

def format_visit_dt(start_dt: dt.datetime | None) -> tuple[str, str]:
    if not start_dt: return "", ""
    date_str = start_dt.strftime("%B %d, %Y").replace(" 0", " ")
    time_str = start_dt.strftime("%I:%M %p").lstrip("0")
    return date_str, time_str

# =========================
# Cleaners (fix "nan", "none", blanks) + label passthrough
# =========================
def _is_blank(v):
    if v is None:
        return True
    s = str(v).strip().lower()
    return s in {"", "none", "n/a", "na", "null", "nil", "nan"}

def clean_email(v):
    if _is_blank(v):
        return None
    s = str(v).strip()
    return s if "@" in s else None

def clean_phone(v):
    if _is_blank(v):
        return None
    return str(v).strip()

def exact_label(v):
    """Return the guardian cell EXACTLY as typed (trimmed)."""
    if _is_blank(v):
        return ""
    return str(v).strip()

def coalesce_g2_email(row: pd.Series) -> str | None:
    primary = clean_email(row.get(COL_G2_EMAIL_PRIMARY))
    if primary:
        return primary
    typo = clean_email(row.get(COL_G2_EMAIL_TYPO))
    return typo

# Try to pull guardian-specific addresses from a variety of likely column names
def get_guardian_address(row: pd.Series, which: int) -> str | None:
    # Common variants; add more if your sheet uses a different header
    if which == 1:
        candidates = ["gaddress", "g1address", "guardian1_address", "guardian1address", "guardianaddress", "gaddress1"]
    else:
        candidates = ["g2address", "guardian2_address", "guardian2address", "gaddress2"]
    for col in candidates:
        if col in row.index:
            addr = str(row.get(col) or "").strip()
            if not _is_blank(addr):
                return addr
    return None

# =========================
# People API helpers
# =========================
def build_display_and_note(guardian_label, ward_first, ward_last, causeno, start_dt, ward_address):
    """Display uses guardian cell EXACTLY as typed; Notes hold details."""
    glabel = (guardian_label or "").strip()
    ward_first = (ward_first or "").strip()
    ward_last = (ward_last or "").strip()
    cause = (causeno or "").strip()

    display = f"{glabel or 'Guardian'} for {ward_first} {ward_last}"

    date_str, time_str = format_visit_dt(start_dt)
    note_lines = [
        f"Guardian for {ward_first} {ward_last}",
        f"Cause: {cause}",
    ]
    if date_str and time_str:
        note_lines.append(f"Visit: {date_str} at {time_str} ({TZ_NAME})")
    if ward_address:
        note_lines.append(f"Ward address: {ward_address}")
    note = "\n".join(note_lines)

    return display, note

def get_or_create_group(service, label: str) -> str | None:
    if not label:
        return None
    try:
        groups = service.contactGroups().list(pageSize=200).execute().get("contactGroups", [])
        for g in groups:
            if g.get("name") == label:
                return g.get("resourceName")
        newg = service.contactGroups().create(body={"contactGroup": {"name": label}}).execute()
        return newg.get("resourceName")
    except Exception as e:
        logging.warning(f"Could not get/create contact group '{label}': {e}")
        return None

def search_existing(service, query: str):
    try:
        resp = service.people().searchContacts(
            query=query,
            pageSize=5,
            readMask="names,emailAddresses,phoneNumbers,biographies,addresses,memberships"
        ).execute()
        people = resp.get("results", [])
        if not people:
            return None
        return people[0].get("person")
    except Exception:
        return None

def person_patch_for_update(existing, add_body, group_resource):
    """
    Build a patch that actually updates fields:
    - Always update names (so display changes take effect).
    - Append new emails/phones (dedup by value).
    - Merge Notes (biographies).
    - Preserve existing addresses unless we're adding a new one.
    - Ensure membership in the contact group.
    """
    person = {"resourceName": existing.get("resourceName"), "etag": existing.get("etag")}

    # 1) Names — overwrite so the visible name changes
    if add_body.get("names"):
        person["names"] = add_body["names"]
    elif existing.get("names"):
        person["names"] = existing["names"]

    # 2) Emails — add missing
    existing_emails = {e.get("value") for e in existing.get("emailAddresses", [])}
    new_emails = [e for e in add_body.get("emailAddresses", []) if e.get("value") not in existing_emails]
    person["emailAddresses"] = existing.get("emailAddresses", []) + new_emails

    # 3) Phones — add missing
    existing_phones = {p.get("value") for p in existing.get("phoneNumbers", [])}
    new_phones = [p for p in add_body.get("phoneNumbers", []) if p.get("value") not in existing_phones]
    person["phoneNumbers"] = existing.get("phoneNumbers", []) + new_phones

    # 4) Notes — merge ours in if not already present
    existing_note = ""
    if existing.get("biographies"):
        existing_note = existing["biographies"][0].get("value", "")
    our_note = add_body.get("biographies", [{}])[0].get("value", "")
    if our_note and our_note not in existing_note:
        merged_note = (existing_note + "\n" + our_note).strip() if existing_note else our_note
        person["biographies"] = [{"value": merged_note}]
    elif existing.get("biographies"):
        person["biographies"] = existing["biographies"]

    # 5) Address — keep existing; if none, add ours (home)
    if existing.get("addresses"):
        person["addresses"] = existing["addresses"]
    else:
        if add_body.get("addresses"):
            person["addresses"] = add_body["addresses"]

    # 6) Group membership — ensure it's present
    memberships = existing.get("memberships", []) if existing.get("memberships") else []
    if group_resource:
        has_group = any(
            m.get("contactGroupMembership", {}).get("contactGroupResourceName") == group_resource
            for m in memberships
        )
        if not has_group:
            memberships.append({"contactGroupMembership": {"contactGroupResourceName": group_resource}})
    if memberships:
        person["memberships"] = memberships

    return person


def create_or_update_contact(service, guardian_label, email, phone, ward_first, ward_last, causeno, ward_address, contact_address, start_dt, group_resource, dry_run=False):
    display, note = build_display_and_note(guardian_label, ward_first, ward_last, causeno, start_dt, ward_address)

    # Build add_body (contact fields) — NO splitting; show display string as visible name
    names = [{"givenName": display}]
    email_addrs = [{"value": email}] if email else []
    phones = [{"value": phone, "type": "mobile"}] if phone else []
    addresses = [{"formattedValue": contact_address, "type": "home"}] if contact_address else []
    bios = [{"value": note}] if note else []
    memberships = [{"contactGroupMembership": {"contactGroupResourceName": group_resource}}] if group_resource else []

    add_body = {
        "names": names,
        "emailAddresses": email_addrs,
        "phoneNumbers": phones,
        "addresses": addresses,
        "biographies": bios,
    }
    if memberships:
        add_body["memberships"] = memberships

    # Dedupe search: email > phone > display label
    existing = None
    if email:
        existing = search_existing(service, email)
    if (existing is None) and phone:
        existing = search_existing(service, phone)
    if (existing is None) and display:
        existing = search_existing(service, display)

    if dry_run:
        action = "Would UPDATE" if existing else "Would CREATE"
        logging.info(f"{action} contact: '{display}' email={email} phone={phone} address={contact_address}")
        return True

    if existing:
        try:
            patch = person_patch_for_update(existing, add_body, group_resource)
            service.people().updateContact(
                resourceName=patch["resourceName"],
                updatePersonFields="names,emailAddresses,phoneNumbers,biographies,addresses,memberships",
                body=patch,
            ).execute()
            logging.info(f"Updated contact: '{display}'")
            return True
        except Exception as e:
            logging.warning(f"Failed to update contact '{display}': {e}")
            return False
    else:
        try:
            service.people().createContact(body=add_body).execute()
            logging.info(f"Created contact: '{display}'")
            return True
        except Exception as e:
            logging.warning(f"Failed to create contact '{display}': {e}")
            return False

# =========================
# Main
# =========================
def main():
    parser = argparse.ArgumentParser(description="Create/Update guardian contacts in Google Contacts.")
    parser.add_argument("--mode", choices=["test_last_row", "live"], required=True)
    parser.add_argument("--dry-run", action="store_true", help="Plan only; do not hit People API; do not write Excel")
    args = parser.parse_args()

    logging.info("=== Start: add_guardians_to_contacts ===")
    logging.info(f"Excel: {EXCEL_PATH}")
    logging.info(f"Mode: {args.mode}  Dry-run: {args.dry_run}")

    # Load data
    try:
        df = read_df(EXCEL_PATH)
    except Exception as e:
        logging.exception(f"Failed to read workbook: {e}"); sys.exit(2)

    required = [COL_CAUSENO, COL_WARDFIRST, COL_WARDLAST]
    missing = [c for c in required if c not in df.columns]
    if missing:
        logging.error(f"Missing expected columns: {missing}"); sys.exit(3)

    # Helper: has any guardian contact method (using cleaners)
    def row_has_any_contact(r):
        g1_email = clean_email(r.get(COL_G1_EMAIL))
        g1_phone = clean_phone(r.get(COL_G1_PHONE))
        g2_email = coalesce_g2_email(r)
        g2_phone = clean_phone(r.get(COL_G2_PHONE))
        return any([g1_email, g1_phone, g2_email, g2_phone])

    # Pick rows
    if args.mode == "test_last_row":
        candidates = df[df.apply(row_has_any_contact, axis=1)]
        if candidates.empty:
            logging.info("No rows with guardian email/phone found. Nothing to test.")
            return
        idx = candidates.index[-1]
        rows = [(idx, df.loc[idx])]
        logging.info(f"Test mode: using last row with any guardian email/phone index={idx}")
    else:
        if COL_CONTACT_ADDED in df.columns:
            eligible = df[(df[COL_CONTACT_ADDED].isna()) | (df[COL_CONTACT_ADDED].astype(str).str.strip() == "")]
        else:
            eligible = df
        eligible = eligible[eligible.apply(row_has_any_contact, axis=1)]
        rows = list(eligible.iterrows())
        if not rows:
            logging.info("No eligible rows to process (Contact_added not blank or no guardian contact info).")
            return

    # People service (only if not dry-run)
    people = None
    group_resource = None
    if not args.dry_run:
        people = get_people_service()
        if CONTACT_GROUP_LABEL:
            group_resource = get_or_create_group(people, CONTACT_GROUP_LABEL)
            if group_resource:
                logging.info(f"Using contact group: {CONTACT_GROUP_LABEL} ({group_resource})")

    # Prepare Excel writing
    contact_added_col_idx = None
    processed_rows = 0

    for idx, row in rows:
        causeno    = str(row.get(COL_CAUSENO, "") or "").strip()
        ward_first = str(row.get(COL_WARDFIRST, "") or "").strip()
        ward_last  = str(row.get(COL_WARDLAST, "") or "").strip()
        ward_addr  = str(row.get(COL_WADDRESS, "") or "").strip()

        # Visit datetime (optional, for Notes)
        start_dt = None
        if COL_VISITDATE in df.columns and COL_VISITTIME in df.columns:
            start_dt = combine_date_time_local(row.get(COL_VISITDATE), row.get(COL_VISITTIME), TZ_NAME)

        # Guardian 1 (label + own address)
        g1_label  = exact_label(row.get(COL_G1_NAME))
        g1_email  = clean_email(row.get(COL_G1_EMAIL))
        g1_phone  = clean_phone(row.get(COL_G1_PHONE))
        g1_addr   = get_guardian_address(row, which=1)

        # Guardian 2 (label + own address)
        g2_label  = exact_label(row.get(COL_G2_NAME))
        g2_email  = coalesce_g2_email(row)
        g2_phone  = clean_phone(row.get(COL_G2_PHONE))
        g2_addr   = get_guardian_address(row, which=2)

        any_success = False
        if g1_email or g1_phone:
            ok = create_or_update_contact(
                people, g1_label, g1_email, g1_phone,
                ward_first, ward_last, causeno,
                ward_addr, g1_addr, start_dt, group_resource,
                dry_run=args.dry_run
            )
            any_success = any_success or ok
        if g2_email or g2_phone:
            ok = create_or_update_contact(
                people, g2_label, g2_email, g2_phone,
                ward_first, ward_last, causeno,
                ward_addr, g2_addr, start_dt, group_resource,
                dry_run=args.dry_run
            )
            any_success = any_success or ok

        if any_success and (not args.dry_run):
            if contact_added_col_idx is None:
                contact_added_col_idx = ensure_contact_added_column(EXCEL_PATH, COL_CONTACT_ADDED)
            try:
                write_Y_at_row(EXCEL_PATH, contact_added_col_idx, idx)
                processed_rows += 1
                logging.info(f"Marked {COL_CONTACT_ADDED}='Y' at row index {idx}")
            except Exception as e:
                logging.warning(f"Failed to write 'Y' for row index {idx}: {e}")

    logging.info(f"Done. Rows marked Y={processed_rows}")

if __name__ == "__main__":
    import sys
    import traceback
    try:
        main()
        print("\n[OK] Add Contacts SUCCESS")
        sys.exit(0)
    except KeyboardInterrupt:
        logging.warning("Interrupted by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\n[FAIL] Add Contacts FAILED: {e}")
        traceback.print_exc()
        sys.exit(1)
