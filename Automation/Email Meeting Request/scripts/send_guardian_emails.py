#!/usr/bin/env python3
# -*- coding: utf-8 -*-

r"""
send_guardian_emails.py (Sheet1-only, SAFE WRITE)
- Trigger to send: datesubmitted is BLANK AND emailsent is BLANK/MISSING
- Default behavior: READ-ONLY (no Excel writes)
- If you pass --confirm-write, it writes ONLY the 'emailsent' column (appended at the end if missing)
- Builds email from approved template; sends to Guardian1 (To) and Guardian2 (Cc) when valid
- Saves a .txt copy of each email in the ward's folder (matched by causeno)
- Logs a CSV run file in \runs\ next to the workbook
"""

import os
import sys
import time
import argparse
import re
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta, MO
import pandas as pd
from pathlib import Path

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from email.mime.text import MIMEText
import base64

from openpyxl import load_workbook  # surgical Excel write (only 'emailsent')

# Dynamic path management - works from any installation location
_script_dir = Path(__file__).parent.parent.parent.parent  # Go up to app root

try:
    sys.path.insert(0, str(_script_dir / "Scripts"))
    from app_paths import get_app_paths

    _app_paths = get_app_paths(str(_script_dir))
    WORKBOOK_PATH = str(_app_paths.EXCEL_PATH)
    GUARDIAN_BASE = str(_app_paths.APP_ROOT)
    NEW_CLIENTS_DIR = str(_app_paths.NEW_CLIENTS_DIR)
    _CONFIG_DIR = _app_paths.CONFIG_DIR
except Exception:
    # Fallback to hardcoded paths
    WORKBOOK_PATH = r"C:\GoogleSync\GuardianShip_App\App Data\ward_guardian_info.xlsx"
    GUARDIAN_BASE = r"C:\GoogleSync\GuardianShip_App"
    NEW_CLIENTS_DIR = r"C:\GoogleSync\GuardianShip_App\New Clients"
    _CONFIG_DIR = Path(r"C:\GoogleSync\GuardianShip_App\Config")

SHEET1_NAME   = "Sheet1"
TIMEZONE      = "America/Chicago"

DEFAULT_MODE   = "draft"   # "draft" or "send" (default safe)
THROTTLE_SEC   = 2.0
LIMIT_DEFAULT  = 999999
EMAIL_SENT_COL = "emailsent"  # created if missing; updated only with --confirm-write

SUBJECT_TEMPLATE = "Court Visitor meeting for {WardFirst} {WardLast} (Travis County)"

BODY_TEMPLATE = """Hello {GuardianName},

I’ve been assigned by the Travis County Probate Court as a Court Visitor for {WardFirst} {WardLast}. Court Visitors meet wards where they live to confirm their needs are being met. You can read more about the program here: https://www.traviscountytx.gov/probate/court-vistor.

I would like to schedule a visit {NextWeekPhrase} and I’m holding {PreferredDays} as first options. You’re welcome to attend (not required). If you can’t attend, a brief phone call beforehand works fine.

A few quick items to reply with:
1) Which day/time works best ({PreferredDaysShort}), and who (if anyone) will be present.
2) Exact address of the ward’s residence (street number, city).
3) Any entry instructions—for example, gated community or private gate codes, buzzer instructions, or special directions if the house number is hard to see. Landmarks or parking tips are very helpful. Please also let me know if there are protective pets that should be secured.

Thank you in advance for your time and help. I look forward to meeting you and the ward.

Warm regards,
May Ehresman
Court Visitor (Travis County)
317-339-9963 | mayehres@gmail.com
"""

# Exact Sheet1 headers
COLS = {
    "ward_first": "wardfirst",
    "ward_last": "wardlast",
    "guardian_name": "guardian1",
    "guardian_email": "gemail",
    "guardian2_email": "g2eamil",   # keep sheet's typo exactly
    "cause_no": "causeno",
    "date_submitted": "datesubmitted",  # trigger; NEVER modify
}

# Gmail OAuth - check app Config folder first, then fall back to configlocal
SCOPES = ["https://www.googleapis.com/auth/gmail.modify"]

# Try app Config folder first
APP_CONFIG_DIR = str(_CONFIG_DIR / "API")
APP_CLIENT_SECRETS = os.path.join(APP_CONFIG_DIR, "gmail_oauth_client.json")
APP_TOKEN_PATH = os.path.join(APP_CONFIG_DIR, "gmail_token.json")

# Fall back to legacy configlocal
LEGACY_CLIENT_SECRETS = r"C:\configlocal\API\gmail_oauth_client.json"
LEGACY_TOKEN_PATH = r"C:\configlocal\API\gmail_token.json"

# Use app config if exists, otherwise legacy
if os.path.exists(APP_CLIENT_SECRETS):
    CLIENT_SECRETS = APP_CLIENT_SECRETS
    TOKEN_PATH = APP_TOKEN_PATH
else:
    CLIENT_SECRETS = LEGACY_CLIENT_SECRETS
    TOKEN_PATH = LEGACY_TOKEN_PATH
# ======================================

# Email cleaning/validation
INVALID_MARKERS = {"none", "n/a", "na", "null", "nan", "(none)", "-"}

EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-']+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")

def clean_email(value) -> str:
    if value is None:
        return ""
    s = str(value)
    s = s.replace("\u00A0", " ").replace("\u200B", "").strip()
    s = s.strip().strip(",;")
    s = s.lower()
    if s in INVALID_MARKERS:
        return ""
    return s

def is_valid_email(s: str) -> bool:
    return bool(s and EMAIL_RE.match(s))

def load_sheet1(path, sheet_name):
    df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    return df

def choose_week_and_days():
    today = date.today()
    next_mon = today + relativedelta(weekday=MO(+1))
    next_sun = next_mon + timedelta(days=6)
    default_week = f"{next_mon.isoformat()}..{next_sun.isoformat()}"
    default_days = "Wed,Thu"

    print(f"\nProposed week: {default_week}  (Mon..Sun)")
    week_in = input("Use this week? Press Enter, or enter YYYY-MM-DD..YYYY-MM-DD: ").strip()
    if not week_in:
        week_in = default_week

    days_in = input(f"Preferred days (comma) [default {default_days}]: ").strip()
    if not days_in:
        days_in = default_days

    try:
        start_s, end_s = week_in.split("..")
        start_d = date.fromisoformat(start_s)
        end_d = date.fromisoformat(end_s)
        week_phrase = "next week" if (start_d == next_mon and end_d == next_sun) \
                      else f"the week of {start_d.strftime('%b %d')}–{end_d.strftime('%b %d')}"
    except Exception:
        week_phrase = "next week"

    days_list = [d.strip() for d in days_in.split(",") if d.strip()]
    preferred_days_short = "/".join(days_list)
    preferred_days_phrase = ", ".join(days_list)

    return week_phrase, preferred_days_phrase, preferred_days_short

def gmail_service():
    from pathlib import Path
    creds = None
    token_path = Path(TOKEN_PATH)

    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception as e:
                # Token refresh failed (expired/revoked) - delete and re-authenticate
                print(f"Token refresh failed: {e}")
                print("Deleting expired token and starting fresh OAuth flow...")
                if token_path.exists():
                    token_path.unlink()
                creds = None  # Force re-auth below

        if not creds:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRETS, SCOPES)
            creds = flow.run_local_server(port=0)

        with open(TOKEN_PATH, "w", encoding="utf-8") as f:
            f.write(creds.to_json())
    return build("gmail", "v1", credentials=creds)

def build_message(to_email, subject, body, cc_email=None):
    msg = MIMEText(body, "plain", "utf-8")
    msg["To"] = to_email
    if cc_email:
        msg["Cc"] = cc_email
    msg["Subject"] = subject
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
    return {"raw": raw}

def find_case_folder(base_dir, cause_no):
    try:
        for name in os.listdir(base_dir):
            full = os.path.join(base_dir, name)
            if os.path.isdir(full) and str(cause_no) in name:
                return full
    except FileNotFoundError:
        pass
    return None

def save_text_copy(folder, filename, subject, body, to_email, cc_email):
    os.makedirs(folder, exist_ok=True)
    out_path = os.path.join(folder, filename)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"TO: {to_email}\n")
        if cc_email:
            f.write(f"CC: {cc_email}\n")
        f.write(f"SUBJECT: {subject}\n\n")
        f.write(body)
    return out_path

def update_emailsent_openpyxl(workbook_path, sheet_name, emailsent_col_name, causes_marked, today_str):
    """
    Surgical cell-level write:
    - Ensures 'emailsent' header exists at the end (if missing)
    - Stamps today_str only for rows whose causeno is in causes_marked
    - Touches nothing else
    """
    wb = load_workbook(workbook_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]

    # Map headers (row 1)
    header_pos = {}
    for c in range(1, ws.max_column + 1):
        header_val = ws.cell(row=1, column=c).value
        key = (str(header_val).strip() if header_val is not None else "")
        header_pos[key] = c

    # Required headers
    if COLS["cause_no"] not in header_pos:
        raise RuntimeError(f"Header '{COLS['cause_no']}' not found in row 1.")

    # Ensure emailsent column exists (append at end if missing)
    if emailsent_col_name not in header_pos:
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col, value=emailsent_col_name)
        header_pos[emailsent_col_name] = new_col

    col_cause    = header_pos[COLS["cause_no"]]
    col_emailsent= header_pos[emailsent_col_name]

    # Build cause_no -> row indices
    cause_to_rows = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_cause).value
        key = (str(v).strip() if v is not None else "")
        if key:
            cause_to_rows.setdefault(key, []).append(r)

    # Stamp emailsent for processed causes
    for cause in causes_marked:
        for r in cause_to_rows.get(str(cause).strip(), []):
            ws.cell(row=r, column=col_emailsent, value=today_str)

    wb.save(workbook_path)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default=WORKBOOK_PATH)
    ap.add_argument("--mode", choices=["draft","send"], default=DEFAULT_MODE)
    ap.add_argument("--limit", type=int, default=LIMIT_DEFAULT)
    ap.add_argument("--throttle", type=float, default=THROTTLE_SEC)
    ap.add_argument("--confirm-write", action="store_true",
                    help="If set, stamp 'emailsent' (YYYY-MM-DD). Otherwise READ-ONLY (no Excel writes).")
    ap.add_argument("--week", type=str, default=None,
                    help="Week range in format YYYY-MM-DD..YYYY-MM-DD (e.g., 2025-11-03..2025-11-09)")
    ap.add_argument("--days", type=str, default=None,
                    help="Preferred days comma-separated (e.g., Wed,Thu)")
    args = ap.parse_args()

    # Load Sheet1
    df = load_sheet1(args.workbook, SHEET1_NAME)

    # Guard: datesubmitted must exist (trigger)
    if COLS["date_submitted"] not in df.columns:
        print(f"ERROR: '{COLS['date_submitted']}' column not found on Sheet1.")
        sys.exit(1)

    # Build eligibility: datesubmitted blank AND emailsent blank/missing
    if EMAIL_SENT_COL not in df.columns:
        df[EMAIL_SENT_COL] = ""  # local df only; NOT writing unless --confirm-write

    elig = df[
        (df[COLS["date_submitted"]].isna() | (df[COLS["date_submitted"]].astype(str).str.strip() == "")) &
        (df[EMAIL_SENT_COL].isna() | (df[EMAIL_SENT_COL].astype(str).str.strip() == ""))
    ].copy()

    if elig.empty:
        print("No eligible rows: need datesubmitted blank AND emailsent blank.")
        return

    # Get week and days from command line or prompt interactively
    if args.week and args.days:
        # Use command-line arguments (GUI mode)
        week_in = args.week
        days_in = args.days

        try:
            start_s, end_s = week_in.split("..")
            start_d = date.fromisoformat(start_s)
            end_d = date.fromisoformat(end_s)

            # Calculate default next week for comparison
            today = date.today()
            next_mon = today + relativedelta(weekday=MO(+1))
            next_sun = next_mon + timedelta(days=6)

            week_phrase = "next week" if (start_d == next_mon and end_d == next_sun) \
                          else f"the week of {start_d.strftime('%b %d')}-{end_d.strftime('%b %d')}"
        except Exception:
            week_phrase = "next week"

        days_list = [d.strip() for d in days_in.split(",") if d.strip()]
        preferred_days_short = "/".join(days_list)
        preferred_days_phrase = ", ".join(days_list)

        print(f"Using week: {week_in}")
        print(f"Using preferred days: {days_in}")
    else:
        # Interactive mode (console)
        week_phrase, preferred_days_phrase, preferred_days_short = choose_week_and_days()

    # Gmail auth
    try:
        service = gmail_service()
    except Exception as e:
        print("Gmail auth failed:", e)
        sys.exit(1)

    run_ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    log_rows = []
    sent = 0
    processed_causes = []  # causes to mark in emailsent (only if --confirm-write)

    for _, row in elig.iterrows():
        if sent >= args.limit:
            break

        cause      = str(row.get(COLS["cause_no"], "")).strip()
        ward_first = str(row.get(COLS["ward_first"], "")).strip()
        ward_last  = str(row.get(COLS["ward_last"], "")).strip()
        g_name     = str(row.get(COLS["guardian_name"], "")).strip() or "Guardian"

        primary_raw   = row.get(COLS["guardian_email"], "")
        secondary_raw = row.get(COLS["guardian2_email"], "") if COLS["guardian2_email"] in df.columns else ""

        g_email  = clean_email(primary_raw)
        g2_email = clean_email(secondary_raw)

        valid_primary   = is_valid_email(g_email)
        valid_secondary = is_valid_email(g2_email)

        # To / Cc selection
        to_email, cc_email = "", None
        if valid_primary:
            to_email = g_email
            if valid_secondary:
                cc_email = g2_email
        elif valid_secondary:
            to_email = g2_email
        else:
            status = f"SKIP:no-valid-email (primary='{primary_raw}', secondary='{secondary_raw}')"
            log_rows.append((cause, ward_last, ward_first, "", status, ""))
            continue

        subject = SUBJECT_TEMPLATE.format(WardFirst=ward_first, WardLast=ward_last)
        body = BODY_TEMPLATE.format(
            GuardianName=g_name,
            WardFirst=ward_first,
            WardLast=ward_last,
            NextWeekPhrase=week_phrase,
            PreferredDays=preferred_days_phrase,
            PreferredDaysShort=preferred_days_short,
        )

        try:
            message = build_message(to_email, subject, body, cc_email=cc_email)
            if args.mode == "draft":
                service.users().drafts().create(userId="me", body={"message": message}).execute()
                status = "DRAFT"
            else:
                service.users().messages().send(userId="me", body=message).execute()
                status = "SENT"

            # Save text copy
            case_folder = find_case_folder(NEW_CLIENTS_DIR, cause) or os.path.join(GUARDIAN_BASE, "_Correspondence_Pending")
            fname = f"Meeting Email - {ward_last}, {ward_first} - {datetime.now().strftime('%Y-%m-%d')}.txt"
            saved_path = save_text_copy(case_folder, fname, subject, body, to_email, cc_email)

            log_rows.append((cause, ward_last, ward_first, to_email, status, saved_path))
            processed_causes.append(cause)
            sent += 1
            time.sleep(args.throttle)

        except HttpError as he:
            log_rows.append((cause, ward_last, ward_first, to_email, f"ERROR:{he}", ""))
        except Exception as e:
            log_rows.append((cause, ward_last, ward_first, to_email, f"ERROR:{e}", ""))

    # READ-ONLY by default. Only stamp emailsent if --confirm-write is provided.
    if args.confirm_write and processed_causes:
        try:
            today_str = date.today().strftime("%Y-%m-%d")
            update_emailsent_openpyxl(args.workbook, SHEET1_NAME, EMAIL_SENT_COL, processed_causes, today_str)
            print(f"[WRITE] emailsent updated for {len(processed_causes)} case(s).")
        except Exception as e:
            print("WARNING: Could not write 'emailsent' via openpyxl:", e)
    else:
        print("[READ-ONLY] Skipped updating 'emailsent' (run with --confirm-write to enable).")

    # Log CSV next to workbook
    log_dir = os.path.join(os.path.dirname(args.workbook), "runs")
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, f"meeting_email_log_{run_ts}.csv")
    pd.DataFrame(
        log_rows,
        columns=["causeno","wardlast","wardfirst","to_email","status","saved_copy_path"]
    ).to_csv(log_path, index=False, encoding="utf-8")

    print(f"Done. {sent} item(s) processed. Log: {log_path}")

if __name__ == "__main__":
    import sys
    import traceback
    try:
        main()
        print("\n[OK] Meeting Requests SUCCESS")
        sys.exit(0)
    except Exception as e:
        print(f"\n[FAIL] Meeting Requests FAILED: {e}")
        traceback.print_exc()
        sys.exit(1)
