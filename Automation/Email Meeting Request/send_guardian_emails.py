#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
send_guardian_emails.py  (Sheet1-only)
- Reads Excel Sheet1, finds rows with datesubmitted blank
- Prompts you to confirm the week/days (defaults: next week; Wed/Thu)
- Builds email from the approved template
- Sends via Gmail API (draft or send)
- Updates Sheet1[datesubmitted] after a successful send
- Saves a .txt copy of the email into the person's folder under:
  C:\GoogleSync\Guardianship Files\*{causeno}*

Prereqs (one-time):
  py -3 -m pip install pandas openpyxl google-auth-oauthlib google-api-python-client python-dateutil
  Enable Gmail API in your Google Cloud project
  Create OAuth Desktop Client, download JSON to: C:\configlocal\API\gmail_oauth_client.json
On first run, a browser will open; your token will be saved to C:\configlocal\API\gmail_token.json
"""

import os
import sys
import time
import argparse
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
import pandas as pd

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from email.mime.text import MIMEText
import base64

# ======= CONFIG YOU MAY EDIT =======
WORKBOOK_PATH = r"C:\GoogleSync\Guardianship Files\data files\ward_guardian_info.xlsx"
GUARDIAN_BASE = r"C:\GoogleSync\Guardianship Files"
SHEET1_NAME = "Sheet1"    # Status + details + datesubmitted are all here now
TIMEZONE = "America/Chicago"

# Run behavior
DEFAULT_MODE = "send"     # "draft" or "send"
THROTTLE_SEC = 2.0
LIMIT_DEFAULT = 999999

# Email templates (approved Version A)
SUBJECT_TEMPLATE = "Court Visitor meeting for {WardFirst} {WardLast} (Travis County)"

BODY_TEMPLATE = """Hello {GuardianName},

I’ve been assigned by the Travis County Probate Court as a Court Visitor for {WardFirst} {WardLast}. Court Visitors meet wards where they live to confirm their needs are being met. You can read more about the program here: https://www.traviscountytx.gov/probate/court-vistor.

I would like to schedule a visit {NextWeekPhrase} and I’m holding {PreferredDays} as first options. You’re welcome to attend (not required). If you can’t attend, a brief phone call beforehand works fine.

A few quick items to reply with:
1) Which day/time works best ({PreferredDaysShort}), and who (if anyone) will be present.
2) Exact address of the ward’s residence (street number, city).
3) Any entry instructions—for example, gated community or private gate codes, buzzer instructions, or special directions if the house number is hard to see. Landmarks or parking tips are very helpful. Please also let me know if there are protective pets that should be secured.

Thank you in advance for your time and help. I look forward to meeting you and the ward.

Warm regards,
May Ehresman
Court Visitor (Travis County)
317-339-9963 | mayehres@gmail.com
"""

# Column names (exactly as in your Sheet1)
COLS = {
    "ward_first": "wardfirst",
    "ward_last": "wardlast",
    "guardian_name": "guardian1",
    "guardian_email": "gemail",
    "guardian2_email": "g2eamil",   # keep the sheet's typo exactly
    "cause_no": "causeno",
    "date_submitted": "datesubmitted",
}

# Gmail API scopes + OAuth file locations
SCOPES = ["https://www.googleapis.com/auth/gmail.modify"]
CLIENT_SECRETS = r"C:\configlocal\API\gmail_oauth_client.json"  # downloaded from Cloud Console
TOKEN_PATH     = r"C:\configlocal\API\gmail_token.json"         # created on first auth
# ===================================


def load_sheet1(path, sheet_name):
    df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def choose_week_and_days():
    today = date.today()
    # Next Monday .. next Sunday by default
    next_mon = today + relativedelta(weekday=relativedelta.MO(+1))
    next_sun = next_mon + timedelta(days=6)
    default_week = f"{next_mon.isoformat()}..{next_sun.isoformat()}"
    default_days = "Wed,Thu"

    print(f"\nProposed week: {default_week}  (Mon..Sun)")
    week_in = input("Use this week? Press Enter, or enter YYYY-MM-DD..YYYY-MM-DD: ").strip()
    if not week_in:
        week_in = default_week

    days_in = input(f"Preferred days (comma) [default {default_days}]: ").strip()
    if not days_in:
        days_in = default_days

    # Phrase building
    try:
        start_s, end_s = week_in.split("..")
        start_d = date.fromisoformat(start_s)
        end_d = date.fromisoformat(end_s)
        # Default wording is "next week"; if user changed week, show explicit dates
        next_default_mon = next_mon
        next_default_sun = next_sun
        if (start_d != next_default_mon) or (end_d != next_default_sun):
            week_phrase = f"the week of {start_d.strftime('%b %d')}–{end_d.strftime('%b %d')}"
        else:
            week_phrase = "next week"
    except Exception:
        week_phrase = "next week"

    days_list = [d.strip() for d in days_in.split(",") if d.strip()]
    preferred_days_short = "/".join(days_list)
    preferred_days_phrase = ", ".join(days_list)

    return week_phrase, preferred_days_phrase, preferred_days_short


def gmail_service():
    """
    Loads credentials from TOKEN_PATH if present; otherwise runs OAuth using
    CLIENT_SECRETS and saves the token to TOKEN_PATH. Returns Gmail API client.
    """
    creds = None
    if os.path.exists(TOKEN_PATH):
        creds = Credentials.from_authorized_user_file(TOKEN_PATH, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception:
                creds = None
        if not creds:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRETS, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_PATH, "w", encoding="utf-8") as f:
            f.write(creds.to_json())

    return build("gmail", "v1", credentials=creds)


def build_message(to_email, subject, body, cc_email=None):
    msg = MIMEText(body, "plain", "utf-8")
    msg["To"] = to_email
    if cc_email:
        msg["Cc"] = cc_email
    msg["Subject"] = subject
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode("utf-8")
    return {"raw": raw}


def find_case_folder(base_dir, cause_no):
    """Find first folder under base_dir whose name contains the cause number."""
    try:
        for name in os.listdir(base_dir):
            full = os.path.join(base_dir, name)
            if os.path.isdir(full) and str(cause_no) in name:
                return full
    except FileNotFoundError:
        pass
    return None


def save_text_copy(folder, filename, subject, body, to_email, cc_email):
    os.makedirs(folder, exist_ok=True)
    out_path = os.path.join(folder, filename)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"TO: {to_email}\n")
        if cc_email:
            f.write(f"CC: {cc_email}\n")
        f.write(f"SUBJECT: {subject}\n\n")
        f.write(body)
    return out_path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default=WORKBOOK_PATH)
    ap.add_argument("--mode", choices=["draft","send"], default=DEFAULT_MODE)
    ap.add_argument("--limit", type=int, default=LIMIT_DEFAULT)
    ap.add_argument("--throttle", type=float, default=THROTTLE_SEC)
    args = ap.parse_args()

    # Load Sheet1 only
    df = load_sheet1(args.workbook, SHEET1_NAME)

    # Eligible: datesubmitted blank/NaN
    if COLS["date_submitted"] not in df.columns:
        print(f"ERROR: '{COLS['date_submitted']}' column not found on Sheet1.")
        sys.exit(1)

    elig = df[
        df[COLS["date_submitted"]].isna() |
        (df[COLS["date_submitted"]].astype(str).str.strip() == "")
    ].copy()

    if elig.empty:
        print("No rows with empty 'datesubmitted' found on Sheet1. Nothing to do.")
        return

    week_phrase, preferred_days, preferred_days_short = choose_week_and_days()

    # Gmail auth
    try:
        service = gmail_service()
    except Exception as e:
        print("Gmail auth failed:", e)
        sys.exit(1)

    run_ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    log_rows = []
    sent = 0

    for _, row in elig.iterrows():
        if sent >= args.limit:
            break

        cause = str(row.get(COLS["cause_no"], "")).strip()
        ward_first = str(row.get(COLS["ward_first"], "")).strip()
        ward_last  = str(row.get(COLS["ward_last"], "")).strip()
        g_name     = str(row.get(COLS["guardian_name"], "")).strip() or "Guardian"
        g_email    = str(row.get(COLS["guardian_email"], "")).strip()
        g2_email   = ""
        if COLS["guardian2_email"] in df.columns:
            g2_email = str(row.get(COLS["guardian2_email"], "")).strip()

        if not g_email or g_email.lower() in ("nan", "none"):
            status = "SKIP:no-primary-email"
            log_rows.append((cause, ward_last, ward_first, g_email, status, ""))
            continue

        subject = SUBJECT_TEMPLATE.format(WardFirst=ward_first, WardLast=ward_last)
        body = BODY_TEMPLATE.format(
            GuardianName=g_name,
            WardFirst=ward_first,
            WardLast=ward_last,
            NextWeekPhrase=week_phrase,
            PreferredDays=preferred_days,
            PreferredDaysShort=preferred_days_short,
        )

        try:
            message = build_message(g_email, subject, body, cc_email=g2_email if g2_email else None)
            if args.mode == "draft":
                service.users().drafts().create(userId="me", body={"message": message}).execute()
                status = "DRAFT"
            else:
                service.users().messages().send(userId="me", body=message).execute()
                status = "SENT"

            # Save text copy
            case_folder = find_case_folder(GUARDIAN_BASE, cause) or os.path.join(GUARDIAN_BASE, "_Unmatched Emails")
            fname = f"Meeting Email - {ward_last}, {ward_first} - {datetime.now().strftime('%Y-%m-%d')}.txt"
            saved_path = save_text_copy(case_folder, fname, subject, body, g_email, g2_email if g2_email else None)

            # Update datesubmitted in Sheet1 (only)
            df.loc[
                df[COLS["cause_no"]].astype(str).str.strip() == cause,
                COLS["date_submitted"]
            ] = date.today().strftime("%Y-%m-%d")

            log_rows.append((cause, ward_last, ward_first, g_email, status, saved_path))
            sent += 1
            time.sleep(args.throttle)

        except HttpError as he:
            log_rows.append((cause, ward_last, ward_first, g_email, f"ERROR:{he}", ""))
        except Exception as e:
            log_rows.append((cause, ward_last, ward_first, g_email, f"ERROR:{e}", ""))

    # Write back updates using openpyxl to preserve formatting
    try:
        from openpyxl import load_workbook

        wb = load_workbook(args.workbook)
        ws = wb[SHEET1_NAME]

        # Find the causeno and datesubmitted column indices
        causeno_col = None
        datesubmitted_col = None

        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value == COLS["cause_no"]:
                causeno_col = col_idx
            elif cell.value == COLS["date_submitted"]:
                datesubmitted_col = col_idx

        if causeno_col and datesubmitted_col:
            # Update datesubmitted for processed cases
            for _, row in df.iterrows():
                cause = str(row.get(COLS["cause_no"], "")).strip()
                date_val = row.get(COLS["date_submitted"])

                # Only update if we changed it (not NaN)
                if pd.notna(date_val):
                    # Find matching row in Excel
                    for row_idx in range(2, ws.max_row + 1):
                        cell_value = str(ws.cell(row_idx, causeno_col).value or "").strip()
                        if cell_value == cause:
                            ws.cell(row_idx, datesubmitted_col, date_val)
                            break

            # Save with preserved formatting
            wb.save(args.workbook)
            wb.close()
        else:
            print("WARNING: Could not find causeno or datesubmitted columns")
            wb.close()

    except Exception as e:
        print("WARNING: Could not write back to workbook:", e)

    # Log CSV next to workbook
    log_dir = os.path.join(os.path.dirname(args.workbook), "runs")
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, f"meeting_email_log_{run_ts}.csv")
    pd.DataFrame(log_rows, columns=["causeno","wardlast","wardfirst","to_email","status","saved_copy_path"]).to_csv(
        log_path, index=False, encoding="utf-8"
    )

    print(f"Done. {sent} item(s) processed. Log: {log_path}")


if __name__ == "__main__":
    main()
