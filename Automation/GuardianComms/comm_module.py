#!/usr/bin/env python3
"""
GuardianAutomation Communications Module (Excel + Gmail-ready)
===============================================================
- Ensures/creates Excel sheets for Clients, Contacts, ClientContacts, Messages, MessageRecipients, MessageLinks.
- Adds required columns if missing; backfills ClientID where needed.
- Template rendering with lightweight placeholder replacement.
- PREVIEW mode: renders per-recipient HTML previews + logs with Status=PREVIEW.
- SEND mode (optional): uses Gmail SMTP if EMAIL_USER and EMAIL_APP_PASSWORD are set.
  (Recommended: enable 2FA and create an App Password in Google account settings.)
- Saves logs to Excel and .eml-like HTML files to Emails/sent/YYYY/YYYY-MM.
- Provider-agnostic logging (Provider='gmail' for SMTP too; thread URL left blank without Gmail API).

Usage examples:
  # First time: initialize workbook and folders
  py comm_module.py init --workbook "C:\path\to\your\Clients.xlsx"

  # Preview a campaign to Active clients with consent
  py comm_module.py preview --workbook "C:\path\to\your\Clients.xlsx" --template simple_reminder --subject "Friendly Reminder"

  # Send a campaign (requires EMAIL_USER + EMAIL_APP_PASSWORD env vars)
  py comm_module.py send --workbook "C:\path\to\your\Clients.xlsx" --template simple_reminder --subject "Friendly Reminder"

  # Show a client's history (by CauseNo or ClientID substring match)
  py comm_module.py history --workbook "C:\path\to\your\Clients.xlsx" --client "C-2024-00123"
"""

import argparse
import os
import sys
import uuid
import json
import smtplib
import re
from pathlib import Path
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception as e:
    print("ERROR: You need 'openpyxl' installed. Try: pip install openpyxl", file=sys.stderr)
    raise

# ----------------------- Configuration -----------------------

REQUIRED_SHEETS = {
    "Clients": [
        "ClientID","CauseNo","WardFirst","WardLast","Status",
        "PrimaryEmail","CCEmail","ConsentToEmail","DoNotEmailReason",
        "LastContacted","NextFollowUp"
    ],
    "Contacts": [
        "ContactID","DisplayName","EmailPrimary","EmailAlt","Phone","Org",
        "RoleType","ConsentToEmail","Notes"
    ],
    "ClientContacts": [
        "ClientID","ContactID","Relationship","IsPrimary"
    ],
    "Messages": [
        "LogID","ClientID","CauseNo","Timestamp","Provider","GmailMessageId","GmailThreadId","ThreadURL",
        "From","To","CC","Subject","Snippet","BodyPreview","HasAttachments","TemplateID","TemplateVersion",
        "Module","Direction","Status","ErrorNote"
    ],
    "MessageRecipients": [
        "LogID","ContactID","Type"  # TO/CC/BCC/FROM
    ],
    "MessageLinks": [
        "LogID","ClientID"
    ],
    "TemplatesIndex": [
        "TemplateID","Version","Purpose","SubjectDefault","FilePath"
    ]
}

DEFAULT_TEMPLATE_ID = "simple_reminder"
DEFAULT_TEMPLATE_HTML = """<!DOCTYPE html>
<html>
  <body>
    <p>Hello {GuardianName},</p>
    <p>This is a friendly reminder about {ReminderTopic} for {WardFirst} {WardLast} (Cause #{CauseNo}).</p>
    <p>If you have questions, reply to this email.</p>
    <p>Thank you,<br>{SenderName}</p>
  </body>
</html>"""

# ----------------------- Helpers -----------------------

def ensure_email_dirs(base_dir: Path):
    emails = base_dir / 'Emails'
    (emails / 'templates').mkdir(parents=True, exist_ok=True)
    (emails / 'previews').mkdir(parents=True, exist_ok=True)
    (emails / 'sent').mkdir(parents=True, exist_ok=True)


def ensure_workbook(path: Path):
    created = False
    if not path.exists():
        wb = openpyxl.Workbook()
        # create first sheet as Clients
        ws = wb.active
        ws.title = "Clients"
        ws.append(REQUIRED_SHEETS["Clients"])
        # add other sheets
        for name, cols in REQUIRED_SHEETS.items():
            if name == "Clients":
                continue
            ws2 = wb.create_sheet(title=name)
            ws2.append(cols)
        wb.save(path)
        created = True
    else:
        wb = openpyxl.load_workbook(path)
        # ensure required sheets and columns
        for name, cols in REQUIRED_SHEETS.items():
            if name not in wb.sheetnames:
                ws = wb.create_sheet(title=name)
                ws.append(cols)
            else:
                ws = wb[name]
                # collect existing headers
                existing = [cell.value for cell in ws[1]]
                changed = False
                # add any missing columns at the end
                for col in cols:
                    if col not in existing:
                        ws.cell(row=1, column=len(existing)+1, value=col)
                        existing.append(col)
                        changed = True
                if changed:
                    # no need to re-order columns now; keep simple
                    pass
        wb.save(path)
    return created

def _col_index_map(ws):
    headers = [c.value for c in ws[1]]
    return {h: i+1 for i, h in enumerate(headers)}

def _get_value(row, idx):
    return row[idx-1].value if idx and idx-1 < len(row) else None

def backfill_client_ids(path: Path):
    wb = openpyxl.load_workbook(path)
    ws = wb["Clients"]
    cmap = _col_index_map(ws)
    if "ClientID" not in cmap:
        raise RuntimeError("Clients sheet missing ClientID header")
    changed = False
    for r in range(2, ws.max_row+1):
        cid = ws.cell(row=r, column=cmap["ClientID"]).value
        cause = ws.cell(row=r, column=cmap.get("CauseNo")).value if cmap.get("CauseNo") else None
        if not cid:
            new_id = str(uuid.uuid4())
            ws.cell(row=r, column=cmap["ClientID"]).value = new_id
            changed = True
        # normalize Status, ConsentToEmail default
        if "Status" in cmap and not ws.cell(row=r, column=cmap["Status"]).value:
            ws.cell(row=r, column=cmap["Status"]).value = "A"
            changed = True
        if "ConsentToEmail" in cmap and ws.cell(row=r, column=cmap["ConsentToEmail"]).value is None:
            ws.cell(row=r, column=cmap["ConsentToEmail"]).value = "Yes"
            changed = True
    if changed:
        wb.save(path)

def ensure_template_index(path: Path, base_dir: Path):
    wb = openpyxl.load_workbook(path)
    ws = wb["TemplatesIndex"]
    cmap = _col_index_map(ws)
    # check if DEFAULT_TEMPLATE_ID is present
    found = False
    for r in range(2, ws.max_row+1):
        tid = ws.cell(row=r, column=cmap["TemplateID"]).value
        if tid == DEFAULT_TEMPLATE_ID:
            found = True
            break
    if not found:
        # write one row
        row = ws.max_row + 1 if ws.max_row > 1 else 2
        file_path = base_dir / "Emails" / "templates" / f"{DEFAULT_TEMPLATE_ID}.html"
        ws.cell(row=row, column=cmap["TemplateID"]).value = DEFAULT_TEMPLATE_ID
        ws.cell(row=row, column=cmap["Version"]).value = "1.0.0"
        ws.cell(row=row, column=cmap["Purpose"]).value = "General gentle reminder"
        ws.cell(row=row, column=cmap["SubjectDefault"]).value = "Friendly Reminder"
        ws.cell(row=row, column=cmap["FilePath"]).value = str(file_path)
        wb.save(path)

def write_default_template(base_dir: Path):
    tpath = base_dir / "Emails" / "templates" / f"{DEFAULT_TEMPLATE_ID}.html"
    tpath.parent.mkdir(parents=True, exist_ok=True)
    if not tpath.exists():
        tpath.write_text(DEFAULT_TEMPLATE_HTML, encoding="utf-8")

def render_template(html_text: str, fields: dict) -> str:
    # Simple {Field} replacement, graceful if missing
    def repl(match):
        key = match.group(1)
        return str(fields.get(key, ""))
    # support {Field} style
    return re.sub(r"\{([A-Za-z0-9_]+)\}", repl, html_text)

def collect_recipients_for_clients(path: Path, active_only=True):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
   
    ws = wb["Clients"]
    cmap = _col_index_map(ws)
    rows = []
    for r in range(2, ws.max_row+1):
        status = ws.cell(row=r, column=cmap.get("Status")).value if cmap.get("Status") else None
        consent = (ws.cell(row=r, column=cmap.get("ConsentToEmail")).value or "").strip().lower()
        to_addr = (ws.cell(row=r, column=cmap.get("PrimaryEmail")).value or "").strip()
        cc_addr = (ws.cell(row=r, column=cmap.get("CCEmail")).value or "").strip()
        if active_only and (status or "").upper() != "A":
            continue
        if consent not in ("yes", "y", "true", "1"):
            continue
        if not to_addr:
            continue
        entry = {
            "ClientID": ws.cell(row=r, column=cmap["ClientID"]).value,
            "CauseNo": ws.cell(row=r, column=cmap.get("CauseNo")).value,
            "WardFirst": ws.cell(row=r, column=cmap.get("WardFirst")).value,
            "WardLast": ws.cell(row=r, column=cmap.get("WardLast")).value,
            "PrimaryEmail": to_addr,
            "CCEmail": cc_addr,
            "GuardianName": ws.cell(row=r, column=cmap.get("GuardianName")).value if cmap.get("GuardianName") else "",
        }
        rows.append(entry)
    return rows

def log_message(path: Path, msg_row: dict, recipients: list, links: list):
    wb = openpyxl.load_workbook(path)
    ws_m = wb["Messages"]
    ws_r = wb["MessageRecipients"]
    ws_l = wb["MessageLinks"]
    cmap_m = _col_index_map(ws_m)
    cmap_r = _col_index_map(ws_r)
    cmap_l = _col_index_map(ws_l)

    # append Messages
    mrow = ws_m.max_row + 1 if ws_m.max_row > 1 else 2
    for k, v in msg_row.items():
        if k in cmap_m:
            ws_m.cell(row=mrow, column=cmap_m[k]).value = v

    # append MessageRecipients
    for rec in recipients:
        rrow = ws_r.max_row + 1 if ws_r.max_row > 1 else 2
        ws_r.cell(row=rrow, column=cmap_r["LogID"]).value = msg_row["LogID"]
        ws_r.cell(row=rrow, column=cmap_r["ContactID"]).value = rec.get("ContactID", "")  # optional today
        ws_r.cell(row=rrow, column=cmap_r["Type"]).value = rec.get("Type", "TO")

    # append MessageLinks
    for link in links:
        lrow = ws_l.max_row + 1 if ws_l.max_row > 1 else 2
        ws_l.cell(row=lrow, column=cmap_l["LogID"]).value = msg_row["LogID"]
        ws_l.cell(row=lrow, column=cmap_l["ClientID"]).value = link.get("ClientID", "")

    wb.save(path)

def save_preview_file(base_dir: Path, client_entry: dict, subject: str, body_html: str):
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    slug = re.sub(r"[^A-Za-z0-9_-]+","_", (subject or "email"))
    fname = f"{ts}_{client_entry.get('CauseNo','NA')}_{slug}.html"
    outdir = base_dir / "Emails" / "previews"
    outdir.mkdir(parents=True, exist_ok=True)
    fpath = outdir / fname
    fpath.write_text(body_html, encoding="utf-8")
    return str(fpath)

def save_sent_file(base_dir: Path, client_entry: dict, subject: str, body_html: str):
    now = datetime.now()
    outdir = base_dir / "Emails" / "sent" / f"{now:%Y}" / f"{now:%Y-%m}"
    outdir.mkdir(parents=True, exist_ok=True)
    ts = now.strftime("%Y%m%d-%H%M%S")
    slug = re.sub(r"[^A-Za-z0-9_-]+","_", (subject or "email"))
    fname = f"{ts}_{client_entry.get('CauseNo','NA')}_{slug}.html"
    fpath = outdir / fname
    fpath.write_text(body_html, encoding="utf-8")
    return str(fpath)

def send_via_gmail_smtp(sender, app_password, to_addr, cc_addr, subject, html_body):
    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From'] = sender
    msg['To'] = to_addr
    if cc_addr:
        msg['Cc'] = cc_addr
    part = MIMEText(html_body, 'html')
    msg.attach(part)
    recipients = [to_addr] + ([cc_addr] if cc_addr else [])
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
        server.login(sender, app_password)
        server.sendmail(sender, recipients, msg.as_string())
    # We don't get Gmail threadId here; Message-Id is inside headers but not returned by smtplib
    return {}

def do_init(args):
    wb_path = Path(args.workbook).expanduser()
    base_dir = Path(args.base).expanduser() if args.base else wb_path.parent
    created = ensure_workbook(wb_path)
    backfill_client_ids(wb_path)
    ensure_email_dirs(base_dir)
    write_default_template(base_dir)
    ensure_template_index(wb_path, base_dir)
    print(f"[OK] Workbook ensured at: {wb_path}")
    print(f"[OK] Default template written: {base_dir / 'Emails' / 'templates' / (DEFAULT_TEMPLATE_ID + '.html')}")
    print("[TIP] Add your real templates and update TemplatesIndex when ready.")

def _load_template_by_id(wb_path: Path, template_id: str):
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb["TemplatesIndex"]
    cmap = _col_index_map(ws)
    for r in range(2, ws.max_row+1):
        tid = ws.cell(row=r, column=cmap["TemplateID"]).value
        if tid == template_id:
            fpath = ws.cell(row=r, column=cmap["FilePath"]).value
            subject_default = ws.cell(row=r, column=cmap["SubjectDefault"]).value
            version = ws.cell(row=r, column=cmap["Version"]).value
            html = Path(fpath).read_text(encoding="utf-8")
            return html, subject_default, version, fpath
    raise RuntimeError(f"TemplateID not found in TemplatesIndex: {template_id}")

def _build_fields(entry, sender_name):
    fields = {
        "GuardianName": entry.get("GuardianName") or "there",
        "WardFirst": entry.get("WardFirst") or "",
        "WardLast": entry.get("WardLast") or "",
        "CauseNo": entry.get("CauseNo") or "",
        "SenderName": sender_name or "Court Visitor",
        "ReminderTopic": "your upcoming item",
    }
    return fields

def do_preview_or_send(args, mode="preview"):
    wb_path = Path(args.workbook).expanduser()
    base_dir = Path(args.base).expanduser() if args.base else wb_path.parent
    template_html, subject_default, template_version, template_path = _load_template_by_id(wb_path, args.template)
    subject = args.subject or subject_default or "Notification"
    sender_display = args.sender_name or "Court Visitor"

    entries = collect_recipients_for_clients(wb_path, active_only=(not args.include_completed))
    if not entries:
        print("No eligible recipients found. Check Status, ConsentToEmail, and PrimaryEmail columns.")
        return

    # optional: dedupe by email
    seen = set()
    deduped = []
    for e in entries:
        key = (e["PrimaryEmail"].lower(), e["CCEmail"].lower() if e["CCEmail"] else "")
        if key in seen:
            continue
        seen.add(key)
        deduped.append(e)
    entries = deduped

    print(f"{'Previewing' if mode=='preview' else 'Sending'} {len(entries)} message(s)…")

    sender_email = os.environ.get("EMAIL_USER")
    app_password = os.environ.get("EMAIL_APP_PASSWORD")
    smtp_enabled = (mode == "send") and sender_email and app_password

    for e in entries:
        fields = _build_fields(e, sender_display)
        body_html = render_template(template_html, fields)
        if mode == "preview" or not smtp_enabled:
            out = save_preview_file(base_dir, e, subject, body_html)
            status = "PREVIEW"
            provider = "gmail"
            msg_ids = {}
        else:
            # SEND
            msg_ids = send_via_gmail_smtp(sender_email, app_password, e["PrimaryEmail"], e["CCEmail"], subject, body_html)
            out = save_sent_file(base_dir, e, subject, body_html)
            status = "SENT"
            provider = "gmail"

        # log
        log_id = str(uuid.uuid4())
        msg_row = {
            "LogID": log_id,
            "ClientID": e["ClientID"],
            "CauseNo": e.get("CauseNo",""),
            "Timestamp": datetime.now().isoformat(timespec="seconds"),
            "Provider": provider,
            "GmailMessageId": msg_ids.get("message_id",""),
            "GmailThreadId": msg_ids.get("thread_id",""),
            "ThreadURL": "",  # requires Gmail API for accurate thread links
            "From": sender_email or "",
            "To": e["PrimaryEmail"],
            "CC": e["CCEmail"],
            "Subject": subject,
            "Snippet": re.sub(r"<[^>]+>", " ", body_html)[:200],
            "BodyPreview": re.sub(r"<[^>]+>", " ", body_html).splitlines()[0][:200],
            "HasAttachments": "N",
            "TemplateID": args.template,
            "TemplateVersion": template_version,
            "Module": args.module or "Email",
            "Direction": "OUTBOUND",
            "Status": status,
            "ErrorNote": ""
        }
        recipients = [
            {"ContactID": "", "Type": "TO"},
        ]
        if e.get("CCEmail"):
            recipients.append({"ContactID":"", "Type":"CC"})
        links = [{"ClientID": e["ClientID"]}]

        log_message(wb_path, msg_row, recipients, links)

    if mode == "preview":
        print("Preview complete. Files saved under Emails/previews. Logs written to Messages/MessageRecipients/MessageLinks.")
    else:
        if not smtp_enabled:
            print("SEND requested but EMAIL_USER or EMAIL_APP_PASSWORD not set. Performed PREVIEW instead.")
        else:
            print("Send complete. Files saved under Emails/sent/YYYY/YYYY-MM. Logs updated.")

def do_history(args):
    wb_path = Path(args.workbook).expanduser()
    key = args.client.strip().lower()
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    ws_m = wb["Messages"]
    cmap = _col_index_map(ws_m)

    matches = []
    for r in range(2, ws_m.max_row+1):
        cid = (ws_m.cell(row=r, column=cmap.get("ClientID")).value or "").lower()
        cause = (ws_m.cell(row=r, column=cmap.get("CauseNo")).value or "").lower()
        if key in cid or key in cause:
            row = {h: ws_m.cell(row=r, column=idx).value for h, idx in cmap.items()}
            matches.append(row)

    matches.sort(key=lambda x: x.get("Timestamp") or "", reverse=True)
    for m in matches[:50]:
        print(f"[{m.get('Timestamp')}] {m.get('Subject')}  To:{m.get('To')}  CC:{m.get('CC')}  Status:{m.get('Status')}")
        snippet = (m.get("Snippet") or "")[:160]
        print(f"   {snippet}…")
    if not matches:
        print("No messages found for that ClientID/CauseNo.")

def main():
    ap = argparse.ArgumentParser(description="GuardianAutomation Communications Module")
    ap.add_argument("--base", help="Base folder for Emails/ (defaults to workbook folder)")

    sub = ap.add_subparsers(dest="cmd", required=True)

    ap_init = sub.add_parser("init", help="Create/ensure sheets and default template")
    ap_init.add_argument("--workbook", required=True, help="Path to Excel workbook")
    ap_init.set_defaults(func=do_init)

    ap_prev = sub.add_parser("preview", help="Render previews and log as PREVIEW")
    ap_prev.add_argument("--workbook", required=True)
    ap_prev.add_argument("--template", default="simple_reminder")
    ap_prev.add_argument("--subject", default=None)
    ap_prev.add_argument("--sender-name", default=None)
    ap_prev.add_argument("--module", default="Email")
    ap_prev.add_argument("--include-completed", action="store_true", help="Include completed clients too")
    ap_prev.set_defaults(func=lambda args: do_preview_or_send(args, mode="preview"))

    ap_send = sub.add_parser("send", help="Send via Gmail SMTP (requires env EMAIL_USER, EMAIL_APP_PASSWORD)")
    ap_send.add_argument("--workbook", required=True)
    ap_send.add_argument("--template", default="simple_reminder")
    ap_send.add_argument("--subject", default=None)
    ap_send.add_argument("--sender-name", default=None)
    ap_send.add_argument("--module", default="Email")
    ap_send.add_argument("--include-completed", action="store_true", help="Include completed clients too")
    ap_send.set_defaults(func=lambda args: do_preview_or_send(args, mode="send"))

    ap_hist = sub.add_parser("history", help="Print recent history for a client (by ClientID or CauseNo substring)")
    ap_hist.add_argument("--workbook", required=True)
    ap_hist.add_argument("--client", required=True, help="ClientID or CauseNo substring")
    ap_hist.set_defaults(func=do_history)

    args = ap.parse_args()
    args.func(args)

if __name__ == "__main__":
    main()
