#!/usr/bin/env python3
"""
Email CVR to Supervisor
Finds completed CVR documents and emails them to supervisor
"""

import os
import json
import pandas as pd
import win32com.client
from datetime import datetime
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
import pickle

class EmailCVRToSupervisor:
    def __init__(self):
        self.credentials = None
        self.gmail_service = None
        self.supervisor_email = self.load_supervisor_email()
    
    def load_supervisor_email(self):
        """Load supervisor email from config"""
        try:
            config_path = "Config/email_config.json"
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    config = json.load(f)
                return config.get('supervisor_email', 'al.benedict@traviscountytx.gov')
            else:
                # Create default config
                config = {
                    'supervisor_email': 'al.benedict@traviscountytx.gov',
                    'last_used_email': 'al.benedict@traviscountytx.gov'
                }
                os.makedirs("Config", exist_ok=True)
                with open(config_path, 'w') as f:
                    json.dump(config, f, indent=2)
                return config['supervisor_email']
        except Exception as e:
            print(f"Error loading supervisor email config: {e}")
            return 'al.benedict@traviscountytx.gov'

    def save_supervisor_email(self, email):
        """Save supervisor email to config"""
        try:
            config_path = "Config/email_config.json"
            config = {
                'supervisor_email': email,
                'last_used_email': email
            }
            os.makedirs("Config", exist_ok=True)
            with open(config_path, 'w') as f:
                json.dump(config, f, indent=2)
            self.supervisor_email = email
            print(f"Supervisor email updated to: {email}")
        except Exception as e:
            print(f"Error saving supervisor email: {e}")
    
    def authenticate(self):
        """Authenticate with Gmail API"""
        try:
            from pathlib import Path
            SCOPES = ['https://www.googleapis.com/auth/gmail.send']

            # Check for existing credentials - use same paths as other Gmail scripts
            creds_path = r"C:\configlocal\API\gmail_oauth_client.json"
            token_path = Path(r"C:\configlocal\API\gmail_token.json")

            # Try to load existing token (JSON format)
            if token_path.exists():
                self.credentials = Credentials.from_authorized_user_file(str(token_path), SCOPES)

            # If no valid credentials, authenticate
            if not self.credentials or not self.credentials.valid:
                if self.credentials and self.credentials.expired and self.credentials.refresh_token:
                    try:
                        self.credentials.refresh(Request())
                    except Exception as e:
                        # Token refresh failed (expired/revoked) - delete and re-authenticate
                        print(f"Token refresh failed: {e}")
                        print("Deleting expired token and starting fresh OAuth flow...")
                        if token_path.exists():
                            token_path.unlink()
                        self.credentials = None  # Force re-auth below

                if not self.credentials:
                    if os.path.exists(creds_path):
                        flow = InstalledAppFlow.from_client_secrets_file(creds_path, SCOPES)
                        self.credentials = flow.run_local_server(port=0)
                    else:
                        raise Exception("Gmail OAuth credentials not found. Please set up Gmail API first.")

                # Save the token (JSON format)
                with open(token_path, 'w') as token:
                    token.write(self.credentials.to_json())

            self.gmail_service = build('gmail', 'v1', credentials=self.credentials)
            return True

        except Exception as e:
            print(f"Authentication error: {str(e)}")
            return False
    
    def find_completed_cvrs(self):
        """Find CVR documents that need to be emailed (no date submitted)"""
        try:
            # Read Excel file
            excel_file = "App Data/ward_guardian_info.xlsx"
            df = pd.read_excel(excel_file)
            
            # Find rows where CVR is created but not yet submitted
            cvr_created = df['CVR created?'] == 'Y'
            no_date_submitted = df['datesubmitted'].isna() | (df['datesubmitted'] == '')
            
            # Get cases that need emailing
            cases_to_email = df[cvr_created & no_date_submitted]
            
            print(f"Found {len(cases_to_email)} CVR documents to email")
            
            # Find corresponding CVR files
            cvr_files = []
            for _, case in cases_to_email.iterrows():
                cause_no = str(case.get('causeno', '')).strip()
                ward_last = str(case.get('wardlast', '')).strip()
                ward_first = str(case.get('wardfirst', '')).strip()
                ward_middle = str(case.get('wardmiddle', '')).strip()

                if cause_no and ward_last and ward_first:
                    # Try to find folder by cause number (more robust than exact name matching)
                    case_folder = None
                    new_clients_dir = "New Clients"

                    # Search for folder containing the cause number
                    if os.path.exists(new_clients_dir):
                        for folder_name in os.listdir(new_clients_dir):
                            if cause_no in folder_name and os.path.isdir(os.path.join(new_clients_dir, folder_name)):
                                case_folder = os.path.join(new_clients_dir, folder_name)
                                break

                    if case_folder:
                        print(f"\nFound folder: {os.path.basename(case_folder)}")
                    else:
                        print(f"\n[WARN] No folder found for cause {cause_no}")

                    if case_folder and os.path.exists(case_folder):
                        print(f"  Folder exists! Searching for CVR...")
                        cvr_found = False
                        for file in os.listdir(case_folder):
                            if file.endswith('.docx') and 'court visitor report' in file.lower():
                                cvr_path = os.path.join(case_folder, file)
                                cvr_files.append({
                                    'path': cvr_path,
                                    'case': case,
                                    'filename': file
                                })
                                print(f"  [OK] Found CVR: {file}")
                                cvr_found = True
                                break
                        if not cvr_found:
                            print(f"  [WARN] Folder exists but no CVR file found")
                            print(f"  Files in folder: {[f for f in os.listdir(case_folder) if f.endswith('.docx')]}")
                    else:
                        print(f"  [WARN] Folder does not exist")
                else:
                    print(f"\n[WARN] Missing case data: cause={cause_no}, last={ward_last}, first={ward_first}")
            
            return cvr_files
            
        except Exception as e:
            print(f"Error finding CVR documents: {str(e)}")
            return []
    
    def create_email_message(self, cvr_files):
        """Create email message with attachments"""
        try:
            if not cvr_files:
                return None
            
            # Create subject line with CVR names
            cvr_names = [f["filename"].replace('.docx', '') for f in cvr_files]
            subject = f"Court Visitor Reports: {', '.join(cvr_names)}"
            
            # Email body
            body = """Hello,

Attached please find the following completed Court Visitor Reports for your review.

Please let me know if you have any questions.

Respectfully,
May Ehresman"""
            
            # Create message
            message = {
                'raw': self.create_raw_message(subject, body, cvr_files)
            }
            
            return message
            
        except Exception as e:
            print(f"Error creating email message: {str(e)}")
            return None
    
    def create_raw_message(self, subject, body, cvr_files):
        """Create raw email message with attachments"""
        import base64
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders
        
        msg = MIMEMultipart()
        msg['To'] = self.supervisor_email
        msg['Subject'] = subject
        
        # Add body
        msg.attach(MIMEText(body, 'plain'))
        
        # Add attachments
        for cvr_file in cvr_files:
            try:
                with open(cvr_file['path'], 'rb') as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                    encoders.encode_base64(part)
                    part.add_header(
                        'Content-Disposition',
                        f'attachment; filename= {cvr_file["filename"]}'
                    )
                    msg.attach(part)
            except Exception as e:
                print(f"Error attaching {cvr_file['filename']}: {e}")
        
        # Encode message
        raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode('utf-8')
        return raw_message
    
    def send_email(self, message):
        """Send email via Gmail API"""
        try:
            if not message:
                return False
            
            # Send email
            sent_message = self.gmail_service.users().messages().send(
                userId='me',
                body=message
            ).execute()
            
            print(f"Email sent successfully. Message ID: {sent_message['id']}")
            return True
            
        except Exception as e:
            print(f"Error sending email: {str(e)}")
            return False
    
    def move_folders_to_completed(self, cvr_files):
        """Move case folders from New Clients to Completed after emailing"""
        import shutil

        try:
            completed_folder = "Completed"
            os.makedirs(completed_folder, exist_ok=True)

            moved_count = 0
            for cvr_file in cvr_files:
                # Get the folder path from the CVR path
                cvr_path = cvr_file['path']
                case_folder = os.path.dirname(cvr_path)
                folder_name = os.path.basename(case_folder)

                # Destination path
                dest_folder = os.path.join(completed_folder, folder_name)

                # Move folder
                if os.path.exists(case_folder):
                    try:
                        # Remove Correspondence subfolder if it exists (can cause lock issues)
                        corr_folder = os.path.join(case_folder, 'Correspondence')
                        if os.path.exists(corr_folder):
                            shutil.rmtree(corr_folder, ignore_errors=True)

                        # If destination exists, remove it first
                        if os.path.exists(dest_folder):
                            shutil.rmtree(dest_folder, ignore_errors=True)

                        shutil.move(case_folder, dest_folder)
                        print(f"  Moved to Completed: {folder_name}")
                        moved_count += 1
                    except Exception as e:
                        print(f"  [WARN] Could not move {folder_name}: {e}")

            print(f"\n[OK] Moved {moved_count} folder(s) to Completed")
            return True

        except Exception as e:
            print(f"Error moving folders to Completed: {str(e)}")
            return False

    def update_excel_submission_dates(self, cvr_files):
        """Update Excel with submission dates for emailed CVRs"""
        try:
            excel_file = "App Data/ward_guardian_info.xlsx"
            df = pd.read_excel(excel_file)
            
            # Update datesubmitted for each CVR that was emailed
            current_date = datetime.now().strftime('%Y-%m-%d')

            # Use openpyxl to preserve Excel formatting
            from openpyxl import load_workbook

            wb = load_workbook(excel_file)
            ws = wb.active

            # Find the causeno and datesubmitted column indices
            causeno_col = None
            datesubmitted_col = None

            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value == 'causeno':
                    causeno_col = col_idx
                elif cell.value == 'datesubmitted':
                    datesubmitted_col = col_idx

            if not causeno_col or not datesubmitted_col:
                print("ERROR: Could not find causeno or datesubmitted columns")
                wb.close()
                return False

            # Update datesubmitted for matching cases
            for cvr_file in cvr_files:
                case = cvr_file['case']
                cause_no = str(case.get('causeno', '')).strip()

                # Find matching row
                for row_idx in range(2, ws.max_row + 1):
                    cell_value = str(ws.cell(row_idx, causeno_col).value or '').strip()
                    if cell_value == cause_no:
                        ws.cell(row_idx, datesubmitted_col, current_date)
                        print(f"Updated submission date for case {cause_no}")
                        break

            # Save with preserved formatting
            wb.save(excel_file)
            wb.close()
            print("Excel file updated with submission dates (formatting preserved)")
            return True
            
        except Exception as e:
            print(f"Error updating Excel: {str(e)}")
            return False
    
    def show_email_preview(self, cvr_files):
        """Show email preview dialog for user confirmation"""
        try:
            import tkinter as tk
            from tkinter import messagebox, ttk

            # Create preview window using Toplevel (no separate root needed when called from GUI)
            preview_window = tk.Toplevel()
            preview_window.title("Email Preview - CVR Documents")
            preview_window.geometry("600x500")
            preview_window.resizable(True, True)
            preview_window.grab_set()  # Make it modal

            # Center the window
            preview_window.transient()
            preview_window.lift()

            # Main frame
            main_frame = ttk.Frame(preview_window, padding="20")
            main_frame.pack(fill='both', expand=True)
            
            # Title
            title_label = ttk.Label(main_frame, text="📧 Email Preview - CVR Documents", 
                                  font=('Segoe UI', 14, 'bold'))
            title_label.pack(pady=(0, 20))
            
            # Email details frame
            details_frame = ttk.LabelFrame(main_frame, text="Email Details", padding="10")
            details_frame.pack(fill='x', pady=(0, 20))

            # Recipient - now editable
            ttk.Label(details_frame, text="To:", font=('Segoe UI', 10, 'bold')).grid(row=0, column=0, sticky='w', padx=(0, 10))
            email_var = tk.StringVar(value=self.supervisor_email)
            email_entry = ttk.Entry(details_frame, textvariable=email_var, width=40, font=('Segoe UI', 10))
            email_entry.grid(row=0, column=1, sticky='ew', padx=(0, 10))
            details_frame.columnconfigure(1, weight=1)
            
            # Subject
            subject = f"Court Visitor Reports: {', '.join([f['filename'].replace('.docx', '') for f in cvr_files])}"
            ttk.Label(details_frame, text="Subject:", font=('Segoe UI', 10, 'bold')).grid(row=1, column=0, sticky='w', padx=(0, 10), pady=(10, 0))
            ttk.Label(details_frame, text=subject, font=('Segoe UI', 10)).grid(row=1, column=1, sticky='w', pady=(10, 0))
            
            # Attachments
            ttk.Label(details_frame, text="Attachments:", font=('Segoe UI', 10, 'bold')).grid(row=2, column=0, sticky='nw', padx=(0, 10), pady=(10, 0))
            
            # Attachments list
            attachments_text = tk.Text(details_frame, height=4, width=50, wrap='word')
            attachments_text.grid(row=2, column=1, sticky='w', pady=(10, 0))
            for cvr_file in cvr_files:
                attachments_text.insert(tk.END, f"• {cvr_file['filename']}\n")
            attachments_text.config(state='disabled')
            
            # Email body
            body_frame = ttk.LabelFrame(main_frame, text="Email Body", padding="10")
            body_frame.pack(fill='both', expand=True, pady=(0, 20))
            
            body_text = tk.Text(body_frame, height=6, wrap='word')
            body_text.pack(fill='both', expand=True)
            
            email_body = """Hello,

Attached please find the following completed Court Visitor Reports for your review.

Please let me know if you have any questions.

Respectfully,
May Ehresman"""
            
            body_text.insert(tk.END, email_body)
            body_text.config(state='disabled')
            
            # Buttons frame
            buttons_frame = ttk.Frame(main_frame)
            buttons_frame.pack(fill='x')
            
            # Store the result
            result = {'confirmed': False}

            def confirm_send():
                # Get the email from the entry widget
                entered_email = email_var.get().strip()

                # Validate email is not empty
                if not entered_email:
                    from tkinter import messagebox
                    messagebox.showerror("Invalid Email", "Please enter a recipient email address.")
                    return

                # Update supervisor email if it changed
                if entered_email != self.supervisor_email:
                    self.save_supervisor_email(entered_email)

                result['confirmed'] = True
                preview_window.destroy()

            def cancel_send():
                result['confirmed'] = False
                preview_window.destroy()
            
            # Buttons
            ttk.Button(buttons_frame, text="✉️ Send Email", command=confirm_send).pack(side='right', padx=(10, 0))
            ttk.Button(buttons_frame, text="❌ Cancel", command=cancel_send).pack(side='right')
            
            # Wait for user response
            preview_window.wait_window()

            return result['confirmed']

        except Exception as e:
            print(f"Error showing email preview: {str(e)}")
            import traceback
            traceback.print_exc()
            # Fallback to simple confirmation
            return messagebox.askyesno("Send CVR Email?",
                                     f"Send {len(cvr_files)} CVR documents to {self.supervisor_email}?")
    
    def process_cvr_email(self):
        """Main method to process CVR email"""
        try:
            print("Starting CVR email process...")
            
            # Authenticate
            if not self.authenticate():
                print("Authentication failed")
                return False
            
            # Find CVR documents to email
            cvr_files = self.find_completed_cvrs()
            if not cvr_files:
                print("No CVR documents found to email")
                return False
            
            # Show email preview and get confirmation
            if not self.show_email_preview(cvr_files):
                print("Email sending cancelled by user")
                return False
            
            # Create email message
            message = self.create_email_message(cvr_files)
            if not message:
                print("Failed to create email message")
                return False
            
            # Send email
            if self.send_email(message):
                # Update Excel with submission dates
                self.update_excel_submission_dates(cvr_files)

                # Move folders to Completed
                print("\nMoving case folders to Completed...")
                self.move_folders_to_completed(cvr_files)

                print("\n[OK] CVR email process completed successfully")
                return True
            else:
                print("Failed to send email")
                return False
                
        except Exception as e:
            print(f"Error in CVR email process: {str(e)}")
            return False

if __name__ == "__main__":
    import sys
    email_sender = EmailCVRToSupervisor()
    result = email_sender.process_cvr_email()
    print(f"Result: {result}")

    # Exit with proper error code so GUI can detect failure
    if not result:
        print("\n[FAIL] Email CVR to Supervisor FAILED - see errors above")
        sys.exit(1)
    else:
        print("\n[OK] Email CVR to Supervisor SUCCESS")
        sys.exit(0)
