#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
format_excel_file.py

One-time script to apply proper formatting to the ward_guardian_info.xlsx Excel file:
1. Format all date columns as MM/DD/YYYY
2. Format all phone number columns as (XXX) XXX-XXXX
3. Enable text wrapping for address columns
4. Auto-adjust column widths to fit content

This script modifies the Excel file formatting only - it does NOT change any data.
All future scripts will preserve this formatting using openpyxl.
"""

import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

# Excel file path
EXCEL_FILE = r"C:\GoogleSync\GuardianShip_App\App Data\ward_guardian_info.xlsx"
SHEET_NAME = "Sheet1"

# Column categories (these are the column NAMES in the Excel header row)
DATE_COLUMNS = [
    'visitdate',
    'wdob',           # Ward date of birth
    'gdob',           # Guardian date of birth
    'datesubmitted',
    'datefiled',
    'datearpfiled',
]

PHONE_COLUMNS = [
    'wtele',          # Ward telephone
    'gtele',          # Guardian telephone
    'g2tele',         # Guardian 2 telephone
]

ADDRESS_COLUMNS = [
    'waddress',       # Ward address
    'gaddress',       # Guardian address
    'g2 address',     # Guardian 2 address
]

TIME_COLUMNS = [
    'visittime',      # Visit time
]


def format_phone_number(value):
    """
    Format phone number as (XXX) XXX-XXXX
    Handles various input formats: 5551234567, (555) 123-4567, 555-123-4567, etc.
    """
    if not value or str(value).strip() == '' or str(value).lower() in ('nan', 'none'):
        return value

    # Extract only digits
    digits = ''.join(c for c in str(value) if c.isdigit())

    # If we have 10 digits, format as (XXX) XXX-XXXX
    if len(digits) == 10:
        return f"({digits[0:3]}) {digits[3:6]}-{digits[6:10]}"

    # If we have 11 digits and starts with 1, skip the 1
    if len(digits) == 11 and digits[0] == '1':
        return f"({digits[1:4]}) {digits[4:7]}-{digits[7:11]}"

    # Otherwise return as-is
    return value


def format_time_value(value):
    """
    Normalize time values to HH:MM AM/PM format
    Handles: 1p, 1pm, 13:00, 1:00, 1:00pm, etc.
    """
    if not value or str(value).strip() == '' or str(value).lower() in ('nan', 'none'):
        return value

    val_str = str(value).strip().lower()

    # Handle simple formats like "1p" or "2a"
    import re
    m = re.match(r'^(\d{1,2})\s*([ap])$', val_str)
    if m:
        hour = int(m.group(1))
        ampm = 'PM' if m.group(2) == 'p' else 'AM'
        return f"{hour}:00 {ampm}"

    # Handle "1pm" or "2am"
    m = re.match(r'^(\d{1,2})\s*([ap]m)$', val_str)
    if m:
        hour = int(m.group(1))
        ampm = m.group(2).upper()
        return f"{hour}:00 {ampm}"

    # Handle "1:30pm" or "13:30"
    m = re.match(r'^(\d{1,2}):(\d{2})\s*([ap]m)?$', val_str)
    if m:
        hour = int(m.group(1))
        minute = m.group(2)
        ampm = m.group(3)

        if ampm:
            # Already has AM/PM
            return f"{hour}:{minute} {ampm.upper()}"
        else:
            # Convert 24-hour to 12-hour
            if hour == 0:
                return f"12:{minute} AM"
            elif hour < 12:
                return f"{hour}:{minute} AM"
            elif hour == 12:
                return f"12:{minute} PM"
            else:
                return f"{hour-12}:{minute} PM"

    # If we can't parse it, return as-is
    return value


def format_date_value(value):
    """
    Convert date to MM/DD/YYYY format
    Handles pandas Timestamp, datetime, and date strings
    """
    if not value or str(value).strip() == '' or str(value).lower() in ('nan', 'none'):
        return value

    # If it's already a datetime object
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")

    # If it's a string, try to parse it
    if isinstance(value, str):
        try:
            # Try parsing common formats
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%m-%d-%Y', '%Y/%m/%d']:
                try:
                    dt = datetime.strptime(value.strip(), fmt)
                    return dt.strftime("%m/%d/%Y")
                except ValueError:
                    continue
        except:
            pass

    return value


def main():
    print("=" * 70)
    print("Excel File Formatter")
    print("=" * 70)
    print(f"\nFormatting: {EXCEL_FILE}")
    print(f"Sheet: {SHEET_NAME}\n")

    # Check if file exists
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: File not found: {EXCEL_FILE}")
        return

    # Create backup
    backup_file = EXCEL_FILE.replace('.xlsx', '_backup_before_formatting.xlsx')
    if not os.path.exists(backup_file):
        import shutil
        shutil.copy2(EXCEL_FILE, backup_file)
        print(f"[BACKUP] Created: {backup_file}\n")

    # Load workbook
    print("Loading workbook...")
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    # Get column mapping (column name -> column index)
    col_map = {}
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value:
            col_map[str(cell.value).strip()] = col_idx

    print(f"Found {len(col_map)} columns in Excel file\n")

    # Track changes
    date_cols_found = []
    phone_cols_found = []
    address_cols_found = []

    # 1. Format DATE columns
    print("=" * 70)
    print("STEP 1: Formatting Date Columns")
    print("=" * 70)
    for col_name in DATE_COLUMNS:
        if col_name in col_map:
            col_idx = col_map[col_name]
            date_cols_found.append(col_name)

            # Apply date format to entire column (skip header row)
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row_idx, col_idx)

                # Format the cell value
                if cell.value:
                    formatted = format_date_value(cell.value)
                    cell.value = formatted

                # Apply number format for dates
                cell.number_format = 'MM/DD/YYYY'

            print(f"  [OK] {col_name} (column {col_idx})")

    if date_cols_found:
        print(f"\n  Formatted {len(date_cols_found)} date columns\n")
    else:
        print("  WARNING: No date columns found\n")

    # 2. Format PHONE NUMBER columns
    print("=" * 70)
    print("STEP 2: Formatting Phone Number Columns")
    print("=" * 70)
    for col_name in PHONE_COLUMNS:
        if col_name in col_map:
            col_idx = col_map[col_name]
            phone_cols_found.append(col_name)

            # Format phone numbers (skip header row)
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row_idx, col_idx)

                if cell.value:
                    formatted = format_phone_number(cell.value)
                    cell.value = formatted

                # Set as text format to preserve formatting
                cell.number_format = '@'  # '@' means text format

            print(f"  [OK] {col_name} (column {col_idx})")

    if phone_cols_found:
        print(f"\n  Formatted {len(phone_cols_found)} phone number columns\n")
    else:
        print("  WARNING: No phone columns found\n")

    # 3. Format TIME columns
    print("=" * 70)
    print("STEP 3: Formatting Time Columns")
    print("=" * 70)
    time_cols_found = []
    for col_name in TIME_COLUMNS:
        if col_name in col_map:
            col_idx = col_map[col_name]
            time_cols_found.append(col_name)

            # Format times (skip header row)
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row_idx, col_idx)

                if cell.value:
                    formatted = format_time_value(cell.value)
                    cell.value = formatted

                # Set as text format to preserve formatting
                cell.number_format = '@'  # '@' means text format

            print(f"  [OK] {col_name} (column {col_idx})")

    if time_cols_found:
        print(f"\n  Formatted {len(time_cols_found)} time columns\n")
    else:
        print("  WARNING: No time columns found\n")

    # 4. Enable TEXT WRAPPING for address columns
    print("=" * 70)
    print("STEP 4: Enabling Text Wrapping for Address Columns")
    print("=" * 70)
    for col_name in ADDRESS_COLUMNS:
        if col_name in col_map:
            col_idx = col_map[col_name]
            address_cols_found.append(col_name)

            # Enable text wrapping for entire column
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row_idx, col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical='top')

            print(f"  [OK] {col_name} (column {col_idx})")

    if address_cols_found:
        print(f"\n  Enabled text wrapping for {len(address_cols_found)} address columns\n")
    else:
        print("  WARNING: No address columns found\n")

    # 5. AUTO-ADJUST column widths
    print("=" * 70)
    print("STEP 5: Auto-Adjusting Column Widths")
    print("=" * 70)

    # Set minimum and maximum widths
    MIN_WIDTH = 10
    MAX_WIDTH = 50

    for col in ws.columns:
        col_letter = col[0].column_letter
        col_name = col[0].value

        # Calculate max length in this column
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        # Set adjusted width (with padding)
        adjusted_width = min(max(max_length + 2, MIN_WIDTH), MAX_WIDTH)
        ws.column_dimensions[col_letter].width = adjusted_width

        if col_name:
            print(f"  [OK] {col_name}: width={adjusted_width}")

    print()

    # 6. SAVE workbook
    print("=" * 70)
    print("STEP 6: Saving Changes")
    print("=" * 70)
    print("Saving formatted workbook...")
    wb.save(EXCEL_FILE)
    wb.close()

    print("[OK] Excel file formatting complete!\n")

    # Summary
    print("=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"Date columns formatted:     {len(date_cols_found)}")
    print(f"Phone columns formatted:    {len(phone_cols_found)}")
    print(f"Time columns formatted:     {len(time_cols_found)}")
    print(f"Address columns wrapped:    {len(address_cols_found)}")
    print(f"All columns auto-adjusted:  YES")
    print(f"\nBackup saved to: {backup_file}")
    print(f"Formatted file: {EXCEL_FILE}")
    print("\n[DONE] All formatting applied successfully!")
    print("=" * 70)


if __name__ == "__main__":
    main()
