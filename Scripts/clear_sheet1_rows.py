"""
Clear Sheet1 rows (keep headers) in ward_guardian_info.xlsx
CRITICAL: Only clear Sheet1, preserve all other sheets
"""
import openpyxl
from pathlib import Path

excel_path = Path(r"C:\GoogleSync\GuardianShip_App\App Data\ward_guardian_info.xlsx")

if not excel_path.exists():
    print(f"Excel file not found: {excel_path}")
    exit(1)

print(f"Opening: {excel_path}")
wb = openpyxl.load_workbook(excel_path)

if "Sheet1" not in wb.sheetnames:
    print("Sheet1 not found!")
    exit(1)

sheet = wb["Sheet1"]

# Get the number of rows
max_row = sheet.max_row
print(f"Sheet1 has {max_row} rows (including header)")

if max_row > 1:
    # Delete rows 2 onwards (keep row 1 which is the header)
    sheet.delete_rows(2, max_row - 1)
    print(f"Deleted {max_row - 1} data rows, kept header row")
else:
    print("Sheet1 already empty (only has header)")

# Save
wb.save(excel_path)
print(f"Saved: {excel_path}")
print("\nSheet1 cleared! All other sheets preserved.")
