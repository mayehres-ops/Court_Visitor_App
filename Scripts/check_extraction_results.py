"""
Check what was extracted to the Excel file
"""
import openpyxl
from pathlib import Path

excel_path = Path(r"C:\GoogleSync\GuardianShip_App\App Data\ward_guardian_info.xlsx")

if not excel_path.exists():
    print(f"Excel file not found: {excel_path}")
    exit(1)

wb = openpyxl.load_workbook(excel_path)
sheet = wb["Sheet1"]

# Get headers
headers = [cell.value for cell in sheet[1]]
print(f"Headers: {headers}")
print()

# Get all data rows
print("=" * 120)
print("EXTRACTION RESULTS:")
print("=" * 120)

for row_idx in range(2, sheet.max_row + 1):
    row_data = {}
    for col_idx, header in enumerate(headers, start=1):
        cell_value = sheet.cell(row_idx, col_idx).value
        if cell_value:
            row_data[header] = cell_value

    if row_data:
        print(f"\nRow {row_idx - 1}:")
        print("-" * 80)

        # Print key fields
        key_fields = ['causeno', 'wardfirst', 'wardlast', 'guardian1', 'Guardian2',
                      'gaddress', 'g2 address', 'DateARPfiled', 'Dateappointed']

        for field in key_fields:
            if field in row_data:
                value = row_data[field]
                print(f"  {field:15s}: {value}")

        # Print other fields
        print()
        for key, value in row_data.items():
            if key not in key_fields:
                print(f"  {key:15s}: {value}")

print("\n" + "=" * 120)
print("SUMMARY:")
print("=" * 120)

# Count Guardian2 extractions
g2_count = 0
datearpfiled_count = 0

for row_idx in range(2, sheet.max_row + 1):
    # Check Guardian2 column
    g2_col_idx = headers.index('Guardian2') + 1 if 'Guardian2' in headers else None
    if g2_col_idx:
        g2_value = sheet.cell(row_idx, g2_col_idx).value
        if g2_value and str(g2_value).strip():
            g2_count += 1

    # Check DateARPfiled column
    datearp_col_idx = headers.index('DateARPfiled') + 1 if 'DateARPfiled' in headers else None
    if datearp_col_idx:
        datearp_value = sheet.cell(row_idx, datearp_col_idx).value
        if datearp_value and str(datearp_value).strip():
            datearpfiled_count += 1

total_rows = sheet.max_row - 1  # Exclude header

print(f"Total cases extracted: {total_rows}")
print(f"Cases with Guardian2: {g2_count}/{total_rows}")
print(f"Cases with DateARPfiled: {datearpfiled_count}/{total_rows}")
print()
