"""
GuardianShip Easy App - PDF to Excel Extractor
Windows-only, single-file script for parsing ORDER and ARP PDFs
"""

import os
import re
import sys
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Tuple, List

# Core dependencies
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    import pdfplumber
    from pdf2image import convert_from_path
    import pytesseract
    from PIL import Image
except ImportError as e:
    print(f"ERROR: Missing required package: {e}")
    print("Please install: pip install openpyxl pdfplumber pdf2image pytesseract Pillow")
    sys.exit(1)

# Optional Vision OCR
try:
    from google.cloud import vision
    VISION_AVAILABLE = True
except ImportError:
    VISION_AVAILABLE = False

# ========== CONFIGURATION ==========
BASE_PATH = Path(r"C:\GoogleSync\GuardianShip_App")
INPUT_FOLDER = BASE_PATH / "New Files"
APP_DATA_FOLDER = BASE_PATH / "App Data"
WORKBOOK_NAME = "ward_guardian_info.xlsx"
WORKBOOK_PATH = APP_DATA_FOLDER / WORKBOOK_NAME
LOG_FOLDER = BASE_PATH / "Logs"
BACKUP_FOLDER = BASE_PATH / "Backups"
DEBUG_TEXT_FOLDER = LOG_FOLDER / "debug_texts"
LOG_FILE = LOG_FOLDER / "extraction_log.txt"

TESSERACT_PATH = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
POPPLER_PATH = r"C:\Program Files\poppler-24.08.0\Library\bin"  # Adjust version as needed
VISION_CREDS_PATH = Path(r"C:\configlocal\API\google_service_account.json")

# Excel header order (EXACT)
EXCEL_HEADERS = [
    "wardlast", "wardfirst", "causeno", "visitdate", "visittime", "wtele",
    "liveswith", "waddress", "wdob", "guardian1", "gaddress", "gemail",
    "gtele", "Relationship", "gdob", "Guardian2", "g2 address", "g2eamil",
    "g2tele", "g2Relationship", "g2dob", "datesubmited", "Dateappointed",
    "miles", "expensesubmited", "expenspd", "DateARPfiled", "last_updated"
]

# Fields that ALWAYS get overwritten on upsert
ALWAYS_OVERWRITE = {"causeno", "Dateappointed", "last_updated"}

# ========== LOGGING ==========
def log(message: str):
    """Append timestamped message to log file and print to console."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_line = f"[{timestamp}] {message}"
    print(log_line)
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(log_line + "\n")

def save_debug_text(causeno: str, label: str, text: str):
    """Save debug text snapshot for troubleshooting."""
    DEBUG_TEXT_FOLDER.mkdir(parents=True, exist_ok=True)
    safe_causeno = re.sub(r'[^\w\-]', '_', causeno)
    filename = DEBUG_TEXT_FOLDER / f"{safe_causeno}__{label}.txt"
    with open(filename, "w", encoding="utf-8") as f:
        f.write(text)

# ========== PREFLIGHT CHECKS ==========
def run_preflight():
    """Run environment checks before processing."""
    log("=" * 60)
    log("PREFLIGHT CHECKS")
    log("=" * 60)
    
    # Python & pip
    log(f"Python: {sys.executable}")
    log(f"Version: {sys.version}")
    
    # Tesseract
    tesseract_paths = []
    if os.path.exists(TESSERACT_PATH):
        tesseract_paths.append(TESSERACT_PATH)
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
        log(f"Tesseract: {TESSERACT_PATH} ✓")
    else:
        log(f"WARNING: Tesseract not found at {TESSERACT_PATH}")
    
    # Check for duplicates
    if sys.platform == "win32":
        import subprocess
        try:
            result = subprocess.run(["where", "tesseract"], capture_output=True, text=True)
            if result.returncode == 0:
                found = result.stdout.strip().split("\n")
                if len(found) > 1:
                    log(f"WARNING: Multiple Tesseract installs found: {found}")
        except:
            pass
    
    # Poppler
    if os.path.exists(POPPLER_PATH):
        log(f"Poppler: {POPPLER_PATH} ✓")
    else:
        log(f"WARNING: Poppler not found at {POPPLER_PATH}")
    
    # Package versions
    try:
        log(f"openpyxl: {openpyxl.__version__}")
    except:
        log("openpyxl: version unknown")
    
    try:
        log(f"pdfplumber: {pdfplumber.__version__}")
    except:
        log("pdfplumber: version unknown")
    
    # Vision OCR
    if VISION_CREDS_PATH.exists() and VISION_AVAILABLE:
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = str(VISION_CREDS_PATH)
        log(f"Vision: enabled ({VISION_CREDS_PATH})")
    else:
        log("Vision: disabled (no creds or package not installed)")
    
    log("=" * 60)

# ========== NORMALIZATION HELPERS ==========
def normalize_causeno(raw: str) -> str:
    """Normalize cause number to NN-NNNNNN format."""
    if not raw:
        return ""
    
    # Extract digits
    digits = re.findall(r'\d+', raw)
    if len(digits) >= 2:
        # Take last two groups: year (2 digits) and number (6+ digits)
        year = digits[-2][-2:]  # Last 2 digits of year
        number = digits[-1].zfill(6)  # Pad number to 6 digits
        return f"{year}-{number}"
    elif len(digits) == 1 and len(digits[0]) >= 8:
        # Single string like "23123456"
        full = digits[0]
        return f"{full[-8:-6]}-{full[-6:]}"
    
    return raw.strip()

def normalize_date(raw: str) -> str:
    """Convert various date formats to MM/DD/YYYY."""
    if not raw:
        return ""
    
    raw = raw.strip()
    
    # Replace dots and dashes with slashes
    raw = raw.replace(".", "/").replace("-", "/")
    
    # Try numeric formats first (MM/DD/YYYY, M/D/YY, etc.)
    numeric_patterns = [
        (r'(\d{1,2})/(\d{1,2})/(\d{4})', lambda m: f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"),
        (r'(\d{1,2})/(\d{1,2})/(\d{2})', lambda m: f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{clamp_year(m.group(3))}"),
    ]
    
    for pattern, formatter in numeric_patterns:
        match = re.search(pattern, raw)
        if match:
            return formatter(match)
    
    # Try Month DD, YYYY format
    month_names = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
        'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12,
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'jun': 6, 'jul': 7, 'aug': 8,
        'sep': 9, 'sept': 9, 'oct': 10, 'nov': 11, 'dec': 12
    }
    
    month_pattern = r'(january|february|march|april|may|june|july|august|september|october|november|december|jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)\s+(\d{1,2}),?\s+(\d{4}|\d{2})'
    match = re.search(month_pattern, raw, re.IGNORECASE)
    if match:
        month_str, day, year = match.groups()
        month = month_names[month_str.lower()]
        year = clamp_year(year) if len(year) == 2 else year
        return f"{month:02d}/{int(day):02d}/{year}"
    
    return raw

def clamp_year(yy: str) -> str:
    """Convert 2-digit year to 4-digit (20YY or 19YY)."""
    year = int(yy)
    if year >= 0 and year <= 30:
        return f"20{year:02d}"
    else:
        return f"19{year:02d}"

def normalize_phone(raw: str) -> str:
    """Normalize phone to (XXX) XXX-XXXX format."""
    if not raw:
        return ""
    
    digits = re.sub(r'\D', '', raw)
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return raw

def normalize_address(raw: str) -> str:
    """Clean up address: strip labels, collapse whitespace."""
    if not raw:
        return ""
    
    # Remove common labels
    cleaned = re.sub(r'Address:?|\\(no P\.O\. Box\\)|City/State/Zip:?', '', raw, flags=re.IGNORECASE)
    
    # Collapse multiple spaces/newlines
    cleaned = re.sub(r'\s+', ' ', cleaned)
    
    return cleaned.strip()

def normalize_name(raw: str) -> str:
    """Clean name: strip boilerplate, title-case."""
    if not raw:
        return ""
    
    # Remove common boilerplate
    cleaned = re.sub(r'\b(the ward|an incapacitated person|incapacitated person|estate of)\b', '', raw, flags=re.IGNORECASE)
    
    # Handle "Last, First" format
    if ',' in cleaned:
        parts = [p.strip() for p in cleaned.split(',')]
        if len(parts) == 2:
            cleaned = f"{parts[1]} {parts[0]}"
    
    # Title case
    cleaned = ' '.join(word.capitalize() for word in cleaned.split())
    
    return cleaned.strip()

# ========== TEXT EXTRACTION ==========
def extract_text_pdfplumber(pdf_path: Path) -> str:
    """Extract text using pdfplumber (fast method)."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                text += page.extract_text() or ""
            return text
    except Exception as e:
        log(f"  pdfplumber failed: {e}")
        return ""

def extract_text_tesseract(pdf_path: Path, page_num: int = 0) -> str:
    """Extract text using Tesseract OCR with PSM mode selection."""
    try:
        images = convert_from_path(pdf_path, first_page=page_num+1, last_page=page_num+1, poppler_path=POPPLER_PATH)
        if not images:
            return ""
        
        img = images[0]
        
        # Try both PSM 4 and 6, choose better result
        text_psm4 = pytesseract.image_to_string(img, config='--psm 4')
        text_psm6 = pytesseract.image_to_string(img, config='--psm 6')
        
        # Simple heuristic: count "guardian" mentions (case-insensitive)
        score4 = text_psm4.lower().count('guardian')
        score6 = text_psm6.lower().count('guardian')
        
        return text_psm4 if score4 >= score6 else text_psm6
    except Exception as e:
        log(f"  Tesseract OCR failed: {e}")
        return ""

def extract_text_vision(pdf_path: Path, page_num: int = 0) -> str:
    """Extract text using Google Vision OCR (optional handwriting fallback)."""
    if not VISION_AVAILABLE or not VISION_CREDS_PATH.exists():
        return ""
    
    try:
        client = vision.ImageAnnotatorClient()
        images = convert_from_path(pdf_path, first_page=page_num+1, last_page=page_num+1, poppler_path=POPPLER_PATH)
        if not images:
            return ""
        
        # Convert PIL image to bytes
        from io import BytesIO
        img_byte_arr = BytesIO()
        images[0].save(img_byte_arr, format='PNG')
        img_byte_arr = img_byte_arr.getvalue()
        
        image = vision.Image(content=img_byte_arr)
        response = client.document_text_detection(image=image)
        
        if response.error.message:
            raise Exception(response.error.message)
        
        return response.full_text_annotation.text if response.full_text_annotation else ""
    except Exception as e:
        log(f"  Vision OCR failed: {e}")
        return ""

def get_pdf_text(pdf_path: Path, is_arp: bool = False) -> str:
    """Get PDF text with OCR fallback logic."""
    # Try pdfplumber first
    text = extract_text_pdfplumber(pdf_path)
    
    # If too short, use OCR
    if len(text) < 80:
        log(f"  Text too short ({len(text)} chars), using OCR...")
        text = extract_text_tesseract(pdf_path, page_num=0)
        
        # For ARP, try Vision if available and if Tesseract result is poor
        if is_arp and text.lower().count('guardian') < 2:
            vision_text = extract_text_vision(pdf_path, page_num=0)
            if vision_text and vision_text.lower().count('guardian') > text.lower().count('guardian'):
                log(f"  Vision OCR produced better result")
                text = vision_text
    
    return text

# ========== PARSING FUNCTIONS ==========
def parse_order_pdf(text: str, filename: str) -> Dict[str, str]:
    """Parse ORDER PDF for cause number, date appointed, and ward name."""
    data = {}
    
    # Extract cause number
    causeno_match = re.search(r'Cause\s+No\.?\s*[:\s]*([\w\s\-]+)', text, re.IGNORECASE)
    if causeno_match:
        data['causeno'] = normalize_causeno(causeno_match.group(1))
    
    # Extract date appointed (look for "Signed on" or similar)
    date_patterns = [
        r'Signed\s+on[:\s]*([A-Za-z]+\s+\d{1,2},?\s+\d{2,4})',
        r'Signed[:\s]*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
        r'Entered[:\s]*([A-Za-z]+\s+\d{1,2},?\s+\d{2,4})',
        r'Entered[:\s]*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
    ]
    
    for pattern in date_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            data['Dateappointed'] = normalize_date(match.group(1))
            break
    
    # Extract ward name from "IN THE MATTER OF / GUARDIANSHIP OF"
    ward_patterns = [
        r'IN\s+THE\s+MATTER\s+OF\s+(?:THE\s+)?GUARDIANSHIP\s+OF[:\s]*([\w\s,]+?)(?:\s*,?\s*(?:an?\s+)?(?:incapacitated|ward|minor|person|adult))',
        r'GUARDIANSHIP\s+OF[:\s]*([\w\s,]+?)(?:\s*,?\s*(?:an?\s+)?(?:incapacitated|ward|minor|person|adult))',
    ]
    
    for pattern in ward_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            full_name = normalize_name(match.group(1))
            # Split into first/last
            parts = full_name.split()
            if len(parts) >= 2:
                data['wardfirst'] = parts[0]
                data['wardlast'] = ' '.join(parts[1:])
            elif len(parts) == 1:
                data['wardlast'] = parts[0]
            break
    
    return data

def parse_arp_pdf(text: str, filename: str) -> Dict[str, str]:
    """Parse ARP PDF for ward, guardian, and filing information."""
    data = {}
    
    save_debug_text(filename, "full_text", text)
    
    # Extract cause number
    causeno_match = re.search(r'Cause\s+No\.?[:\s]*([\w\s\-]+)', text, re.IGNORECASE)
    if causeno_match:
        data['causeno'] = normalize_causeno(causeno_match.group(1))
    
    # Extract ward information
    ward_section = extract_section(text, r'WARD\s+INFORMATION', r'GUARDIAN\s+INFORMATION', 500)
    if ward_section:
        save_debug_text(filename, "ward_section", ward_section)
        
        # Ward name
        name_match = re.search(r'Name[:\s]*([\w\s,]+?)(?:\n|\r|$)', ward_section, re.IGNORECASE)
        if name_match:
            full_name = normalize_name(name_match.group(1))
            parts = full_name.split()
            if len(parts) >= 2:
                data['wardfirst'] = parts[0]
                data['wardlast'] = ' '.join(parts[1:])
        
        # Ward DOB
        dob_match = re.search(r'Date\s+of\s+Birth[:\s]*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})', ward_section, re.IGNORECASE)
        if dob_match:
            data['wdob'] = normalize_date(dob_match.group(1))
        
        # Ward phone
        phone_match = re.search(r'Telephone[:\s]*([\d\-\(\)\s]+)', ward_section, re.IGNORECASE)
        if phone_match:
            data['wtele'] = normalize_phone(phone_match.group(1))
        
        # Ward address (multi-line)
        data['waddress'] = extract_address(ward_section)
    
    # Extract guardian information
    guardian_section = extract_section(text, r'GUARDIAN\s+INFORMATION', r'(?:PHYSICIAN|LIVING|RESIDENCE)', 800)
    if guardian_section:
        save_debug_text(filename, "guardian_section", guardian_section)
        
        # Guardian 1 name
        name_match = re.search(r'Name[:\s]*([\w\s,\.]+?)(?:\n|\r|$)', guardian_section, re.IGNORECASE)
        if name_match:
            data['guardian1'] = normalize_name(name_match.group(1))
        
        # Guardian DOB
        dob_match = re.search(r'Date\s+of\s+Birth[:\s]*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})', guardian_section, re.IGNORECASE)
        if dob_match:
            data['gdob'] = normalize_date(dob_match.group(1))
        
        # Guardian phone
        phone_match = re.search(r'Telephone[:\s]*([\d\-\(\)\s]+)', guardian_section, re.IGNORECASE)
        if phone_match:
            data['gtele'] = normalize_phone(phone_match.group(1))
        
        # Guardian email
        email_match = re.search(r'Email[:\s]*([\w\.\-]+@[\w\.\-]+)', guardian_section, re.IGNORECASE)
        if email_match:
            data['gemail'] = email_match.group(1).lower()
        
        # Relationship
        rel_match = re.search(r'Relationship[:\s]*([\w\s]+?)(?:\n|\r|$)', guardian_section, re.IGNORECASE)
        if rel_match:
            data['Relationship'] = rel_match.group(1).strip()
        
        # Guardian address
        data['gaddress'] = extract_address(guardian_section)
    
    # Lives with (checkbox logic)
    liveswith_match = re.search(r'Do\s+you\s+reside\s+with\s+the\s+ward\?\s*\[.\]\s*YES', text, re.IGNORECASE)
    if liveswith_match:
        data['liveswith'] = "Guardian"
    
    # Date ARP filed
    filed_patterns = [
        r'Filed[/\s]*Entered[:\s]*([A-Za-z]+\s+\d{1,2},?\s+\d{2,4})',
        r'Filed[:\s]*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
        r'Entered[:\s]*([A-Za-z]+\s+\d{1,2},?\s+\d{2,4})',
    ]
    
    for pattern in filed_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            data['DateARPfiled'] = normalize_date(match.group(1))
            break
    
    return data

def extract_section(text: str, start_pattern: str, end_pattern: str, max_chars: int = 500) -> str:
    """Extract a section of text between two patterns."""
    start_match = re.search(start_pattern, text, re.IGNORECASE)
    if not start_match:
        return ""
    
    start_pos = start_match.end()
    end_match = re.search(end_pattern, text[start_pos:], re.IGNORECASE)
    
    if end_match:
        end_pos = start_pos + end_match.start()
    else:
        end_pos = start_pos + max_chars
    
    return text[start_pos:end_pos]

def extract_address(section: str) -> str:
    """Extract address from section with multi-line support."""
    # Look for "Address (no P.O. Box)" label
    addr_patterns = [
        r'Address\s*\(no\s+P\.O\.\s+Box\)[:\s]*(.*?)(?:City/State/Zip)',
        r'Address[:\s]*(.*?)(?:City/State/Zip|Telephone|Email|$)',
    ]
    
    for pattern in addr_patterns:
        match = re.search(pattern, section, re.IGNORECASE | re.DOTALL)
        if match:
            addr_part = match.group(1).strip()
            
            # Look for city/state/zip
            csz_match = re.search(r'City/State/Zip[:\s]*(.*?)(?:\n|$)', section, re.IGNORECASE)
            if csz_match:
                csz_part = csz_match.group(1).strip()
                full_addr = f"{addr_part} {csz_part}"
            else:
                full_addr = addr_part
            
            return normalize_address(full_addr)
    
    # Fallback: find first street-ish line
    lines = section.split('\n')
    for line in lines:
        if re.search(r'\d+\s+\w+', line):  # Contains number and word (likely street)
            return normalize_address(line)
    
    return ""

# ========== EXCEL OPERATIONS ==========
def ensure_workbook() -> openpyxl.Workbook:
    """Load or create workbook with proper headers."""
    APP_DATA_FOLDER.mkdir(parents=True, exist_ok=True)
    
    if WORKBOOK_PATH.exists():
        wb = load_workbook(WORKBOOK_PATH)
        ws = wb.active
        
        # Verify headers
        current_headers = [cell.value for cell in ws[1]]
        if current_headers != EXCEL_HEADERS:
            log("  WARNING: Headers don't match expected format, fixing...")
            # Insert correct headers
            ws.delete_rows(1)
            ws.insert_rows(1)
            for idx, header in enumerate(EXCEL_HEADERS, start=1):
                ws.cell(row=1, column=idx, value=header)
    else:
        log("  Creating new workbook...")
        wb = Workbook()
        ws = wb.active
        
        # Write headers
        for idx, header in enumerate(EXCEL_HEADERS, start=1):
            ws.cell(row=1, column=idx, value=header)
        
        wb.save(WORKBOOK_PATH)
    
    return wb

def find_row_by_causeno(ws, causeno: str) -> Optional[int]:
    """Find row index for given cause number."""
    causeno_col = EXCEL_HEADERS.index("causeno") + 1
    
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=causeno_col).value
        if cell_value and str(cell_value).strip() == causeno:
            return row_idx
    
    return None

def upsert_row(wb: openpyxl.Workbook, data: Dict[str, str]):
    """Insert or update row with idempotent rules."""
    ws = wb.active
    causeno = data.get('causeno', '')
    
    if not causeno:
        log("  WARNING: No cause number, skipping upsert")
        return
    
    # Add timestamp
    data['last_updated'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Find existing row
    row_idx = find_row_by_causeno(ws, causeno)
    
    if row_idx:
        log(f"  Updating existing row {row_idx} for cause {causeno}")
        
        # Update with idempotent rules
        for header in EXCEL_HEADERS:
            col_idx = EXCEL_HEADERS.index(header) + 1
            current_value = ws.cell(row=row_idx, column=col_idx).value
            new_value = data.get(header, '')
            
            # Always overwrite these fields
            if header in ALWAYS_OVERWRITE:
                if new_value:
                    ws.cell(row=row_idx, column=col_idx, value=new_value)
            # Only fill if blank
            elif not current_value or str(current_value).strip() == '':
                if new_value:
                    ws.cell(row=row_idx, column=col_idx, value=new_value)
    else:
        log(f"  Appending new row for cause {causeno}")
        
        # Append new row
        new_row_idx = ws.max_row + 1
        for header in EXCEL_HEADERS:
            col_idx = EXCEL_HEADERS.index(header) + 1
            value = data.get(header, '')
            ws.cell(row=new_row_idx, column=col_idx, value=value)
    
    # Save workbook
    wb.save(WORKBOOK_PATH)
    
    # Log preview of non-empty fields
    preview = {k: v for k, v in data.items() if v}
    log(f"  Preview: {preview}")

def backup_workbook():
    """Create one-time backup of workbook for this run."""
    if not WORKBOOK_PATH.exists():
        return
    
    BACKUP_FOLDER.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUP_FOLDER / f"{WORKBOOK_NAME.replace('.xlsx', '')}_{timestamp}.xlsx"
    
    shutil.copy2(WORKBOOK_PATH, backup_path)
    log(f"Backup created: {backup_path}")

# ========== MAIN PROCESSING ==========
def process_pdf(pdf_path: Path, wb: openpyxl.Workbook):
    """Process a single PDF file."""
    filename = pdf_path.name.lower()
    
    # Skip approval PDFs
    if any(word in filename for word in ['approval', 'approved', 'approvals']):
        log(f"  Skipping (contains 'approval'): {pdf_path.name}")
        return
    
    log(f"Processing: {pdf_path.name}")
    
    # Determine PDF type
    is_order = 'order' in filename
    is_arp = 'arp' in filename
    
    if not is_order and not is_arp:
        log(f"  Skipping (unknown type): {pdf_path.name}")
        return
    
    # Extract text
    text = get_pdf_text(pdf_path, is_arp=is_arp)
    
    if not text or len(text) < 50:
        log(f"  WARNING: Very little text extracted ({len(text)} chars)")
        return
    
    # Parse based on type
    if is_order:
        data = parse_order_pdf(text, pdf_path.name)
    else:  # is_arp
        data = parse_arp_pdf(text, pdf_path.name)
    
    if not data:
        log(f"  WARNING: No data extracted from {pdf_path.name}")
        return
    
    # Upsert to Excel
    upsert_row(wb, data)
    log(f"  ✓ Completed: {pdf_path.name}")

def main():
    """Main entry point."""
    log("=" * 60)
    log("GuardianShip Easy App - PDF Extractor Starting")
    log("=" * 60)
    
    # Preflight checks
    run_preflight()
    
    # Create necessary folders
    INPUT_FOLDER.mkdir(parents=True, exist_ok=True)
    LOG_FOLDER.mkdir(parents=True, exist_ok=True)
    BACKUP_FOLDER.mkdir(parents=True, exist_ok=True)
    
    # Backup existing workbook
    backup_workbook()
    
    # Ensure workbook exists with proper headers
    wb = ensure_workbook()
    
    # Find all PDFs
    pdf_files = list(INPUT_FOLDER.glob("*.pdf"))
    log(f"Found {len(pdf_files)} PDF file(s) in {INPUT_FOLDER}")
    
    if not pdf_files:
        log("No PDFs to process. Exiting.")
        return
    
    # Process each PDF
    for pdf_path in pdf_files:
        try:
            process_pdf(pdf_path, wb)
        except Exception as e:
            log(f"  ERROR processing {pdf_path.name}: {e}")
            import traceback
            log(f"  Traceback: {traceback.format_exc()}")
    
    log("=" * 60)
    log("Processing complete!")
    log("=" * 60)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"FATAL ERROR: {e}")
        import traceback
        log(f"Traceback: {traceback.format_exc()}")
        sys.exit(1)